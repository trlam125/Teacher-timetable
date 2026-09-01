from __future__ import annotations

import csv
import io
import re
import unicodedata
import zipfile
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable
from xml.etree import ElementTree as ET

from openpyxl import load_workbook


MAX_ARCHIVE_UNCOMPRESSED_BYTES = 80 * 1024 * 1024
SUPPORTED_EXTENSIONS = {".xlsx", ".xlsm", ".xls", ".csv", ".tsv", ".docx"}


class ScheduleAuditParseError(ValueError):
    pass


@dataclass
class RawLesson:
    day_text: str
    session_text: str
    period_text: str
    class_text: str
    lesson_text: str = ""
    subject_text: str = ""
    teacher_text: str = ""
    room_text: str = ""
    source: str = ""
    origin: str = "aggregate"


def normalize_text(value: Any) -> str:
    text = str(value or "").strip().casefold().replace("đ", "d")
    text = unicodedata.normalize("NFD", text)
    text = "".join(ch for ch in text if unicodedata.category(ch) != "Mn")
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _archive_size_guard(content: bytes) -> None:
    try:
        with zipfile.ZipFile(io.BytesIO(content)) as zf:
            total = sum(max(0, item.file_size) for item in zf.infolist())
    except zipfile.BadZipFile as exc:
        raise ScheduleAuditParseError("File nén không hợp lệ hoặc đã bị hỏng.") from exc
    if total > MAX_ARCHIVE_UNCOMPRESSED_BYTES:
        raise ScheduleAuditParseError("File có dữ liệu giải nén quá lớn để kiểm tra an toàn.")


def _read_excel_tables(content: bytes, suffix: str) -> list[tuple[str, list[list[str]]]]:
    if suffix in {".xlsx", ".xlsm"}:
        _archive_size_guard(content)
        try:
            workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        except Exception as exc:
            raise ScheduleAuditParseError("Không đọc được file Excel. Hãy kiểm tra file có bị hỏng không.") from exc
        tables: list[tuple[str, list[list[str]]]] = []
        try:
            for sheet in workbook.worksheets:
                rows = [[_cell_text(value) for value in row] for row in sheet.iter_rows(values_only=True)]
                tables.append((sheet.title, rows))
        finally:
            workbook.close()
        return tables

    try:
        import xlrd  # type: ignore
    except ImportError as exc:
        raise ScheduleAuditParseError(
            "File .xls cần thư viện xlrd. Hãy chạy lại pip install -r requirements.txt hoặc lưu file thành .xlsx."
        ) from exc
    try:
        book = xlrd.open_workbook(file_contents=content, on_demand=True)
    except Exception as exc:
        raise ScheduleAuditParseError("Không đọc được file Excel .xls.") from exc
    tables = []
    try:
        for sheet in book.sheets():
            rows = [[_cell_text(sheet.cell_value(r, c)) for c in range(sheet.ncols)] for r in range(sheet.nrows)]
            tables.append((sheet.name, rows))
    finally:
        book.release_resources()
    return tables


def _decode_delimited(content: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "cp1258", "cp1252"):
        try:
            return content.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise ScheduleAuditParseError("Không nhận diện được bảng mã của file CSV/TSV.")


def _read_delimited_table(content: bytes, suffix: str) -> list[tuple[str, list[list[str]]]]:
    text = _decode_delimited(content)
    sample = text[:8192]
    delimiter = "\t" if suffix == ".tsv" else ","
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        delimiter = dialect.delimiter
    except csv.Error:
        pass
    rows = [[_cell_text(value) for value in row] for row in csv.reader(io.StringIO(text), delimiter=delimiter)]
    return [("Dữ liệu", rows)]


def _docx_text(element: ET.Element, namespace: dict[str, str]) -> str:
    paragraph_tag = f"{{{namespace['w']}}}p"
    paragraphs = [element] if element.tag == paragraph_tag else element.findall(".//w:p", namespace)
    if paragraphs:
        chunks = [
            "".join(node.text or "" for node in paragraph.findall(".//w:t", namespace)).strip()
            for paragraph in paragraphs
        ]
        return re.sub(r"\s+", " ", " ".join(chunk for chunk in chunks if chunk)).strip()
    texts = [node.text or "" for node in element.findall(".//w:t", namespace)]
    return re.sub(r"\s+", " ", "".join(texts)).strip()


def _read_docx_tables(content: bytes) -> list[tuple[str, list[list[str]]]]:
    _archive_size_guard(content)
    namespace = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
    try:
        with zipfile.ZipFile(io.BytesIO(content)) as zf:
            xml = zf.read("word/document.xml")
    except (KeyError, zipfile.BadZipFile) as exc:
        raise ScheduleAuditParseError("Không đọc được tài liệu Word .docx.") from exc
    try:
        root = ET.fromstring(xml)
    except ET.ParseError as exc:
        raise ScheduleAuditParseError("Nội dung XML của tài liệu Word không hợp lệ.") from exc

    body = root.find("w:body", namespace)
    if body is None:
        return []
    tables: list[tuple[str, list[list[str]]]] = []
    last_paragraph = ""
    table_no = 0
    paragraph_tag = f"{{{namespace['w']}}}p"
    table_tag = f"{{{namespace['w']}}}tbl"
    for child in body:
        if child.tag == paragraph_tag:
            text = _docx_text(child, namespace)
            if text:
                last_paragraph = text
            continue
        if child.tag != table_tag:
            continue
        table_no += 1
        rows: list[list[str]] = []
        for tr in child.findall("w:tr", namespace):
            row = [_docx_text(tc, namespace) for tc in tr.findall("w:tc", namespace)]
            rows.append(row)
        title = last_paragraph or f"Bảng {table_no}"
        tables.append((title, rows))
    return tables


def read_tables(filename: str, content: bytes) -> tuple[str, list[tuple[str, list[list[str]]]]]:
    suffix = Path(filename or "").suffix.lower()
    if suffix not in SUPPORTED_EXTENSIONS:
        if suffix == ".doc":
            raise ScheduleAuditParseError("Word .doc đời cũ chưa được hỗ trợ. Hãy lưu lại dưới dạng .docx.")
        raise ScheduleAuditParseError(
            "Định dạng chưa hỗ trợ. Dùng .xlsx, .xlsm, .xls, .csv, .tsv hoặc .docx."
        )
    if suffix in {".xlsx", ".xlsm", ".xls"}:
        return suffix.lstrip(".").upper(), _read_excel_tables(content, suffix)
    if suffix in {".csv", ".tsv"}:
        return suffix.lstrip(".").upper(), _read_delimited_table(content, suffix)
    return "DOCX", _read_docx_tables(content)


def _alias_map(items: Iterable[dict[str, Any]], fields: tuple[str, ...]) -> dict[str, int]:
    aliases: dict[str, int] = {}
    ambiguous: set[str] = set()
    for item in items:
        item_id = int(item["id"])
        for field in fields:
            alias = normalize_text(item.get(field, ""))
            if not alias:
                continue
            if alias in aliases and aliases[alias] != item_id:
                ambiguous.add(alias)
            else:
                aliases[alias] = item_id
    for alias in ambiguous:
        aliases.pop(alias, None)
    return aliases


def _exact_alias(value: str, aliases: dict[str, int]) -> int | None:
    return aliases.get(normalize_text(value))


def _find_alias(value: str, aliases: dict[str, int]) -> int | None:
    text = normalize_text(value)
    if not text:
        return None
    exact = aliases.get(text)
    if exact is not None:
        return exact
    padded = f" {text} "
    matches: list[tuple[int, str, int]] = []
    for alias, item_id in aliases.items():
        if len(alias) < 2:
            continue
        if f" {alias} " in padded:
            matches.append((len(alias), alias, item_id))
    if not matches:
        return None
    matches.sort(reverse=True)
    best_length = matches[0][0]
    best_ids = {item_id for length, _alias, item_id in matches if length == best_length}
    return next(iter(best_ids)) if len(best_ids) == 1 else None


def _header_kind(value: str) -> str:
    text = normalize_text(value)
    if text in {"thu", "ngay", "day", "weekday"} or text.startswith("thu "):
        return "day"
    if text in {"buoi", "ca", "session", "shift"}:
        return "session"
    if text in {"tiet", "tiet hoc", "period", "lesson", "lesson no"}:
        return "period"
    if text in {"lop", "lop hoc", "class", "classroom"}:
        return "class"
    if text in {"mon", "mon hoc", "subject"}:
        return "subject"
    if text in {"giao vien", "gv", "teacher"}:
        return "teacher"
    if text in {"phong", "phong hoc", "room"}:
        return "room"
    if text in {"noi dung", "lich hoc", "bai hoc", "lesson info", "noi dung tiet"}:
        return "lesson"
    return ""


def _day_index(value: str) -> int | None:
    text = normalize_text(value)
    if not text:
        return None
    if text in {"cn", "chu nhat", "sunday", "sun"}:
        return 6
    english = {
        "monday": 0, "mon": 0, "tuesday": 1, "tue": 1, "wednesday": 2, "wed": 2,
        "thursday": 3, "thu": 3, "friday": 4, "fri": 4, "saturday": 5, "sat": 5,
    }
    if text in english:
        return english[text]
    match = re.search(r"(?:thu\s*)?([2-8])\b", text)
    if match:
        number = int(match.group(1))
        return 6 if number == 8 else number - 2
    if text.isdigit():
        number = int(text)
        if 2 <= number <= 8:
            return 6 if number == 8 else number - 2
    return None


def _session_index(value: str, session_count: int) -> int | None:
    if session_count <= 1:
        return 0
    text = normalize_text(value)
    if not text:
        return None
    # File TKB thuc te thuong ghi "Sang 2" / "Chieu 2", trong do so 2 la thu
    # chu khong phai so buoi. Uu tien tu Sang/Chieu truoc khi doc bat ky chu so nao.
    if text == "s" or text.startswith("sang") or "morning" in text:
        return 0
    if text == "c" or text.startswith("chieu") or "afternoon" in text:
        return 1 if session_count > 1 else 0
    if any(token in text for token in ("buoi 1", "ca 1")):
        return 0
    if any(token in text for token in ("buoi 2", "ca 2")):
        return 1 if session_count > 1 else 0
    match = re.search(r"(?:buoi|ca|session)\s*([1-9])\b", text)
    if match:
        index = int(match.group(1)) - 1
        return index if 0 <= index < session_count else None
    return None


def _period_number(value: str) -> int | None:
    text = normalize_text(value)
    if not text:
        return None
    match = re.search(r"\b(\d{1,2})\b", text)
    return int(match.group(1)) if match else None


def _resolve_slot(day_text: str, session_text: str, period_text: str, project: dict[str, Any]) -> int | None:
    days = int(project["days"])
    sessions = int(project["sessions"])
    periods = int(project["periods"])
    day = _day_index(day_text)
    period = _period_number(period_text)
    if day is None or period is None or not (0 <= day < days):
        return None
    session = _session_index(session_text, sessions)
    if session is None and 1 <= period <= sessions * periods:
        session = (period - 1) // periods
        period = ((period - 1) % periods) + 1
    if session is None:
        session = 0 if sessions == 1 else None
    if session is None or not (0 <= session < sessions) or not (1 <= period <= periods):
        return None
    return day * sessions * periods + session * periods + period - 1


def slot_label(slot: int, project: dict[str, Any]) -> str:
    periods = int(project["periods"])
    sessions = int(project["sessions"])
    per_day = periods * sessions
    day = slot // per_day
    inside = slot % per_day
    session = inside // periods
    period = inside % periods + 1
    day_name = ["Thứ 2", "Thứ 3", "Thứ 4", "Thứ 5", "Thứ 6", "Thứ 7", "CN"][day]
    if sessions == 1:
        return f"{day_name}, tiết {period}"
    session_name = "Sáng" if session == 0 else ("Chiều" if session == 1 else f"Buổi {session + 1}")
    return f"{day_name}, {session_name.lower()}, tiết {period}"


def _trim_matrix(rows: list[list[str]]) -> list[list[str]]:
    cleaned = []
    for row in rows:
        values = list(row)
        while values and not _cell_text(values[-1]):
            values.pop()
        cleaned.append([_cell_text(value) for value in values])
    while cleaned and not any(cleaned[-1]):
        cleaned.pop()
    return cleaned


def _parse_long_table(title: str, rows: list[list[str]]) -> list[RawLesson]:
    for header_index, row in enumerate(rows[:30]):
        kinds: dict[str, int] = {}
        for col, value in enumerate(row):
            kind = _header_kind(value)
            if kind and kind not in kinds:
                kinds[kind] = col
        if not {"day", "period", "class"}.issubset(kinds):
            continue
        if not ({"subject", "lesson", "teacher"} & set(kinds)):
            continue
        parsed: list[RawLesson] = []
        last_day = ""
        last_session = ""
        for row_no, values in enumerate(rows[header_index + 1 :], start=header_index + 2):
            get = lambda key: values[kinds[key]] if key in kinds and kinds[key] < len(values) else ""
            day = get("day") or last_day
            session = get("session") or last_session
            period = get("period")
            class_text = get("class")
            subject = get("subject")
            lesson = get("lesson")
            teacher = get("teacher")
            room = get("room")
            if get("day"):
                last_day = get("day")
            if get("session"):
                last_session = get("session")
            if not period or not class_text or not (subject or lesson or teacher):
                continue
            parsed.append(RawLesson(
                day, session, period, class_text, lesson, subject, teacher, room,
                f"{title} · dòng {row_no}", "aggregate",
            ))
        if parsed:
            return parsed
    return []


def _parse_wide_table(title: str, rows: list[list[str]], class_aliases: dict[str, int]) -> list[RawLesson]:
    ignored_headers = {
        "gv nghi", "giao vien nghi", "ghi chu", "note", "notes", "phong", "room", "stt", "so tt",
    }
    best: tuple[int, int, dict[str, int], list[tuple[int, str]]] | None = None
    for header_index, row in enumerate(rows[:30]):
        kinds: dict[str, int] = {}
        for col, value in enumerate(row):
            kind = _header_kind(value)
            if kind in {"day", "session", "period"} and kind not in kinds:
                kinds[kind] = col
        if "period" not in kinds or "day" not in kinds:
            continue
        class_cols: list[tuple[int, str]] = []
        period_col = kinds["period"]
        for col, value in enumerate(row):
            header = _cell_text(value)
            norm = normalize_text(header)
            if col in kinds.values() or not header or norm in ignored_headers:
                continue
            recognized = _exact_alias(header, class_aliases) is not None
            if recognized or col > period_col:
                class_cols.append((col, header))
        score = sum(3 if _exact_alias(header, class_aliases) is not None else 1 for _col, header in class_cols)
        if class_cols and (best is None or score > best[0]):
            best = (score, header_index, kinds, class_cols)
    if best is None:
        return []
    _score, header_index, kinds, class_cols = best
    parsed: list[RawLesson] = []
    last_day = ""
    last_session = ""
    for row_no, values in enumerate(rows[header_index + 1 :], start=header_index + 2):
        get = lambda col: values[col] if col is not None and col < len(values) else ""
        day_value = get(kinds.get("day"))
        session_value = get(kinds.get("session"))
        period = get(kinds.get("period"))
        if day_value:
            last_day = day_value
        if session_value:
            last_session = session_value
        day = day_value or last_day
        session = session_value or last_session
        if not period or _period_number(period) is None:
            continue
        for col, class_text in class_cols:
            lesson = get(col)
            if not lesson or normalize_text(lesson) in {"x", "trong", "nghi", "off", "none", "na"}:
                continue
            parsed.append(RawLesson(
                day, session, period, class_text, lesson_text=lesson,
                source=f"{title} · dòng {row_no}", origin="class",
            ))
    return parsed


def _parse_class_day_grid(title: str, rows: list[list[str]], class_aliases: dict[str, int]) -> list[RawLesson]:
    class_id = _exact_alias(title, class_aliases)
    if class_id is None:
        for alias in sorted(class_aliases, key=len, reverse=True):
            if alias and f" {alias} " in f" {normalize_text(title)} ":
                class_id = class_aliases[alias]
                break
    if class_id is None:
        return []
    class_alias = next((alias for alias, item_id in class_aliases.items() if item_id == class_id), title)
    for header_index, row in enumerate(rows[:25]):
        day_cols = [(col, value) for col, value in enumerate(row) if _day_index(value) is not None]
        if len(day_cols) < 2:
            continue
        period_col = next((col for col, value in enumerate(row) if _header_kind(value) == "period"), 0)
        session_col = next((col for col, value in enumerate(row) if _header_kind(value) == "session"), None)
        parsed: list[RawLesson] = []
        last_session = ""
        for row_no, values in enumerate(rows[header_index + 1 :], start=header_index + 2):
            period = values[period_col] if period_col < len(values) else ""
            if not period or _period_number(period) is None:
                continue
            session_value = values[session_col] if session_col is not None and session_col < len(values) else ""
            if session_value:
                last_session = session_value
            session = session_value or last_session
            for col, day_text in day_cols:
                lesson = values[col] if col < len(values) else ""
                if lesson and normalize_text(lesson) not in {"x", "trong", "nghi", "off", "none", "na"}:
                    parsed.append(RawLesson(
                        day_text, session, period, class_alias, lesson_text=lesson,
                        source=f"{title} · dòng {row_no}", origin="class",
                    ))
        if parsed:
            return parsed
    return []


def _entity_id_from_heading(
    title: str,
    rows_before_header: list[list[str]],
    aliases: dict[str, int],
) -> int | None:
    candidates: list[int] = []
    for value in [title, *(cell for row in rows_before_header for cell in row)]:
        item_id = _find_alias(_cell_text(value), aliases)
        if item_id is not None:
            candidates.append(item_id)
    unique = set(candidates)
    return next(iter(unique)) if len(unique) == 1 else None


def _parse_teacher_day_grid(
    title: str,
    rows: list[list[str]],
    class_aliases: dict[str, int],
    teacher_aliases: dict[str, int],
    classes_by_id: dict[int, dict[str, Any]],
    teachers_by_id: dict[int, dict[str, Any]],
) -> list[RawLesson]:
    for header_index, row in enumerate(rows[:25]):
        day_cols = [(col, value) for col, value in enumerate(row) if _day_index(value) is not None]
        if len(day_cols) < 2:
            continue
        teacher_id = _entity_id_from_heading(title, rows[:header_index], teacher_aliases)
        if teacher_id is None:
            continue
        teacher_name = teachers_by_id.get(teacher_id, {}).get("name", title)
        period_col = next((col for col, value in enumerate(row) if _header_kind(value) == "period"), 0)
        session_col = next((col for col, value in enumerate(row) if _header_kind(value) == "session"), None)
        parsed: list[RawLesson] = []
        last_session = ""
        for row_no, values in enumerate(rows[header_index + 1 :], start=header_index + 2):
            period = values[period_col] if period_col < len(values) else ""
            if not period or _period_number(period) is None:
                continue
            session_value = values[session_col] if session_col is not None and session_col < len(values) else ""
            if session_value:
                last_session = session_value
            session = session_value or last_session
            for col, day_text in day_cols:
                lesson = values[col] if col < len(values) else ""
                if not lesson or normalize_text(lesson) in {"x", "trong", "nghi", "off", "none", "na"}:
                    continue
                class_id = _find_alias(lesson, class_aliases)
                class_text = classes_by_id.get(class_id, {}).get("name", lesson) if class_id is not None else lesson
                parsed.append(RawLesson(
                    day_text, session, period, class_text,
                    lesson_text=lesson, teacher_text=teacher_name,
                    source=f"{title} · dòng {row_no}", origin="teacher",
                ))
        if parsed:
            return parsed
    return []


def _parse_teacher_wide_table(
    title: str,
    rows: list[list[str]],
    class_aliases: dict[str, int],
    teacher_aliases: dict[str, int],
    classes_by_id: dict[int, dict[str, Any]],
    teachers_by_id: dict[int, dict[str, Any]],
) -> list[RawLesson]:
    best: tuple[int, int, dict[str, int], list[tuple[int, int]]] | None = None
    for header_index, row in enumerate(rows[:30]):
        kinds: dict[str, int] = {}
        for col, value in enumerate(row):
            kind = _header_kind(value)
            if kind in {"day", "session", "period"} and kind not in kinds:
                kinds[kind] = col
        if "period" not in kinds or "day" not in kinds:
            continue
        teacher_cols: list[tuple[int, int]] = []
        for col, value in enumerate(row):
            if col in kinds.values():
                continue
            teacher_id = _exact_alias(_cell_text(value), teacher_aliases)
            if teacher_id is not None:
                teacher_cols.append((col, teacher_id))
        if teacher_cols and (best is None or len(teacher_cols) > best[0]):
            best = (len(teacher_cols), header_index, kinds, teacher_cols)
    if best is None:
        return []

    _score, header_index, kinds, teacher_cols = best
    parsed: list[RawLesson] = []
    last_day = ""
    last_session = ""
    for row_no, values in enumerate(rows[header_index + 1 :], start=header_index + 2):
        get = lambda col: values[col] if col is not None and col < len(values) else ""
        day_value = get(kinds.get("day"))
        session_value = get(kinds.get("session"))
        period = get(kinds.get("period"))
        if day_value:
            last_day = day_value
        if session_value:
            last_session = session_value
        day = day_value or last_day
        session = session_value or last_session
        if not period or _period_number(period) is None:
            continue
        for col, teacher_id in teacher_cols:
            lesson = get(col)
            if not lesson or normalize_text(lesson) in {"x", "trong", "nghi", "off", "none", "na"}:
                continue
            class_id = _find_alias(lesson, class_aliases)
            class_text = classes_by_id.get(class_id, {}).get("name", lesson) if class_id is not None else lesson
            teacher_name = teachers_by_id.get(teacher_id, {}).get("name", _cell_text(rows[header_index][col]))
            parsed.append(RawLesson(
                day, session, period, class_text,
                lesson_text=lesson, teacher_text=teacher_name,
                source=f"{title} · dòng {row_no}", origin="teacher",
            ))
    return parsed


def _header_entity_ids(rows: list[list[str]], aliases: dict[str, int]) -> set[int]:
    best: set[int] = set()
    for row in rows[:30]:
        has_axis = any(_header_kind(value) in {"day", "period"} for value in row)
        if not has_axis:
            continue
        ids = {
            item_id for value in row
            if (item_id := _exact_alias(_cell_text(value), aliases)) is not None
        }
        if len(ids) > len(best):
            best = ids
    return best


def parse_tables(
    tables: list[tuple[str, list[list[str]]]],
    classes: list[dict[str, Any]],
    teachers: list[dict[str, Any]],
) -> tuple[list[RawLesson], list[str], dict[str, Any]]:
    class_aliases = _alias_map(classes, ("name",))
    teacher_aliases = _alias_map(teachers, ("name", "short_name"))
    classes_by_id = {int(item["id"]): item for item in classes}
    teachers_by_id = {int(item["id"]): item for item in teachers}
    all_class_ids = set(classes_by_id)
    all_teacher_ids = set(teachers_by_id)
    class_scope: set[int] = set()
    teacher_scope: set[int] = set()
    layouts: set[str] = set()
    full_project = False
    all_rows: list[RawLesson] = []
    warnings: list[str] = []

    for title, raw_rows in tables:
        rows = _trim_matrix(raw_rows)
        if not rows:
            continue
        parsed: list[RawLesson] = []
        layout = ""

        parsed = _parse_long_table(title, rows)
        if parsed:
            layout = "Bảng tổng hợp"
            title_class_id = _find_alias(title, class_aliases)
            title_teacher_id = _find_alias(title, teacher_aliases)
            if title_class_id is not None:
                class_scope.add(title_class_id)
            elif title_teacher_id is not None:
                teacher_scope.add(title_teacher_id)
            else:
                full_project = True

        if not parsed:
            parsed = _parse_class_day_grid(title, rows, class_aliases)
            if parsed:
                layout = "Theo lớp/học sinh"
                for raw in parsed:
                    class_id = _exact_alias(raw.class_text, class_aliases)
                    if class_id is not None:
                        class_scope.add(class_id)

        if not parsed:
            parsed = _parse_teacher_day_grid(
                title, rows, class_aliases, teacher_aliases, classes_by_id, teachers_by_id,
            )
            if parsed:
                layout = "Theo giáo viên"
                for raw in parsed:
                    teacher_id = _find_alias(raw.teacher_text, teacher_aliases)
                    if teacher_id is not None:
                        teacher_scope.add(teacher_id)

        if not parsed:
            parsed = _parse_teacher_wide_table(
                title, rows, class_aliases, teacher_aliases, classes_by_id, teachers_by_id,
            )
            if parsed:
                layout = "Bảng tổng hợp theo giáo viên"
                teacher_scope.update(_header_entity_ids(rows, teacher_aliases))

        if not parsed:
            parsed = _parse_wide_table(title, rows, class_aliases)
            if parsed:
                layout = "Bảng tổng hợp theo lớp"
                class_scope.update(_header_entity_ids(rows, class_aliases))

        if parsed:
            all_rows.extend(parsed)
            layouts.add(layout)
        else:
            looks_like_schedule = (
                _find_alias(title, class_aliases) is not None
                or _find_alias(title, teacher_aliases) is not None
                or any(
                    _header_kind(value) in {"day", "period"}
                    for row in rows[:30]
                    for value in row
                )
            )
            if looks_like_schedule:
                warnings.append(f"Không nhận diện được cấu trúc thời khóa biểu trong “{title}”.")

    if class_scope and class_scope == all_class_ids:
        full_project = True
    if teacher_scope and teacher_scope == all_teacher_ids:
        full_project = True

    detection = {
        "layouts": sorted(layouts),
        "full_project": full_project,
        "class_ids": sorted(class_scope),
        "teacher_ids": sorted(teacher_scope),
    }
    return all_rows, warnings, detection


def _issue(
    code: str,
    severity: str,
    title: str,
    detail: str,
    *,
    slot: int | None = None,
    project: dict[str, Any] | None = None,
    source: str = "",
    entity: str = "",
) -> dict[str, Any]:
    return {
        "code": code,
        "severity": severity,
        "title": title,
        "detail": detail,
        "slot": slot,
        "slot_label": slot_label(slot, project) if slot is not None and project is not None else "",
        "source": source,
        "entity": entity,
    }


def _consecutive_runs(slots: Iterable[int], project: dict[str, Any]) -> list[list[int]]:
    periods = int(project["periods"])
    sessions = int(project["sessions"])
    per_day = periods * sessions
    grouped: dict[tuple[int, int], list[int]] = defaultdict(list)
    for slot in sorted(set(int(value) for value in slots)):
        day = slot // per_day
        inside = slot % per_day
        session = inside // periods
        grouped[(day, session)].append(slot)
    runs: list[list[int]] = []
    for values in grouped.values():
        run: list[int] = []
        for slot in values:
            if run and slot != run[-1] + 1:
                runs.append(run)
                run = []
            run.append(slot)
        if run:
            runs.append(run)
    return runs


def required_double_pattern_ok(slots: Iterable[int], total_periods: int, project: dict[str, Any]) -> bool:
    values = sorted(set(int(slot) for slot in slots))
    if len(values) != total_periods:
        return False
    expected_pairs = total_periods // 2
    expected_single = total_periods % 2
    runs = _consecutive_runs(values, project)
    pair_count = sum(len(run) // 2 for run in runs)
    leftover = sum(len(run) % 2 for run in runs)
    return pair_count >= expected_pairs and leftover == expected_single


def analyze_schedule_file(
    *,
    filename: str,
    content: bytes,
    project: dict[str, Any],
    classes: list[dict[str, Any]],
    subjects: list[dict[str, Any]],
    teachers: list[dict[str, Any]],
    assignments: list[dict[str, Any]],
    include_editable: bool = False,
) -> dict[str, Any]:
    file_format, tables = read_tables(filename, content)
    raw_lessons, parse_warnings, detection = parse_tables(tables, classes, teachers)
    if not raw_lessons:
        raise ScheduleAuditParseError(
            "Không tìm thấy tiết học nào có thể đọc được. File có thể là bảng tổng hợp, TKB theo lớp/học sinh hoặc TKB theo giáo viên; cần có thông tin Thứ/Tiết và lớp tương ứng."
        )

    class_aliases = _alias_map(classes, ("name",))
    subject_aliases = _alias_map(subjects, ("name", "short_name"))
    teacher_aliases = _alias_map(teachers, ("name", "short_name"))
    classes_by_id = {int(item["id"]): item for item in classes}
    subjects_by_id = {int(item["id"]): item for item in subjects}
    teachers_by_id = {int(item["id"]): item for item in teachers}
    assignments_by_id = {int(item["id"]): item for item in assignments}
    assignment_by_pair = {(int(item["class_id"]), int(item["subject_id"])): item for item in assignments}
    assignments_by_class_teacher: dict[tuple[int, int], list[dict[str, Any]]] = defaultdict(list)
    for assignment in assignments:
        assignments_by_class_teacher[(int(assignment["class_id"]), int(assignment["teacher_id"]))].append(assignment)

    issues: list[dict[str, Any]] = []
    recognized: list[dict[str, Any]] = []
    global_blocked = {int(value) for value in project.get("blocked_slots", [])}

    for raw in raw_lessons:
        slot = _resolve_slot(raw.day_text, raw.session_text, raw.period_text, project)
        if slot is None:
            issues.append(_issue(
                "invalid_slot", "error", "Không xác định được tiết học",
                f"Không đổi được “{raw.day_text} / {raw.session_text or '—'} / {raw.period_text}” thành một ô hợp lệ của project.",
                source=raw.source,
            ))
            continue
        class_id = _exact_alias(raw.class_text, class_aliases)
        if class_id is None:
            issues.append(_issue(
                "unknown_class", "error", "Không nhận diện được lớp",
                f"Tên lớp “{raw.class_text}” không khớp với lớp nào trong project.", slot=slot, project=project, source=raw.source,
            ))
            continue

        combined = " ".join(part for part in (raw.lesson_text, raw.subject_text, raw.teacher_text) if part)
        subject_id = _find_alias(raw.subject_text or raw.lesson_text, subject_aliases)
        explicit_teacher_id = _find_alias(raw.teacher_text or raw.lesson_text, teacher_aliases)
        teacher_id = explicit_teacher_id

        if subject_id is None and teacher_id is not None:
            candidates = assignments_by_class_teacher.get((class_id, teacher_id), [])
            if len(candidates) == 1:
                subject_id = int(candidates[0]["subject_id"])
        assignment = assignment_by_pair.get((class_id, subject_id)) if subject_id is not None else None
        if teacher_id is None and assignment is not None:
            teacher_id = int(assignment["teacher_id"])

        if subject_id is None:
            issues.append(_issue(
                "unknown_subject", "error", "Không nhận diện được môn học",
                f"Ô “{combined or raw.lesson_text}” của lớp {classes_by_id[class_id]['name']} không khớp môn học/phân công nào.",
                slot=slot, project=project, source=raw.source, entity=classes_by_id[class_id]["name"],
            ))
            continue
        subject = subjects_by_id.get(subject_id, {"name": raw.subject_text or "?"})
        if assignment is None:
            issues.append(_issue(
                "missing_assignment", "error", "Tiết không có trong phân công",
                f"Lớp {classes_by_id[class_id]['name']} có môn {subject['name']} trong file nhưng project chưa có phân công tương ứng.",
                slot=slot, project=project, source=raw.source, entity=classes_by_id[class_id]["name"],
            ))
            continue

        expected_teacher_id = int(assignment["teacher_id"])
        if explicit_teacher_id is not None and explicit_teacher_id != expected_teacher_id:
            actual = teachers_by_id.get(explicit_teacher_id, {"name": raw.teacher_text or "?"})
            expected = teachers_by_id.get(expected_teacher_id, {"name": "?"})
            issues.append(_issue(
                "wrong_teacher", "error", "Sai giáo viên phân công",
                f"{classes_by_id[class_id]['name']} · {subject['name']} đang ghi {actual['name']}, trong project phân công cho {expected['name']}.",
                slot=slot, project=project, source=raw.source, entity=actual["name"],
            ))
        if teacher_id is None:
            teacher_id = expected_teacher_id

        recognized.append({
            "slot": slot,
            "class_id": class_id,
            "subject_id": subject_id,
            "teacher_id": teacher_id,
            "assignment_id": int(assignment["id"]),
            "room": raw.room_text.strip(),
            "source": raw.source,
            "lesson_text": raw.lesson_text or raw.subject_text,
            "origin": raw.origin,
        })

    # Một workbook có thể chứa đồng thời TKB theo lớp và theo giáo viên.
    # Cùng một tiết xuất hiện ở hai góc nhìn chỉ được tính một lần; bản sao
    # lặp trong cùng một kiểu bảng vẫn được giữ để phát hiện dữ liệu trùng.
    deduplicated: list[dict[str, Any]] = []
    first_by_key: dict[tuple[int, int, int, int], dict[str, Any]] = {}
    for entry in recognized:
        key = (
            int(entry["slot"]), int(entry["class_id"]),
            int(entry["subject_id"]), int(entry["teacher_id"]),
        )
        existing = first_by_key.get(key)
        if existing is not None and entry["origin"] not in existing["_origins"]:
            existing["_origins"].add(entry["origin"])
            if entry["source"] and entry["source"] not in existing["source"]:
                existing["source"] += f"; {entry['source']}"
            if not existing["room"] and entry["room"]:
                existing["room"] = entry["room"]
            continue
        entry["_origins"] = {entry["origin"]}
        deduplicated.append(entry)
        if existing is None:
            first_by_key[key] = entry
    recognized = deduplicated

    conflict_codes_by_draft: dict[int, set[str]] = defaultdict(set)
    conflict_details_by_draft: dict[int, list[str]] = defaultdict(list)
    by_class_slot: dict[tuple[int, int], list[dict[str, Any]]] = defaultdict(list)
    by_teacher_slot: dict[tuple[int, int], list[dict[str, Any]]] = defaultdict(list)
    by_room_slot: dict[tuple[int, str], list[dict[str, Any]]] = defaultdict(list)
    by_assignment: dict[int, list[dict[str, Any]]] = defaultdict(list)
    by_teacher_day: dict[tuple[int, int], list[dict[str, Any]]] = defaultdict(list)
    by_class_subject: dict[tuple[int, int], list[dict[str, Any]]] = defaultdict(list)
    periods_per_day = int(project["sessions"]) * int(project["periods"])

    for entry in recognized:
        slot = int(entry["slot"])
        class_id = int(entry["class_id"])
        teacher_id = int(entry["teacher_id"])
        by_class_slot[(slot, class_id)].append(entry)
        by_teacher_slot[(slot, teacher_id)].append(entry)
        if normalize_text(entry["room"]):
            by_room_slot[(slot, normalize_text(entry["room"]))].append(entry)
        by_assignment[int(entry["assignment_id"])].append(entry)
        by_teacher_day[(teacher_id, slot // periods_per_day)].append(entry)
        by_class_subject[(class_id, int(entry["subject_id"]))].append(entry)

        teacher = teachers_by_id.get(teacher_id)
        school_class = classes_by_id.get(class_id)
        if slot in global_blocked:
            issues.append(_issue(
                "global_blocked", "error", "Xếp vào tiết toàn trường đã khóa",
                f"{school_class['name']} · {subjects_by_id[entry['subject_id']]['name']} nằm trong ô đã khóa toàn trường.",
                slot=slot, project=project, source=entry["source"],
            ))
        if teacher and slot in {int(value) for value in teacher.get("unavailable", [])}:
            issues.append(_issue(
                "teacher_unavailable", "error", "Giáo viên bị xếp vào tiết nghỉ",
                f"{teacher['name']} đang được xếp dạy {school_class['name']} dù tiết này đã đánh dấu không thể dạy.",
                slot=slot, project=project, source=entry["source"], entity=teacher["name"],
            ))
        if school_class and slot in {int(value) for value in school_class.get("unavailable", [])}:
            issues.append(_issue(
                "class_unavailable", "error", "Lớp bị xếp vào tiết nghỉ",
                f"{school_class['name']} có tiết học trong ô đã đánh dấu lớp không học.",
                slot=slot, project=project, source=entry["source"], entity=school_class["name"],
            ))

    for (slot, class_id), rows in by_class_slot.items():
        if len(rows) > 1:
            names = ", ".join(subjects_by_id[row["subject_id"]]["name"] for row in rows)
            issues.append(_issue(
                "class_collision", "error", "Trùng lịch lớp",
                f"Lớp {classes_by_id[class_id]['name']} có {len(rows)} tiết cùng lúc: {names}.",
                slot=slot, project=project, source="; ".join(row["source"] for row in rows), entity=classes_by_id[class_id]["name"],
            ))

    for (slot, teacher_id), rows in by_teacher_slot.items():
        if len(rows) > 1:
            class_names = ", ".join(classes_by_id[row["class_id"]]["name"] for row in rows)
            teacher_name = teachers_by_id.get(teacher_id, {"name": "?"})["name"]
            issues.append(_issue(
                "teacher_collision", "error", "Trùng lịch giáo viên",
                f"{teacher_name} bị xếp dạy đồng thời {len(rows)} lớp: {class_names}.",
                slot=slot, project=project, source="; ".join(row["source"] for row in rows), entity=teacher_name,
            ))

    for (slot, _room_key), rows in by_room_slot.items():
        if len(rows) > 1:
            room = rows[0]["room"]
            class_names = ", ".join(classes_by_id[row["class_id"]]["name"] for row in rows)
            issues.append(_issue(
                "room_collision", "error", "Trùng phòng học",
                f"Phòng {room} được dùng đồng thời cho: {class_names}.", slot=slot, project=project,
                source="; ".join(row["source"] for row in rows), entity=room,
            ))

    missing_count = 0
    extra_count = 0
    class_scope = {int(value) for value in detection.get("class_ids", [])}
    teacher_scope = {int(value) for value in detection.get("teacher_ids", [])}
    if detection.get("full_project") or (not class_scope and not teacher_scope):
        audited_assignments = assignments
    else:
        audited_assignments = [
            assignment for assignment in assignments
            if int(assignment["class_id"]) in class_scope
            or int(assignment["teacher_id"]) in teacher_scope
        ]
    for assignment in audited_assignments:
        assignment_id = int(assignment["id"])
        rows = by_assignment.get(assignment_id, [])
        actual = len(rows)
        expected = int(assignment.get("periods_per_week", 0) or 0)
        label = f"{assignment.get('class_name', '?')} · {assignment.get('subject_name', '?')}"
        if actual < expected:
            missing = expected - actual
            missing_count += missing
            issues.append(_issue(
                "missing_lessons", "error", "Xếp thiếu tiết",
                f"{label}: cần {expected} tiết/tuần nhưng file chỉ có {actual}, thiếu {missing} tiết.", entity=label,
            ))
        elif actual > expected:
            extra = actual - expected
            extra_count += extra
            issues.append(_issue(
                "extra_lessons", "error", "Xếp thừa tiết",
                f"{label}: cần {expected} tiết/tuần nhưng file có {actual}, thừa {extra} tiết.", entity=label,
            ))
        elif expected and assignment.get("block_mode") == "required_double":
            slots = [int(row["slot"]) for row in rows]
            if not required_double_pattern_ok(slots, expected, project):
                issues.append(_issue(
                    "required_double_mismatch", "warning", "Chưa đúng cấu trúc tiết đôi bắt buộc",
                    f"{label} đủ {expected} tiết nhưng chưa tạo đúng số cặp tiết đôi theo cấu hình.", entity=label,
                ))

    for (teacher_id, day), rows in by_teacher_day.items():
        teacher = teachers_by_id.get(teacher_id)
        if not teacher:
            continue
        maximum = int(teacher.get("max_periods_day", 0) or 0)
        if maximum > 0 and len(rows) > maximum:
            day_name = ["Thứ 2", "Thứ 3", "Thứ 4", "Thứ 5", "Thứ 6", "Thứ 7", "CN"][day]
            issues.append(_issue(
                "teacher_daily_overload", "warning", "Giáo viên vượt số tiết tối đa/ngày",
                f"{teacher['name']} có {len(rows)} tiết vào {day_name}, vượt giới hạn {maximum} tiết/ngày.", entity=teacher["name"],
            ))

    for (class_id, subject_id), rows in by_class_subject.items():
        subject = subjects_by_id.get(subject_id)
        if not subject:
            continue
        maximum = int(subject.get("max_consecutive", 0) or 0)
        if maximum <= 0:
            continue
        slots = [int(row["slot"]) for row in rows]
        for run in _consecutive_runs(slots, project):
            if len(run) > maximum:
                issues.append(_issue(
                    "subject_consecutive_overload", "warning", "Môn học vượt số tiết liên tiếp",
                    f"{classes_by_id[class_id]['name']} · {subject['name']} có {len(run)} tiết liên tiếp, vượt giới hạn {maximum}.",
                    slot=run[0], project=project, entity=f"{classes_by_id[class_id]['name']} · {subject['name']}",
                ))

    for warning in parse_warnings:
        issues.append(_issue("unread_table", "warning", "Có bảng/sheet chưa đọc được", warning))

    severity_order = {"error": 0, "warning": 1, "info": 2}
    issues.sort(key=lambda item: (severity_order.get(item["severity"], 9), item.get("slot") is None, item.get("slot") or -1, item["title"]))
    errors = sum(1 for item in issues if item["severity"] == "error")
    warnings = sum(1 for item in issues if item["severity"] == "warning")
    collisions = sum(1 for item in issues if item["code"] in {"teacher_collision", "class_collision", "room_collision"})

    detection["classes"] = [classes_by_id[item_id]["name"] for item_id in detection.get("class_ids", []) if item_id in classes_by_id]
    detection["teachers"] = [teachers_by_id[item_id]["name"] for item_id in detection.get("teacher_ids", []) if item_id in teachers_by_id]
    if detection.get("full_project"):
        detection["scope_label"] = "Toàn bộ project"
    elif detection["classes"] and detection["teachers"]:
        detection["scope_label"] = f"{len(detection['classes'])} lớp · {len(detection['teachers'])} giáo viên"
    elif detection["classes"]:
        detection["scope_label"] = "Lớp: " + ", ".join(detection["classes"])
    elif detection["teachers"]:
        detection["scope_label"] = "Giáo viên: " + ", ".join(detection["teachers"])
    else:
        detection["scope_label"] = "Phần dữ liệu đã nhận diện"

    result = {
        "ok": True,
        "filename": filename,
        "format": file_format,
        "detection": detection,
        "status": "clean" if errors == 0 and warnings == 0 else ("error" if errors else "warning"),
        "summary": {
            "read_lessons": len(raw_lessons),
            "recognized_lessons": len(recognized),
            "errors": errors,
            "warnings": warnings,
            "collisions": collisions,
            "missing_periods": missing_count,
            "extra_periods": extra_count,
        },
        "issues": issues,
    }
    if include_editable:
        result["editable_lessons"] = [
            {
                "draft_id": index,
                "assignment_id": int(entry["assignment_id"]),
                "slot": int(entry["slot"]),
                "source": entry.get("source", ""),
                "room": entry.get("room", ""),
            }
            for index, entry in enumerate(recognized, start=1)
        ]
    return result

# ===== Standalone import/audit (khong phu thuoc project co san) =====

_STANDALONE_SUBJECT_HINTS = {
    "hdtnhn": "HĐTNHN",
    "tnhn": "TNHN",
    "hdtn": "HĐTN",
    "toan": "Toán",
    "ngu van": "Ngữ văn",
    "n van": "Ngữ văn",
    "van": "Ngữ văn",
    "tieng anh": "Tiếng Anh",
    "t anh": "Tiếng Anh",
    "anh": "Tiếng Anh",
    "vat ly": "Vật lý",
    "ly": "Vật lý",
    "hoa hoc": "Hóa học",
    "hoa": "Hóa học",
    "sinh hoc": "Sinh học",
    "sinh": "Sinh học",
    "lich su": "Lịch sử",
    "su": "Lịch sử",
    "dia ly": "Địa lý",
    "dia": "Địa lý",
    "lsdl": "LSĐL",
    "gdcd": "GDCD",
    "gdktpl": "GDKTPL",
    "gddp": "GDĐP",
    "tin hoc": "Tin học",
    "tin": "Tin học",
    "cong nghe": "Công nghệ",
    "c nghe": "Công nghệ",
    "cn": "Công nghệ",
    "the duc": "Thể dục",
    "gdtc": "GDTC",
    "am nhac": "Âm nhạc",
    "nt nhac": "Âm nhạc",
    "my thuat": "Mỹ thuật",
    "nt mt": "Mỹ thuật",
    "quoc phong": "Quốc phòng",
    "gdqp": "GDQP",
    "khtn": "KHTN",
    "khxh": "KHXH",
    "trai nghiem": "Trải nghiệm",
    "chao co": "Chào cờ",
    "sinh hoat": "Sinh hoạt",
}

# Cac mau o TKB pho bien: "N.Van Que", "KHTN(S) N.Tam", "HDTNHN Que"...
# Match prefix tren chuoi goc de giu lai ten giao vien phia sau chinh xac.
_STANDALONE_SUBJECT_PREFIX_PATTERNS: tuple[tuple[re.Pattern[str], str], ...] = tuple(
    (re.compile(pattern, re.IGNORECASE), subject)
    for pattern, subject in (
        (r"^\s*HĐTNHN\s*[-:]?\s*", "HĐTNHN"),
        (r"^\s*TNHN\s*[-:]?\s*", "TNHN"),
        (r"^\s*HĐTN\s*[-:]?\s*", "HĐTN"),
        (r"^\s*N\.?\s*Văn\s*[-:]?\s*", "Ngữ văn"),
        (r"^\s*(?:Ngữ\s*)?Văn\s*[-:]?\s*", "Ngữ văn"),
        (r"^\s*T\.?\s*Anh\s*[-:]?\s*", "Tiếng Anh"),
        (r"^\s*(?:Tiếng\s*)?Anh\s*[-:]?\s*", "Tiếng Anh"),
        (r"^\s*C\.?\s*Nghệ\s*[-:]?\s*", "Công nghệ"),
        (r"^\s*CN\s*[-:]?\s*", "Công nghệ"),
        (r"^\s*LSĐL(?:\s*\([^)]*\))?\s*[-:]?\s*", "LSĐL"),
        (r"^\s*KHTN(?:\s*\([^)]*\))?\s*[-:]?\s*", "KHTN"),
        (r"^\s*NT\s*\(\s*Nhạc\s*\)\s*[-:]?\s*", "Âm nhạc"),
        (r"^\s*NT\s*\(\s*MT\s*\)\s*[-:]?\s*", "Mỹ thuật"),
        (r"^\s*GDĐP\s*[-:]?\s*", "GDĐP"),
        (r"^\s*GDTC\s*[-:]?\s*", "GDTC"),
        (r"^\s*GDCD\s*[-:]?\s*", "GDCD"),
        (r"^\s*Tin(?:\s*học)?\s*[-:]?\s*", "Tin học"),
        (r"^\s*Toán\s*[-:]?\s*", "Toán"),
    )
)

def _standalone_subject_prefix(value: str) -> tuple[str, str]:
    text = _cell_text(value)
    for pattern, subject in _STANDALONE_SUBJECT_PREFIX_PATTERNS:
        match = pattern.match(text)
        if match:
            return subject, text[match.end():].strip(" -–—|;/")
    return "", ""



def _standalone_clean_entity_heading(value: str) -> str:
    text = _cell_text(value)
    text = re.sub(
        r"^\s*(?:tkb|thoi\s*khoa\s*bieu)\s*(?:cua\s*)?(?:lop|giao\s*vien|gv)?\s*[:\-]?\s*",
        "",
        text,
        flags=re.IGNORECASE,
    ).strip()
    text = re.sub(r"^\s*(?:lop|giao\s*vien|gv)\s*[:\-]?\s*", "", text, flags=re.IGNORECASE).strip()
    return text or _cell_text(value)


def _standalone_class_token(value: str) -> str:
    text = _cell_text(value)
    if not text:
        return ""
    patterns = [
        r"(?i)(?:\blop\s*)?(\d{1,2}\s*[A-Z]{1,3}\s*\d{0,2})\b",
        r"(?i)(?:\blop\s*)?(\d{1,2}\s*[/\-]\s*\d{1,2})\b",
    ]
    for pattern in patterns:
        match = re.search(pattern, text)
        if match:
            return re.sub(r"\s+", "", match.group(1)).upper()
    normalized = normalize_text(text)
    match = re.fullmatch(r"(?:lop\s*)?(\d{1,2})\s*([a-z]{1,3})\s*(\d{0,2})", normalized)
    if match:
        return f"{match.group(1)}{match.group(2).upper()}{match.group(3)}"
    return ""


def _standalone_looks_like_class(value: str) -> bool:
    return bool(_standalone_class_token(value))


def _standalone_subject_from_text(value: str) -> str:
    subject, _teacher = _standalone_subject_prefix(value)
    if subject:
        return subject
    text = _cell_text(value)
    norm = normalize_text(text)
    if not norm:
        return ""
    for hint in sorted(_STANDALONE_SUBJECT_HINTS, key=len, reverse=True):
        if re.search(rf"(?:^|\s){re.escape(hint)}(?:\s|$)", norm):
            return _STANDALONE_SUBJECT_HINTS[hint]
    return ""


def _standalone_split_parts(value: str) -> list[str]:
    text = _cell_text(value)
    if not text:
        return []
    parts = [part.strip() for part in re.split(r"\s*(?:\n|\r|\||;|·|•|\s[-–—]\s|\s/\s)\s*", text) if part.strip()]
    return parts or [text]


def _standalone_teacher_from_text(value: str, *, exclude: Iterable[str] = ()) -> str:
    text = _cell_text(value)
    if not text:
        return ""
    explicit = re.search(r"(?i)\b(?:giao\s*vien|gv)\s*[:\-]\s*([^|;/]+)", text)
    if explicit:
        return explicit.group(1).strip()
    excluded = {normalize_text(item) for item in exclude if item}
    _subject, remainder = _standalone_subject_prefix(text)
    if remainder and normalize_text(remainder) not in excluded and not _standalone_looks_like_class(remainder):
        return remainder
    norm_text = normalize_text(text)
    parts = _standalone_split_parts(text)
    subject_norms = set(_STANDALONE_SUBJECT_HINTS)
    for part in reversed(parts):
        norm = normalize_text(part)
        if not norm or norm in excluded or _standalone_looks_like_class(part):
            continue
        if any(re.search(rf"(?:^|\s){re.escape(subject)}(?:\s|$)", norm) for subject in subject_norms):
            continue
        # Ten giao vien thuong co it nhat 2 tu; chap nhan ma GV viet hoa ngan neu co tien to GV.
        if len(norm.split()) >= 2:
            return part.strip()
    return ""


def _standalone_parse_cell(
    value: str,
    *,
    fixed_class: str = "",
    fixed_teacher: str = "",
) -> tuple[str, str, str]:
    text = _cell_text(value)
    class_name = fixed_class or _standalone_class_token(text)
    subject_name = _standalone_subject_from_text(text)
    teacher_name = fixed_teacher
    if not teacher_name:
        teacher_name = _standalone_teacher_from_text(text, exclude=(class_name, subject_name))
    parts = _standalone_split_parts(text)
    if not subject_name:
        for part in parts:
            if class_name and normalize_text(part) == normalize_text(class_name):
                continue
            if teacher_name and normalize_text(part) == normalize_text(teacher_name):
                continue
            if _standalone_looks_like_class(part):
                continue
            norm = normalize_text(part)
            if norm and not norm.startswith(("gv ", "giao vien ")):
                subject_name = part.strip()
                break
    if not subject_name and fixed_teacher and class_name:
        # TKB giao vien co the chi ghi ten lop; van cho phep convert bang mon tam.
        subject_name = f"Mon chua xac dinh - {fixed_teacher}"
    if not teacher_name and fixed_class and subject_name:
        teacher_name = f"GV chua xac dinh - {fixed_class} - {subject_name}"
    return class_name.strip(), subject_name.strip(), teacher_name.strip()


def _standalone_day_index(value: str) -> int | None:
    text = normalize_text(value)
    if not text:
        return None
    exact = {
        "cn": 6, "chu nhat": 6, "sunday": 6, "sun": 6,
        "monday": 0, "mon": 0, "tuesday": 1, "tue": 1,
        "wednesday": 2, "wed": 2, "thursday": 3,
        "friday": 4, "fri": 4, "saturday": 5, "sat": 5,
    }
    if text in exact:
        return exact[text]
    match = re.fullmatch(r"(?:thu\s*)?([2-8])", text)
    if match:
        number = int(match.group(1))
        return 6 if number == 8 else number - 2
    return None

def _standalone_find_grid_header(rows: list[list[str]]) -> tuple[int, list[tuple[int, str]], int, int | None] | None:
    for header_index, row in enumerate(rows[:30]):
        day_cols = [(col, _cell_text(value)) for col, value in enumerate(row) if _standalone_day_index(_cell_text(value)) is not None]
        if len(day_cols) < 2:
            continue
        period_col = next((col for col, value in enumerate(row) if _header_kind(_cell_text(value)) == "period"), None)
        if period_col is None:
            period_col = 0
        session_col = next((col for col, value in enumerate(row) if _header_kind(_cell_text(value)) == "session"), None)
        return header_index, day_cols, period_col, session_col
    return None


def _standalone_parse_grid(title: str, rows: list[list[str]]) -> tuple[list[RawLesson], str]:
    header = _standalone_find_grid_header(rows)
    if header is None:
        return [], ""
    header_index, day_cols, period_col, session_col = header
    heading = _standalone_clean_entity_heading(title)
    sample_cells: list[str] = []
    for values in rows[header_index + 1 : header_index + 18]:
        for col, _day in day_cols:
            if col < len(values) and _cell_text(values[col]):
                sample_cells.append(_cell_text(values[col]))
    class_hits = sum(1 for cell in sample_cells if _standalone_class_token(cell))
    is_class_sheet = _standalone_looks_like_class(heading)
    if not is_class_sheet and sample_cells:
        is_teacher_sheet = class_hits >= max(1, len(sample_cells) // 3)
    else:
        is_teacher_sheet = False
    if not is_class_sheet and not is_teacher_sheet:
        # Ten sheet ngan khong giong tieu de chung thuong la ten lop; neu la ten nguoi
        # ma o co lop thi nhanh teacher o tren da bat duoc.
        is_class_sheet = bool(heading and normalize_text(heading) not in {"du lieu", "thoi khoa bieu", "tkb", "sheet", "sheet1"})
    if not is_class_sheet and not is_teacher_sheet:
        return [], ""

    parsed: list[RawLesson] = []
    last_session = ""
    for row_no, values in enumerate(rows[header_index + 1 :], start=header_index + 2):
        period = values[period_col] if period_col < len(values) else ""
        if not period or _period_number(period) is None:
            continue
        session_value = values[session_col] if session_col is not None and session_col < len(values) else ""
        if session_value:
            last_session = session_value
        session = session_value or last_session
        for col, day_text in day_cols:
            lesson = values[col] if col < len(values) else ""
            if not lesson or normalize_text(lesson) in {"x", "trong", "nghi", "off", "none", "na"}:
                continue
            if is_teacher_sheet:
                class_name, subject_name, teacher_name = _standalone_parse_cell(lesson, fixed_teacher=heading)
                if not class_name:
                    continue
                parsed.append(RawLesson(
                    day_text, session, period, class_name,
                    lesson_text=lesson, subject_text=subject_name, teacher_text=teacher_name,
                    source=f"{title} · dong {row_no}", origin="teacher",
                ))
            else:
                class_name, subject_name, teacher_name = _standalone_parse_cell(lesson, fixed_class=heading)
                parsed.append(RawLesson(
                    day_text, session, period, class_name or heading,
                    lesson_text=lesson, subject_text=subject_name, teacher_text=teacher_name,
                    source=f"{title} · dong {row_no}", origin="class",
                ))
    if not parsed:
        return [], ""
    return parsed, "Theo giao vien" if is_teacher_sheet else "Theo lop/hoc sinh"


def _standalone_parse_wide(title: str, rows: list[list[str]]) -> tuple[list[RawLesson], str]:
    best = None
    for header_index, row in enumerate(rows[:30]):
        kinds: dict[str, int] = {}
        for col, value in enumerate(row):
            kind = _header_kind(_cell_text(value))
            if kind in {"day", "session", "period"} and kind not in kinds:
                kinds[kind] = col
        if "day" not in kinds or "period" not in kinds:
            continue
        entity_cols = [(col, _cell_text(value)) for col, value in enumerate(row) if col not in kinds.values() and _cell_text(value)]
        if not entity_cols:
            continue
        score = len(entity_cols)
        if best is None or score > best[0]:
            best = (score, header_index, kinds, entity_cols)
    if best is None:
        return [], ""
    _score, header_index, kinds, entity_cols = best
    class_header_hits = sum(1 for _col, header in entity_cols if _standalone_looks_like_class(header))
    sampled = []
    for values in rows[header_index + 1 : header_index + 18]:
        for col, _header in entity_cols:
            if col < len(values) and _cell_text(values[col]):
                sampled.append(_cell_text(values[col]))
    cell_class_hits = sum(1 for cell in sampled if _standalone_class_token(cell))
    teacher_wide = class_header_hits == 0 and cell_class_hits >= max(1, len(sampled) // 3) if sampled else False
    class_wide = not teacher_wide

    parsed: list[RawLesson] = []
    last_day = ""
    last_session = ""
    for row_no, values in enumerate(rows[header_index + 1 :], start=header_index + 2):
        get = lambda col: values[col] if col is not None and col < len(values) else ""
        day_value = get(kinds.get("day"))
        session_value = get(kinds.get("session"))
        period = get(kinds.get("period"))
        if day_value:
            last_day = day_value
        if session_value:
            last_session = session_value
        day = day_value or last_day
        session = session_value or last_session
        if not day or not period or _period_number(period) is None:
            continue
        for col, header in entity_cols:
            lesson = get(col)
            if not lesson or normalize_text(lesson) in {"x", "trong", "nghi", "off", "none", "na"}:
                continue
            if teacher_wide:
                class_name, subject_name, teacher_name = _standalone_parse_cell(lesson, fixed_teacher=header)
                if not class_name:
                    continue
                parsed.append(RawLesson(
                    day, session, period, class_name,
                    lesson_text=lesson, subject_text=subject_name, teacher_text=teacher_name,
                    source=f"{title} · dong {row_no}", origin="teacher",
                ))
            else:
                class_name, subject_name, teacher_name = _standalone_parse_cell(lesson, fixed_class=header)
                parsed.append(RawLesson(
                    day, session, period, class_name or header,
                    lesson_text=lesson, subject_text=subject_name, teacher_text=teacher_name,
                    source=f"{title} · dong {row_no}", origin="class",
                ))
    if not parsed:
        return [], ""
    return parsed, "Bang tong hop theo giao vien" if teacher_wide else "Bang tong hop theo lop"


def parse_tables_standalone(
    tables: list[tuple[str, list[list[str]]]],
) -> tuple[list[RawLesson], list[str], dict[str, Any]]:
    all_rows: list[RawLesson] = []
    warnings: list[str] = []
    layouts: set[str] = set()
    for title, raw_rows in tables:
        rows = _trim_matrix(raw_rows)
        if not rows:
            continue
        parsed = _parse_long_table(title, rows)
        layout = "Bang tong hop" if parsed else ""
        if parsed:
            enriched = []
            for item in parsed:
                class_name, subject_name, teacher_name = _standalone_parse_cell(
                    item.lesson_text or item.subject_text or item.teacher_text,
                    fixed_class=item.class_text,
                    fixed_teacher=item.teacher_text,
                )
                item.class_text = class_name or item.class_text
                item.subject_text = item.subject_text or subject_name
                item.teacher_text = item.teacher_text or teacher_name
                enriched.append(item)
            parsed = enriched
        if not parsed:
            parsed, layout = _standalone_parse_grid(title, rows)
        if not parsed:
            parsed, layout = _standalone_parse_wide(title, rows)
        if parsed:
            all_rows.extend(parsed)
            layouts.add(layout)
        elif any(_header_kind(value) in {"day", "period"} for row in rows[:30] for value in row):
            warnings.append(f"Khong nhan dien duoc cau truc thoi khoa bieu trong \"{title}\".")
    return all_rows, warnings, {"layouts": sorted(layouts)}


def infer_standalone_project(
    tables: list[tuple[str, list[list[str]]]],
    raw_lessons: list[RawLesson],
) -> dict[str, Any]:
    day_indexes: set[int] = set()
    explicit_sessions: set[int] = set()
    period_numbers: list[int] = []
    for _title, raw_rows in tables:
        rows = _trim_matrix(raw_rows)
        grid = _standalone_find_grid_header(rows)
        if grid:
            header_index, day_cols, period_col, session_col = grid
            for _col, day_text in day_cols:
                day = _standalone_day_index(day_text)
                if day is not None:
                    day_indexes.add(day)
            for values in rows[header_index + 1 :]:
                if period_col < len(values):
                    number = _period_number(_cell_text(values[period_col]))
                    if number is not None:
                        period_numbers.append(number)
                if session_col is not None and session_col < len(values):
                    session = _session_index(_cell_text(values[session_col]), 2)
                    if session is not None:
                        explicit_sessions.add(session)
        else:
            for header_index, row in enumerate(rows[:30]):
                kinds = {_header_kind(_cell_text(value)): col for col, value in enumerate(row) if _header_kind(_cell_text(value))}
                if "period" not in kinds or "day" not in kinds:
                    continue
                for values in rows[header_index + 1 :]:
                    if kinds["day"] < len(values):
                        day = _standalone_day_index(_cell_text(values[kinds["day"]]))
                        if day is not None:
                            day_indexes.add(day)
                    if kinds["period"] < len(values):
                        number = _period_number(_cell_text(values[kinds["period"]]))
                        if number is not None:
                            period_numbers.append(number)
                    if "session" in kinds and kinds["session"] < len(values):
                        session = _session_index(_cell_text(values[kinds["session"]]), 2)
                        if session is not None:
                            explicit_sessions.add(session)
                break
    for raw in raw_lessons:
        day = _standalone_day_index(raw.day_text)
        if day is not None:
            day_indexes.add(day)
        number = _period_number(raw.period_text)
        if number is not None:
            period_numbers.append(number)
        if raw.session_text:
            session = _session_index(raw.session_text, 2)
            if session is not None:
                explicit_sessions.add(session)
    days = max(day_indexes) + 1 if day_indexes else 6
    days = max(5, min(7, days))
    maximum_period = max(period_numbers, default=5)
    if 1 in explicit_sessions:
        sessions = 2
        periods = max(1, min(8, maximum_period))
    elif maximum_period > 8:
        sessions = 2
        periods = max(1, min(8, (maximum_period + 1) // 2))
    else:
        sessions = 1
        periods = max(1, min(8, maximum_period))
    return {
        "id": 0,
        "name": "TKB import",
        "school_name": "Truong hoc",
        "days": days,
        "sessions": sessions,
        "periods": periods,
        "blocked_slots": [],
    }


def _standalone_short_name(value: str, fallback: str) -> str:
    words = [word for word in re.split(r"\s+", _cell_text(value)) if word]
    if not words:
        return fallback[:20]
    if len(words) == 1:
        return words[0][:20].upper()
    return "".join(word[0] for word in words if word)[:20].upper() or fallback[:20]


def analyze_standalone_schedule_file(
    *,
    filename: str,
    content: bytes,
    include_editable: bool = False,
) -> dict[str, Any]:
    file_format, tables = read_tables(filename, content)
    raw_lessons, parse_warnings, detection = parse_tables_standalone(tables)
    if not raw_lessons:
        raise ScheduleAuditParseError(
            "Khong tim thay tiet hoc nao co the doc duoc. File can co thong tin Thu/Tiet va lop, mon hoac giao vien tuong ung."
        )
    project = infer_standalone_project(tables, raw_lessons)

    issues: list[dict[str, Any]] = []
    normalized_rows: list[dict[str, Any]] = []
    for raw in raw_lessons:
        slot = _resolve_slot(raw.day_text, raw.session_text, raw.period_text, project)
        if slot is None:
            issues.append(_issue(
                "invalid_slot", "error", "Khong xac dinh duoc tiet hoc",
                f"Khong doi duoc '{raw.day_text} / {raw.session_text or '-'} / {raw.period_text}' thanh mot o hop le.",
                source=raw.source,
            ))
            continue
        class_name = _standalone_class_token(raw.class_text) or _standalone_clean_entity_heading(raw.class_text)
        subject_name = _cell_text(raw.subject_text) or _standalone_subject_from_text(raw.lesson_text)
        teacher_name = _cell_text(raw.teacher_text)
        if not class_name:
            issues.append(_issue(
                "unknown_class", "error", "Khong nhan dien duoc lop",
                f"Khong tim thay ten lop trong o '{raw.lesson_text or raw.class_text}'.",
                slot=slot, project=project, source=raw.source,
            ))
            continue
        if not subject_name:
            _class, subject_name, inferred_teacher = _standalone_parse_cell(
                raw.lesson_text, fixed_class=class_name, fixed_teacher=teacher_name,
            )
            teacher_name = teacher_name or inferred_teacher
        if not subject_name:
            subject_name = f"Mon chua xac dinh - {class_name}"
            issues.append(_issue(
                "unknown_subject", "warning", "Chua xac dinh duoc mon hoc",
                f"{class_name} tai {slot_label(slot, project)} khong co ten mon ro rang; he thong tao mon tam de ban co the chinh sua sau khi import.",
                slot=slot, project=project, source=raw.source, entity=class_name,
            ))
        if not teacher_name:
            teacher_name = f"GV chua xac dinh - {class_name} - {subject_name}"
            issues.append(_issue(
                "unknown_teacher", "warning", "Chua xac dinh duoc giao vien",
                f"{class_name} · {subject_name} khong co ten giao vien ro rang; he thong tao giao vien tam de khong lam mat tiet khi convert.",
                slot=slot, project=project, source=raw.source, entity=class_name,
            ))
        normalized_rows.append({
            "slot": slot,
            "class_name": class_name,
            "subject_name": subject_name,
            "teacher_name": teacher_name,
            "room": raw.room_text.strip(),
            "source": raw.source,
            "origin": raw.origin,
            "raw_text": _cell_text(raw.lesson_text) or " · ".join(part for part in (subject_name, teacher_name) if part),
        })

    # Gop cung mot tiet neu cung du lieu xuat hien o hai goc nhin lop/giao vien.
    deduplicated: list[dict[str, Any]] = []
    first_by_key: dict[tuple[int, str, str, str], dict[str, Any]] = {}
    for entry in normalized_rows:
        key = (
            int(entry["slot"]), normalize_text(entry["class_name"]),
            normalize_text(entry["subject_name"]), normalize_text(entry["teacher_name"]),
        )
        existing = first_by_key.get(key)
        if existing is not None and entry["origin"] != existing["origin"]:
            if entry["source"] and entry["source"] not in existing["source"]:
                existing["source"] += f"; {entry['source']}"
            if not existing["room"] and entry["room"]:
                existing["room"] = entry["room"]
            continue
        deduplicated.append(entry)
        if existing is None:
            first_by_key[key] = entry
    normalized_rows = deduplicated

    class_names = sorted({row["class_name"] for row in normalized_rows}, key=lambda value: normalize_text(value))
    subject_names = sorted({row["subject_name"] for row in normalized_rows}, key=lambda value: normalize_text(value))
    teacher_names = sorted({row["teacher_name"] for row in normalized_rows}, key=lambda value: normalize_text(value))
    class_ids = {normalize_text(name): index for index, name in enumerate(class_names, start=1)}
    subject_ids = {normalize_text(name): index for index, name in enumerate(subject_names, start=1)}
    teacher_ids = {normalize_text(name): index for index, name in enumerate(teacher_names, start=1)}

    classes = [{"id": item_id, "name": name, "grade_id": None, "unavailable": []} for name, item_id in ((name, class_ids[normalize_text(name)]) for name in class_names)]
    subjects = [{
        "id": subject_ids[normalize_text(name)], "name": name,
        "short_name": _standalone_short_name(name, f"M{subject_ids[normalize_text(name)]}"),
        "max_consecutive": int(project["periods"]),
    } for name in subject_names]
    teachers = [{
        "id": teacher_ids[normalize_text(name)], "name": name,
        "short_name": _standalone_short_name(name, f"GV{teacher_ids[normalize_text(name)]}"),
        "department_id": None, "max_periods_day": int(project["sessions"]) * int(project["periods"]),
        "unavailable": [], "subject_ids": [],
    } for name in teacher_names]

    assignment_keys: list[tuple[int, int, int]] = []
    assignment_counts: dict[tuple[int, int, int], int] = defaultdict(int)
    for row in normalized_rows:
        key = (
            class_ids[normalize_text(row["class_name"])],
            subject_ids[normalize_text(row["subject_name"])],
            teacher_ids[normalize_text(row["teacher_name"])],
        )
        if key not in assignment_counts:
            assignment_keys.append(key)
        assignment_counts[key] += 1
    assignments: list[dict[str, Any]] = []
    assignment_id_by_key: dict[tuple[int, int, int], int] = {}
    class_by_id = {row["id"]: row for row in classes}
    subject_by_id = {row["id"]: row for row in subjects}
    teacher_by_id = {row["id"]: row for row in teachers}
    for assignment_id, key in enumerate(assignment_keys, start=1):
        class_id, subject_id, teacher_id = key
        assignment_id_by_key[key] = assignment_id
        assignments.append({
            "id": assignment_id,
            "class_id": class_id,
            "subject_id": subject_id,
            "teacher_id": teacher_id,
            "periods_per_week": assignment_counts[key],
            "block_mode": "free",
            "class_name": class_by_id[class_id]["name"],
            "subject_name": subject_by_id[subject_id]["name"],
            "subject_short": subject_by_id[subject_id]["short_name"],
            "teacher_name": teacher_by_id[teacher_id]["name"],
            "teacher_short": teacher_by_id[teacher_id]["short_name"],
        })
        teacher_by_id[teacher_id]["subject_ids"].append(subject_id)

    recognized: list[dict[str, Any]] = []
    for index, row in enumerate(normalized_rows, start=1):
        key = (
            class_ids[normalize_text(row["class_name"])],
            subject_ids[normalize_text(row["subject_name"])],
            teacher_ids[normalize_text(row["teacher_name"])],
        )
        recognized.append({
            "draft_id": index,
            "assignment_id": assignment_id_by_key[key],
            "slot": int(row["slot"]),
            "room": row["room"],
            "source": row["source"],
            "raw_text": row.get("raw_text", ""),
            "class_id": key[0], "subject_id": key[1], "teacher_id": key[2],
        })

    conflict_codes_by_draft: dict[int, set[str]] = defaultdict(set)
    conflict_details_by_draft: dict[int, list[str]] = defaultdict(list)
    by_class_slot: dict[tuple[int, int], list[dict[str, Any]]] = defaultdict(list)
    by_teacher_slot: dict[tuple[int, int], list[dict[str, Any]]] = defaultdict(list)
    by_room_slot: dict[tuple[int, str], list[dict[str, Any]]] = defaultdict(list)
    for entry in recognized:
        by_class_slot[(entry["slot"], entry["class_id"])].append(entry)
        by_teacher_slot[(entry["slot"], entry["teacher_id"])].append(entry)
        room_key = normalize_text(entry["room"])
        if room_key:
            by_room_slot[(entry["slot"], room_key)].append(entry)
    for (slot, class_id), rows in by_class_slot.items():
        if len(rows) > 1:
            detail = f"Lớp {class_by_id[class_id]['name']} có {len(rows)} tiết cùng lúc."
            for row in rows:
                conflict_codes_by_draft[int(row["draft_id"])].add("class_collision")
                conflict_details_by_draft[int(row["draft_id"])].append(detail)
            issues.append(_issue(
                "class_collision", "error", "Trung lich lop",
                detail,
                slot=slot, project=project, source="; ".join(row["source"] for row in rows), entity=class_by_id[class_id]["name"],
            ))
    for (slot, teacher_id), rows in by_teacher_slot.items():
        if len(rows) > 1:
            class_list = ", ".join(class_by_id[row["class_id"]]["name"] for row in rows)
            detail = f"Giáo viên {teacher_by_id[teacher_id]['name']} bị xếp đồng thời: {class_list}."
            for row in rows:
                conflict_codes_by_draft[int(row["draft_id"])].add("teacher_collision")
                conflict_details_by_draft[int(row["draft_id"])].append(detail)
            issues.append(_issue(
                "teacher_collision", "error", "Trung lich giao vien",
                detail,
                slot=slot, project=project, source="; ".join(row["source"] for row in rows), entity=teacher_by_id[teacher_id]["name"],
            ))
    for (slot, _room), rows in by_room_slot.items():
        if len(rows) > 1:
            room = rows[0]["room"]
            detail = f"Phòng {room} đang được dùng cho {len(rows)} lớp cùng lúc."
            for row in rows:
                conflict_codes_by_draft[int(row["draft_id"])].add("room_collision")
                conflict_details_by_draft[int(row["draft_id"])].append(detail)
            issues.append(_issue(
                "room_collision", "error", "Trung phong hoc",
                detail,
                slot=slot, project=project, source="; ".join(row["source"] for row in rows), entity=room,
            ))
    for warning in parse_warnings:
        issues.append(_issue("unread_table", "warning", "Co bang/sheet chua doc duoc", warning))

    severity_order = {"error": 0, "warning": 1, "info": 2}
    issues.sort(key=lambda item: (severity_order.get(item["severity"], 9), item.get("slot") is None, item.get("slot") or -1, item["title"]))
    errors = sum(1 for item in issues if item["severity"] == "error")
    warnings = sum(1 for item in issues if item["severity"] == "warning")
    collisions = sum(1 for item in issues if item["code"] in {"teacher_collision", "class_collision", "room_collision"})
    detection.update({
        "scope_label": f"Doc lap · {len(classes)} lop · {len(teachers)} giao vien · {len(subjects)} mon",
        "classes": [row["name"] for row in classes],
        "teachers": [row["name"] for row in teachers],
    })

    viewer_cells: list[dict[str, Any]] = []
    affected_coordinates: set[tuple[int, int]] = set()
    for entry in recognized:
        class_id = int(entry["class_id"])
        subject_id = int(entry["subject_id"])
        teacher_id = int(entry["teacher_id"])
        codes = sorted(conflict_codes_by_draft.get(int(entry["draft_id"]), set()))
        if codes:
            affected_coordinates.add((int(entry["slot"]), class_id))
        viewer_cells.append({
            "draft_id": int(entry["draft_id"]),
            "slot": int(entry["slot"]),
            "class_id": class_id,
            "class_name": class_by_id[class_id]["name"],
            "subject_name": subject_by_id[subject_id]["name"],
            "teacher_name": teacher_by_id[teacher_id]["name"],
            "room": entry.get("room", ""),
            "source": entry.get("source", ""),
            "raw_text": entry.get("raw_text", ""),
            "conflicts": codes,
            "conflict_details": conflict_details_by_draft.get(int(entry["draft_id"]), []),
        })
    viewer = {
        "days": int(project["days"]),
        "sessions": int(project["sessions"]),
        "periods": int(project["periods"]),
        "classes": [{"id": row["id"], "name": row["name"]} for row in classes],
        "cells": viewer_cells,
        "conflict_cells": len(affected_coordinates),
    }
    result = {
        "ok": True,
        "filename": filename,
        "format": file_format,
        "detection": detection,
        "status": "clean" if errors == 0 and warnings == 0 else ("error" if errors else "warning"),
        "summary": {
            "read_lessons": len(raw_lessons),
            "recognized_lessons": len(recognized),
            "errors": errors,
            "warnings": warnings,
            "collisions": collisions,
            "missing_periods": 0,
            "extra_periods": 0,
            "classes": len(classes),
            "teachers": len(teachers),
            "subjects": len(subjects),
        },
        "issues": issues,
        "viewer": viewer,
        "data": {
            "project": project,
            "classes": classes,
            "subjects": subjects,
            "teachers": teachers,
            "assignments": assignments,
            "lessons": [],
        },
    }
    if include_editable:
        result["editable_lessons"] = [
            {"draft_id": row["draft_id"], "assignment_id": row["assignment_id"], "slot": row["slot"]}
            for row in recognized
        ]
    return result
