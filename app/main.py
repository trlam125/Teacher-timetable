from __future__ import annotations

import io
import json
import random
import secrets
import smtplib
from collections import Counter, defaultdict
from datetime import datetime, timedelta, timezone
from email.message import EmailMessage
from typing import Optional
from urllib.parse import quote

from fastapi import Depends, FastAPI, Form, HTTPException, Request
from fastapi.responses import HTMLResponse, JSONResponse, RedirectResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from itsdangerous import BadSignature, SignatureExpired, URLSafeTimedSerializer
import hashlib
import hmac
import os
from dotenv import load_dotenv
from pydantic import BaseModel, Field
from sqlalchemy import Boolean, ForeignKey, Integer, String, Text, UniqueConstraint, create_engine, func, inspect, select
from sqlalchemy.orm import DeclarativeBase, Mapped, Session, mapped_column, relationship, sessionmaker

load_dotenv()

DATABASE_URL = os.getenv("DATABASE_URL")
SECRET_KEY = os.getenv("SECRET_KEY")

if not DATABASE_URL:
    raise RuntimeError(
        "Thiếu DATABASE_URL. Hãy tạo file .env dựa trên .env.example "
        "và nhập chuỗi kết nối PostgreSQL."
    )
if not DATABASE_URL.startswith(("postgresql://", "postgresql+psycopg://")):
    raise RuntimeError("Project này chỉ hỗ trợ PostgreSQL.")
if not SECRET_KEY:
    raise RuntimeError(
        "Thiếu SECRET_KEY. Hãy tạo khóa bí mật và thêm vào file .env."
    )

# Chuẩn hóa về driver psycopg 3.
if DATABASE_URL.startswith("postgresql://"):
    DATABASE_URL = DATABASE_URL.replace(
        "postgresql://", "postgresql+psycopg://", 1
    )

engine = create_engine(
    DATABASE_URL,
    pool_pre_ping=True,
)
SessionLocal = sessionmaker(bind=engine, expire_on_commit=False)

class Base(DeclarativeBase):
    pass

class User(Base):
    __tablename__ = "users"
    id: Mapped[int] = mapped_column(primary_key=True)
    email: Mapped[str] = mapped_column(String(255), unique=True, index=True)
    password_hash: Mapped[str] = mapped_column(String(255))
    name: Mapped[str] = mapped_column(String(120), default="Giáo viên")
    role: Mapped[str] = mapped_column(String(20), default="pending")
    requested_teacher_name: Mapped[Optional[str]] = mapped_column(String(120), nullable=True)
    requested_project_id: Mapped[Optional[int]] = mapped_column(Integer, nullable=True, index=True)
    teacher_id: Mapped[Optional[int]] = mapped_column(Integer, nullable=True, index=True)
    reset_token_hash: Mapped[Optional[str]] = mapped_column(String(64), nullable=True)
    reset_token_expires_at: Mapped[Optional[str]] = mapped_column(String(40), nullable=True)
    is_superadmin: Mapped[bool] = mapped_column(Boolean, default=False)
    session_version: Mapped[int] = mapped_column(Integer, default=1)

class Project(Base):
    __tablename__ = "projects"
    id: Mapped[int] = mapped_column(primary_key=True)
    owner_id: Mapped[int] = mapped_column(ForeignKey("users.id"), index=True)
    name: Mapped[str] = mapped_column(String(200))
    school_name: Mapped[str] = mapped_column(String(200), default="Trường học")
    days: Mapped[int] = mapped_column(Integer, default=6)
    sessions: Mapped[int] = mapped_column(Integer, default=2)
    periods_per_session: Mapped[int] = mapped_column(Integer, default=5)
    blocked_slots_json: Mapped[str] = mapped_column(Text, default="[]")
    share_token: Mapped[str] = mapped_column(String(64), unique=True, default=lambda: secrets.token_urlsafe(16))
    created_at: Mapped[str] = mapped_column(String(40), default=lambda: datetime.now().isoformat(timespec="seconds"))

class Department(Base):
    __tablename__ = "departments"
    id: Mapped[int] = mapped_column(primary_key=True)
    project_id: Mapped[int] = mapped_column(ForeignKey("projects.id"), index=True)
    name: Mapped[str] = mapped_column(String(120))

class Subject(Base):
    __tablename__ = "subjects"
    id: Mapped[int] = mapped_column(primary_key=True)
    project_id: Mapped[int] = mapped_column(ForeignKey("projects.id"), index=True)
    name: Mapped[str] = mapped_column(String(120))
    short_name: Mapped[str] = mapped_column(String(20))
    max_consecutive: Mapped[int] = mapped_column(Integer, default=2)

class Teacher(Base):
    __tablename__ = "teachers"
    id: Mapped[int] = mapped_column(primary_key=True)
    project_id: Mapped[int] = mapped_column(ForeignKey("projects.id"), index=True)
    department_id: Mapped[Optional[int]] = mapped_column(ForeignKey("departments.id"), nullable=True)
    name: Mapped[str] = mapped_column(String(120))
    short_name: Mapped[str] = mapped_column(String(30))
    max_periods_day: Mapped[int] = mapped_column(Integer, default=5)
    unavailable_json: Mapped[str] = mapped_column(Text, default="[]")

class TeacherAccountLink(Base):
    __tablename__ = "teacher_account_links"
    id: Mapped[int] = mapped_column(primary_key=True)
    user_id: Mapped[int] = mapped_column(ForeignKey("users.id"), index=True)
    teacher_id: Mapped[int] = mapped_column(ForeignKey("teachers.id"), unique=True, index=True)
    __table_args__ = (UniqueConstraint("user_id", "teacher_id", name="uq_teacher_account_link"),)

class Grade(Base):
    __tablename__ = "grades"
    id: Mapped[int] = mapped_column(primary_key=True)
    project_id: Mapped[int] = mapped_column(ForeignKey("projects.id"), index=True)
    name: Mapped[str] = mapped_column(String(80))

class SchoolClass(Base):
    __tablename__ = "classes"
    id: Mapped[int] = mapped_column(primary_key=True)
    project_id: Mapped[int] = mapped_column(ForeignKey("projects.id"), index=True)
    grade_id: Mapped[Optional[int]] = mapped_column(ForeignKey("grades.id"), nullable=True)
    name: Mapped[str] = mapped_column(String(80))
    unavailable_json: Mapped[str] = mapped_column(Text, default="[]")

class Assignment(Base):
    __tablename__ = "assignments"
    id: Mapped[int] = mapped_column(primary_key=True)
    project_id: Mapped[int] = mapped_column(ForeignKey("projects.id"), index=True)
    class_id: Mapped[int] = mapped_column(ForeignKey("classes.id"))
    subject_id: Mapped[int] = mapped_column(ForeignKey("subjects.id"))
    teacher_id: Mapped[int] = mapped_column(ForeignKey("teachers.id"))
    periods_per_week: Mapped[int] = mapped_column(Integer, default=1)
    block_mode: Mapped[str] = mapped_column(String(24), default="free")
    # Giữ cột cũ để migration các project đã tồn tại; giao diện mới không dùng mẫu chuỗi.
    consecutive_pattern: Mapped[str] = mapped_column(String(80), default="")

class FixedLesson(Base):
    __tablename__ = "fixed_lessons"
    id: Mapped[int] = mapped_column(primary_key=True)
    project_id: Mapped[int] = mapped_column(ForeignKey("projects.id"), index=True)
    assignment_id: Mapped[int] = mapped_column(ForeignKey("assignments.id"))
    slot: Mapped[int] = mapped_column(Integer)
    group_size: Mapped[int] = mapped_column(Integer, default=1)

class Lesson(Base):
    __tablename__ = "lessons"
    id: Mapped[int] = mapped_column(primary_key=True)
    project_id: Mapped[int] = mapped_column(ForeignKey("projects.id"), index=True)
    assignment_id: Mapped[int] = mapped_column(ForeignKey("assignments.id"))
    slot: Mapped[int] = mapped_column(Integer)
    locked: Mapped[bool] = mapped_column(Boolean, default=False)
    __table_args__ = (UniqueConstraint("project_id", "assignment_id", "slot", name="uq_lesson"),)

class TeacherPreference(Base):
    __tablename__ = "teacher_preferences"
    id: Mapped[int] = mapped_column(primary_key=True)
    project_id: Mapped[int] = mapped_column(ForeignKey("projects.id"), index=True)
    teacher_id: Mapped[int] = mapped_column(ForeignKey("teachers.id"), index=True)
    preferred_json: Mapped[str] = mapped_column(Text, default="[]")
    unavailable_json: Mapped[str] = mapped_column(Text, default="[]")
    note: Mapped[str] = mapped_column(Text, default="")
    status: Mapped[str] = mapped_column(String(24), default="pending", index=True)
    created_at: Mapped[str] = mapped_column(String(40), default=lambda: datetime.now().isoformat(timespec="seconds"))
    reviewed_at: Mapped[Optional[str]] = mapped_column(String(40), nullable=True)

Base.metadata.create_all(engine)

def migrate_schema():
    """Bổ sung các cột cũ còn thiếu bằng câu lệnh tương thích PostgreSQL."""
    inspector = inspect(engine)
    if "users" not in inspector.get_table_names():
        return

    columns = {column["name"] for column in inspector.get_columns("users")}
    role_was_added = "role" not in columns
    with engine.begin() as connection:
        if role_was_added:
            connection.exec_driver_sql(
                "ALTER TABLE users "
                "ADD COLUMN role VARCHAR(20) NOT NULL DEFAULT 'pending'"
            )
        if "teacher_id" not in columns:
            connection.exec_driver_sql(
                "ALTER TABLE users ADD COLUMN teacher_id INTEGER"
            )
        if "requested_teacher_name" not in columns:
            connection.exec_driver_sql(
                "ALTER TABLE users ADD COLUMN requested_teacher_name VARCHAR(120)"
            )
        if "requested_project_id" not in columns:
            connection.exec_driver_sql(
                "ALTER TABLE users ADD COLUMN requested_project_id INTEGER"
            )
            connection.exec_driver_sql(
                "CREATE INDEX IF NOT EXISTS ix_users_requested_project_id ON users (requested_project_id)"
            )
        if "reset_token_hash" not in columns:
            connection.exec_driver_sql(
                "ALTER TABLE users ADD COLUMN reset_token_hash VARCHAR(64)"
            )
        if "reset_token_expires_at" not in columns:
            connection.exec_driver_sql(
                "ALTER TABLE users ADD COLUMN reset_token_expires_at VARCHAR(40)"
            )
        if "is_superadmin" not in columns:
            connection.exec_driver_sql(
                "ALTER TABLE users ADD COLUMN is_superadmin BOOLEAN NOT NULL DEFAULT FALSE"
            )
        if "session_version" not in columns:
            connection.exec_driver_sql(
                "ALTER TABLE users ADD COLUMN session_version INTEGER NOT NULL DEFAULT 1"
            )
        connection.exec_driver_sql(
            "UPDATE users SET role='pending' WHERE role IS NULL OR role='' OR role='user'"
        )
        # Luôn khôi phục vai trò từ các quan hệ dữ liệu cũ. Một số database cũ
        # đã có cột role nhưng vẫn lưu giá trị legacy "user"; chỉ chạy phần
        # này khi vừa thêm cột role sẽ khiến toàn bộ tài khoản bị kẹt ở pending.
        connection.exec_driver_sql(
            "UPDATE users SET role='teacher' "
            "WHERE teacher_id IS NOT NULL AND role='pending'"
        )
        if "projects" in inspector.get_table_names():
            # Chủ project luôn là quản trị viên. Chạy sau bước teacher để tài
            # khoản vừa là giáo viên cũ vừa sở hữu project vẫn nhận quyền admin.
            connection.exec_driver_sql(
                "UPDATE users SET role='admin' "
                "WHERE id IN (SELECT DISTINCT owner_id FROM projects)"
            )
        if "teacher_account_links" in inspector.get_table_names():
            connection.exec_driver_sql(
                "INSERT INTO teacher_account_links (user_id, teacher_id) "
                "SELECT id, teacher_id FROM users "
                "WHERE role='teacher' AND teacher_id IS NOT NULL "
                "ON CONFLICT (teacher_id) DO NOTHING"
            )
        connection.exec_driver_sql(
            "ALTER TABLE users ALTER COLUMN role SET DEFAULT 'pending'"
        )
        bootstrap_email = os.getenv("BOOTSTRAP_ADMIN_EMAIL", "").strip().lower()
        if bootstrap_email:
            connection.exec_driver_sql(
                "UPDATE users SET is_superadmin=TRUE WHERE lower(email)=lower(%s) AND role='admin' "
                "AND NOT EXISTS (SELECT 1 FROM users WHERE is_superadmin=TRUE)",
                (bootstrap_email,),
            )
        connection.exec_driver_sql(
            "UPDATE users SET is_superadmin=TRUE WHERE id=("
            "SELECT id FROM users WHERE role='admin' ORDER BY id ASC LIMIT 1"
            ") AND NOT EXISTS (SELECT 1 FROM users WHERE is_superadmin=TRUE)"
        )

        assignment_columns = {column["name"] for column in inspector.get_columns("assignments")} if "assignments" in inspector.get_table_names() else set()
        if assignment_columns and "block_mode" not in assignment_columns:
            connection.exec_driver_sql(
                "ALTER TABLE assignments ADD COLUMN block_mode VARCHAR(24) NOT NULL DEFAULT 'free'"
            )
            # Mẫu cũ chỉ gồm 1 và 2, đồng thời có ít nhất một số 2, được
            # chuyển sang bắt buộc tiết đôi. Mẫu có cụm 3+ chuyển thành tự do
            # vì chế độ mới không còn hỗ trợ cụm tùy ý.
            connection.exec_driver_sql(
                "UPDATE assignments SET block_mode='required_double' "
                "WHERE consecutive_pattern ~ '(^|,)[[:space:]]*2[[:space:]]*(,|$)' "
                "AND COALESCE(regexp_replace(consecutive_pattern, '[[:space:]12,]', '', 'g'), '') = ''"
            )
            connection.exec_driver_sql(
                "UPDATE assignments SET consecutive_pattern=''"
            )

        fixed_columns = {column["name"] for column in inspector.get_columns("fixed_lessons")} if "fixed_lessons" in inspector.get_table_names() else set()
        if fixed_columns and "group_size" not in fixed_columns:
            connection.exec_driver_sql(
                "ALTER TABLE fixed_lessons ADD COLUMN group_size INTEGER NOT NULL DEFAULT 1"
            )
        if fixed_columns and "assignments" in inspector.get_table_names():
            # Ở chế độ tự do/ưu tiên, mỗi ghim là một tiết độc lập. Chuyển các
            # ghim cụm cũ thành từng ghim đơn để không tạo tiết khóa mồ côi.
            connection.exec_driver_sql(
                "UPDATE fixed_lessons SET group_size=1 FROM assignments "
                "WHERE fixed_lessons.assignment_id=assignments.id "
                "AND assignments.block_mode<>'required_double'"
            )
            if "lessons" in inspector.get_table_names():
                connection.exec_driver_sql(
                    "INSERT INTO fixed_lessons (project_id, assignment_id, slot, group_size) "
                    "SELECT lessons.project_id, lessons.assignment_id, lessons.slot, 1 "
                    "FROM lessons JOIN assignments ON assignments.id=lessons.assignment_id "
                    "WHERE lessons.locked=TRUE AND assignments.block_mode<>'required_double' "
                    "AND NOT EXISTS (SELECT 1 FROM fixed_lessons "
                    "WHERE fixed_lessons.assignment_id=lessons.assignment_id "
                    "AND fixed_lessons.slot=lessons.slot)"
                )

        # Bản demo cũ từng được tạo chỉ với một buổi. Khôi phục cả sáng và chiều.
        if "projects" in inspector.get_table_names():
            project_columns = {column["name"] for column in inspector.get_columns("projects")}
            if "blocked_slots_json" not in project_columns:
                connection.exec_driver_sql(
                    "ALTER TABLE projects ADD COLUMN blocked_slots_json TEXT NOT NULL DEFAULT '[]'"
                )
            connection.exec_driver_sql(
                "UPDATE projects SET sessions=2 "
                "WHERE name='TKB học kỳ I' AND school_name='THPT Demo' AND sessions=1"
            )

migrate_schema()
class Passwords:
    @staticmethod
    def hash(password: str) -> str:
        salt = secrets.token_hex(16)
        digest = hashlib.pbkdf2_hmac("sha256", password.encode(), salt.encode(), 200_000).hex()
        return f"pbkdf2_sha256${salt}${digest}"
    @staticmethod
    def verify(password: str, encoded: str) -> bool:
        try:
            algo, salt, digest = encoded.split("$", 2)
            if algo != "pbkdf2_sha256": return False
            actual = hashlib.pbkdf2_hmac("sha256", password.encode(), salt.encode(), 200_000).hex()
            return hmac.compare_digest(actual, digest)
        except Exception:
            return False
pwd = Passwords()
signer = URLSafeTimedSerializer(SECRET_KEY, salt="session")
reset_signer = URLSafeTimedSerializer(SECRET_KEY, salt="password-reset")
captcha_signer = URLSafeTimedSerializer(SECRET_KEY, salt="forgot-password-captcha")

app = FastAPI(title="Teacher Timetable")
app.mount("/static", StaticFiles(directory="app/static"), name="static")
templates = Jinja2Templates(directory="app/templates")

DAYS = ["Thứ 2", "Thứ 3", "Thứ 4", "Thứ 5", "Thứ 6", "Thứ 7", "Chủ nhật"]
RESET_TOKEN_TTL_SECONDS = 30 * 60
SESSION_TTL_SECONDS = max(300, int(os.getenv("SESSION_TTL_SECONDS", str(12 * 60 * 60))))
APP_ENV = os.getenv("APP_ENV", "production").strip().lower()
BOOTSTRAP_ADMIN_EMAIL = os.getenv("BOOTSTRAP_ADMIN_EMAIL", "").strip().lower()
BOOTSTRAP_ADMIN_PASSWORD = os.getenv("BOOTSTRAP_ADMIN_PASSWORD", "")
SEED_DEMO_DATA = os.getenv("SEED_DEMO_DATA", "false").strip().lower() in {"1", "true", "yes"}

def new_captcha() -> tuple[str, str]:
    left = secrets.randbelow(8) + 2
    right = secrets.randbelow(8) + 2
    token = captcha_signer.dumps({"answer": left + right})
    return f"{left} + {right} = ?", token

def captcha_is_valid(token: str, answer: str) -> bool:
    try:
        data = captcha_signer.loads(token, max_age=10 * 60)
        return hmac.compare_digest(str(data["answer"]), answer.strip())
    except (BadSignature, SignatureExpired, KeyError, ValueError):
        return False

def send_password_reset_email(recipient: str, reset_url: str) -> bool:
    smtp_host = os.getenv("SMTP_HOST")
    if not smtp_host:
        return False
    smtp_port = int(os.getenv("SMTP_PORT", "587"))
    smtp_user = os.getenv("SMTP_USER", "")
    smtp_password = os.getenv("SMTP_PASSWORD", "")
    smtp_from = os.getenv("SMTP_FROM", smtp_user or "no-reply@smart-tkb.local")
    use_ssl = os.getenv("SMTP_SSL", "false").lower() in {"1", "true", "yes"}

    message = EmailMessage()
    message["Subject"] = "Đặt lại mật khẩu Smart TKB"
    message["From"] = smtp_from
    message["To"] = recipient
    message.set_content(
        "Bạn vừa yêu cầu đặt lại mật khẩu Smart TKB.\n\n"
        f"Mở liên kết sau trong vòng 30 phút:\n{reset_url}\n\n"
        "Nếu bạn không yêu cầu, hãy bỏ qua email này."
    )

    smtp_class = smtplib.SMTP_SSL if use_ssl else smtplib.SMTP
    with smtp_class(smtp_host, smtp_port, timeout=15) as client:
        if not use_ssl and os.getenv("SMTP_STARTTLS", "true").lower() in {"1", "true", "yes"}:
            client.starttls()
        if smtp_user:
            client.login(smtp_user, smtp_password)
        client.send_message(message)
    return True

def db_session():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()

def set_session_cookie(response, user: User):
    response.set_cookie(
        "session",
        signer.dumps({"uid": user.id, "sv": user.session_version}),
        max_age=SESSION_TTL_SECONDS,
        httponly=True,
        samesite="lax",
    )

def current_user(request: Request, db: Session = Depends(db_session)) -> User:
    raw = request.cookies.get("session")
    if not raw:
        raise HTTPException(401)
    try:
        data = signer.loads(raw, max_age=SESSION_TTL_SECONDS)
        user = db.get(User, int(data["uid"]))
        if not user or int(data.get("sv", -1)) != user.session_version:
            raise HTTPException(401)
        return user
    except (BadSignature, SignatureExpired, KeyError, TypeError, ValueError):
        raise HTTPException(401)

def get_project(pid: int, user: User, db: Session) -> Project:
    if user.role != "admin":
        raise HTTPException(403,"Chỉ quản trị viên được thực hiện thao tác này")
    p = db.get(Project, pid)
    if not p or p.owner_id != user.id:
        raise HTTPException(404)
    return p

def get_project_for_update(pid: int, user: User, db: Session) -> Project:
    """Khóa project đến hết transaction để tuần tự hóa mọi thay đổi lịch."""
    if user.role != "admin":
        raise HTTPException(403,"Chỉ quản trị viên được thực hiện thao tác này")
    project = db.scalar(
        select(Project)
        .where(Project.id == pid, Project.owner_id == user.id)
        .with_for_update()
    )
    if not project:
        raise HTTPException(404)
    return project

def admin_project_ids(user: User, db: Session) -> set[int]:
    if user.role != "admin":
        return set()
    return set(db.scalars(select(Project.id).where(Project.owner_id == user.id)).all())

def admin_teacher_ids(user: User, db: Session) -> set[int]:
    project_ids = admin_project_ids(user, db)
    if not project_ids:
        return set()
    return set(db.scalars(select(Teacher.id).where(Teacher.project_id.in_(project_ids))).all())

def account_teacher_ids(account: User, db: Session) -> set[int]:
    ids = set(db.scalars(
        select(TeacherAccountLink.teacher_id).where(TeacherAccountLink.user_id == account.id)
    ).all())
    if account.teacher_id is not None:
        ids.add(account.teacher_id)
    return ids

def account_for_teacher(teacher_id: int, db: Session) -> User | None:
    link = db.scalar(select(TeacherAccountLink).where(TeacherAccountLink.teacher_id == teacher_id))
    if link:
        return db.get(User, link.user_id)
    return db.scalar(select(User).where(User.role == "teacher", User.teacher_id == teacher_id))

def ensure_teacher_link(account: User, teacher: Teacher, db: Session) -> None:
    existing = db.scalar(select(TeacherAccountLink).where(TeacherAccountLink.teacher_id == teacher.id))
    if existing and existing.user_id != account.id:
        raise HTTPException(409, "Giáo viên này đã có tài khoản")
    if not existing:
        db.add(TeacherAccountLink(user_id=account.id, teacher_id=teacher.id))
    if account.teacher_id is None:
        account.teacher_id = teacher.id

def is_bootstrap_admin(user: User, db: Session) -> bool:
    return user.role == "admin" and bool(user.is_superadmin)

def admin_can_manage_account(admin: User, account: User, db: Session) -> bool:
    if admin.role != "admin":
        return False
    if account.id == admin.id:
        return True
    if account.role == "admin":
        return False
    project_ids = admin_project_ids(admin, db)
    if account.role == "teacher":
        teacher_ids = account_teacher_ids(account, db)
        if not teacher_ids:
            return False
        return db.scalar(select(Teacher.id).where(
            Teacher.id.in_(teacher_ids), Teacher.project_id.in_(project_ids)
        )) is not None
    if account.role == "pending":
        if account.requested_project_id in project_ids:
            return True
        return account.requested_project_id is None and is_bootstrap_admin(admin, db)
    return False

def development_reset_links_enabled(request: Request) -> bool:
    host = (request.url.hostname or "").lower()
    return APP_ENV == "development" and host in {"localhost", "127.0.0.1", "::1"}

def slot_meta(project: Project, slot: int):
    ppd = project.sessions * project.periods_per_session
    day = slot // ppd
    inside = slot % ppd
    session = inside // project.periods_per_session
    period = inside % project.periods_per_session
    return day, session, period

def all_slots(project: Project):
    return list(range(project.days * project.sessions * project.periods_per_session))

def parse_slots(text: str):
    try:
        return set(int(x) for x in json.loads(text or "[]"))
    except Exception:
        return set()

BLOCK_MODES = {"free", "preferred_double", "required_double"}

def consecutive_groups(pattern:str,total_periods:int):
    """Đọc mẫu cụm cũ, chỉ dùng cho migration/tương thích dữ liệu cũ."""
    text=(pattern or "").strip()
    if not text: return [1]*total_periods
    try:
        groups=[int(value.strip()) for value in text.split(",") if value.strip()]
    except ValueError as exc:
        raise ValueError("Mẫu tiết liên tiếp cũ không hợp lệ.") from exc
    if not groups or any(value<1 for value in groups) or sum(groups)!=total_periods:
        raise ValueError("Mẫu tiết liên tiếp cũ không hợp lệ.")
    return groups

def normalized_block_mode(value:str,total_periods:int,subject:Subject,project:Project):
    mode=(value or "free").strip().lower()
    aliases={
        "prefer_double":"preferred_double",
        "preferred":"preferred_double",
        "required":"required_double",
        "double":"required_double",
    }
    mode=aliases.get(mode,mode)
    if mode not in BLOCK_MODES:
        raise ValueError("Chế độ xếp tiết không hợp lệ.")
    total_slots=project.days*project.sessions*project.periods_per_session
    if total_periods>total_slots:
        raise ValueError(f"Số tiết/tuần không được vượt quá {total_slots} ô của thời khóa biểu.")
    if mode in {"preferred_double","required_double"} and total_periods>=2:
        if subject.max_consecutive<2:
            raise ValueError(
                f"Môn {subject.name} đang giới hạn tối đa {subject.max_consecutive} tiết liên tiếp; "
                "hãy tăng giới hạn lên ít nhất 2 trước khi chọn chế độ tiết đôi."
            )
        if project.periods_per_session<2:
            raise ValueError("Mỗi buổi phải có ít nhất 2 tiết để dùng chế độ tiết đôi.")
    if mode=="required_double":
        groups=[2]*(total_periods//2)+([1] if total_periods%2 else [])
        if not timetable_pattern_feasible(project,groups):
            raise ValueError(
                "Số cặp tiết bắt buộc không thể phân bố trong số ngày, buổi và tiết hiện có."
            )
    return mode

def assignment_requires_double(assignment:Assignment):
    return getattr(assignment,"block_mode","free")=="required_double"

def assignment_prefers_double(assignment:Assignment):
    return getattr(assignment,"block_mode","free")=="preferred_double"

def assignment_groups(assignment:Assignment):
    total=max(0,int(assignment.periods_per_week or 0))
    if assignment_requires_double(assignment):
        return [2]*(total//2)+([1] if total%2 else [])
    return [1]*total

def assignment_generated_pattern(assignment:Assignment):
    return ",".join(str(value) for value in assignment_groups(assignment)) if assignment_requires_double(assignment) else ""

def assignment_mode_label(assignment:Assignment):
    labels={
        "free":"Tự do",
        "preferred_double":"Ưu tiên tiết đôi",
        "required_double":"Bắt buộc tiết đôi",
    }
    return labels.get(getattr(assignment,"block_mode","free"),"Tự do")

def valid_slots(project: Project, slots: list[int] | set[int]):
    maximum = project.days * project.sessions * project.periods_per_session
    return sorted({int(slot) for slot in slots if 0 <= int(slot) < maximum})

def bounded_int(value,default:int,minimum:int,maximum:int,label:str):
    raw=default if value in (None,"") else value
    if isinstance(raw,bool) or isinstance(raw,float) and not raw.is_integer():
        raise HTTPException(400,f"{label} phải là số nguyên từ {minimum} đến {maximum}")
    try:
        parsed=int(raw)
    except (TypeError,ValueError) as exc:
        raise HTTPException(400,f"{label} phải là số nguyên từ {minimum} đến {maximum}") from exc
    if not minimum<=parsed<=maximum:
        raise HTTPException(400,f"{label} phải nằm trong khoảng từ {minimum} đến {maximum}")
    return parsed

def pattern_slots_match(project:Project,pattern:str,total_periods:int,slots:list[int] | set[int]):
    """Kiểm tra các cụm tiết thực tế có đúng mẫu đã khai báo hay không."""
    try:
        expected=sorted(consecutive_groups(pattern,total_periods))
    except ValueError:
        return False
    if len(slots)!=total_periods:
        return False
    ppd=project.sessions*project.periods_per_session
    groups=defaultdict(list)
    for slot in sorted(set(slots)):
        day=slot//ppd
        inside=slot%ppd
        session=inside//project.periods_per_session
        period=inside%project.periods_per_session
        groups[(day,session)].append(period)
    actual=[]
    for periods in groups.values():
        run=1
        for left,right in zip(periods,periods[1:]):
            if right==left+1:
                run+=1
            else:
                actual.append(run);run=1
        actual.append(run)
    return sorted(actual)==expected

def assignment_run_groups(project: Project, slots: list[int] | set[int]):
    """Trả về các cụm liên tiếp theo đúng ranh giới ngày và buổi."""
    grouped = defaultdict(list)
    for slot in sorted(set(slots)):
        day, session, period = slot_meta(project, slot)
        grouped[(day, session)].append((period, slot))
    runs = []
    for values in grouped.values():
        current = [values[0]] if values else []
        for item in values[1:]:
            if item[0] == current[-1][0] + 1:
                current.append(item)
            else:
                runs.append({"start": current[0][1], "size": len(current), "slots": [x[1] for x in current]})
                current = [item]
        if current:
            runs.append({"start": current[0][1], "size": len(current), "slots": [x[1] for x in current]})
    return sorted(runs, key=lambda item: item["start"])

def _pack_pattern_groups_into_segments(group_sizes:list[int],segments:list[tuple[int,int]]):
    """Xếp các cụm chưa neo vào những đoạn trống, cách nhau ít nhất một tiết."""
    items=sorted((int(size) for size in group_sizes),reverse=True)
    if not items:
        return []
    usable=[(start,length) for start,length in segments if length>0]
    capacities=[length+1 for _start,length in usable]
    if sum(size+1 for size in items)>sum(capacities):
        return None
    allocations=[[] for _segment in usable]
    failed=set()

    def search(index:int):
        if index==len(items):
            return True
        state=(index,tuple(sorted(capacities,reverse=True)))
        if state in failed:
            return False
        size=items[index]
        weight=size+1
        seen_capacities=set()
        for segment_index,capacity in enumerate(capacities):
            if capacity<weight or capacity in seen_capacities:
                continue
            seen_capacities.add(capacity)
            capacities[segment_index]-=weight
            allocations[segment_index].append(size)
            if search(index+1):
                return True
            allocations[segment_index].pop()
            capacities[segment_index]+=weight
        failed.add(state)
        return False

    if not search(0):
        return None
    placements=[]
    for (segment_start,_length),sizes in zip(usable,allocations):
        cursor=segment_start
        for size in sizes:
            placements.append((size,cursor))
            cursor+=size+1
    return placements

def _complete_pattern_placement(
    project:Project,
    expected:list[int],
    slots:list[int] | set[int],
    forced_placements:list[tuple[int,int]] | None=None,
):
    """Tìm một cách đặt đầy đủ các cụm, đồng thời chứa chính xác các tiết đã có."""
    values=list(slots)
    current=set(values)
    maximum=project.days*project.sessions*project.periods_per_session
    if len(values)!=len(current) or any(slot<0 or slot>=maximum for slot in current):
        return None
    if len(current)>sum(expected) or any(size<1 or size>project.periods_per_session for size in expected):
        return None

    ppd=project.sessions*project.periods_per_session
    periods_per_session=project.periods_per_session
    starts_by_size={}
    for size in set(expected):
        starts=[]
        for day in range(project.days):
            for session in range(project.sessions):
                base=day*ppd+session*periods_per_session
                starts.extend(base+period for period in range(periods_per_session-size+1))
        starts_by_size[size]=starts

    def interval(start:int,size:int):
        return set(range(start,start+size))

    def compatible(left_start:int,left_size:int,right_start:int,right_size:int):
        left_slots=interval(left_start,left_size)
        right_slots=interval(right_start,right_size)
        if left_slots.intersection(right_slots):
            return False
        left_day,left_session,_=slot_meta(project,left_start)
        right_day,right_session,_=slot_meta(project,right_start)
        if (left_day,left_session)!=(right_day,right_session):
            return True
        return left_start+left_size!=right_start and right_start+right_size!=left_start

    def free_segments(selected:list[tuple[int,int,set[int]]]):
        forbidden=defaultdict(set)
        for size,start,_covered in selected:
            day,session,period=slot_meta(project,start)
            key=(day,session)
            forbidden[key].update(range(period,period+size))
            if period>0:
                forbidden[key].add(period-1)
            if period+size<periods_per_session:
                forbidden[key].add(period+size)
        result=[]
        for day in range(project.days):
            for session in range(project.sessions):
                blocked=forbidden[(day,session)]
                base=day*ppd+session*periods_per_session
                start=None
                for period in range(periods_per_session+1):
                    is_free=period<periods_per_session and period not in blocked
                    if is_free and start is None:
                        start=period
                    elif not is_free and start is not None:
                        result.append((base+start,period-start))
                        start=None
        return result

    remaining=Counter(expected)
    selected=[]
    forced_covered=set()
    for size,start in forced_placements or []:
        if remaining[size]<=0 or start not in starts_by_size.get(size,[]):
            return None
        covered=current.intersection(interval(start,size))
        if not covered or forced_covered.intersection(covered):
            return None
        if any(not compatible(start,size,other_start,other_size) for other_size,other_start,_ in selected):
            return None
        remaining[size]-=1
        selected.append((size,start,set(covered)))
        forced_covered.update(covered)
    failed=set()

    def search(uncovered:frozenset[int]):
        state=(
            tuple(sorted(uncovered)),
            tuple(sorted(remaining.items())),
            tuple(sorted((size,start) for size,start,_covered in selected)),
        )
        if state in failed:
            return None
        if not uncovered:
            rest=[]
            for size,count in remaining.items():
                rest.extend([size]*count)
            packed=_pack_pattern_groups_into_segments(rest,free_segments(selected))
            if packed is None:
                failed.add(state)
                return None
            return [*selected,*[(size,start,set()) for size,start in packed]]

        target=min(uncovered)
        for size in sorted((value for value,count in remaining.items() if count>0),reverse=True):
            for start in starts_by_size[size]:
                target_slots=interval(start,size)
                covered=current.intersection(target_slots)
                if target not in covered or not covered.issubset(uncovered):
                    continue
                if any(not compatible(start,size,other_start,other_size) for other_size,other_start,_ in selected):
                    continue
                remaining[size]-=1
                selected.append((size,start,set(covered)))
                result=search(frozenset(set(uncovered)-covered))
                if result is not None:
                    return result
                selected.pop()
                remaining[size]+=1
        failed.add(state)
        return None

    return search(frozenset(current-forced_covered))

def timetable_pattern_feasible(project:Project,groups:list[int]):
    return _complete_pattern_placement(project,groups,set()) is not None

def pattern_completion_plan(project: Project, assignment: Assignment, slots: list[int] | set[int]):
    """Lập kế hoạch hoàn thành mẫu tiết từ phần lịch hiện có.

    Các tiết đã đặt có thể là một đoạn liền hoặc nhiều mảnh của cùng một cụm
    (ví dụ đã có tiết 1 và 3 của cụm 3 tiết). Hàm tìm cách bao phủ toàn bộ các
    tiết hiện có bằng những cụm hợp lệ, rồi trả về phần còn thiếu của mỗi cụm.
    """
    current=set(slots)
    maximum=project.days*project.sessions*project.periods_per_session
    if len(current)!=len(list(slots)) or len(current)>assignment.periods_per_week:
        return None
    if any(slot<0 or slot>=maximum for slot in current):
        return None
    if not assignment_requires_double(assignment):
        return [{
            "size":1,
            "anchor_slots":tuple(),
            "candidate_starts":None,
        } for _ in range(assignment.periods_per_week-len(current))]
    expected=assignment_groups(assignment)
    placements=_complete_pattern_placement(project,expected,current)
    if placements is None:
        return None
    ppd=project.sessions*project.periods_per_session
    anchored=[(size,start,set(covered)) for size,start,covered in placements if covered]
    plan = []
    for placement_index,(target_size,start,covered) in enumerate(placements):
        if not covered:
            plan.append({
                "size": target_size,
                "anchor_slots": tuple(),
                "candidate_starts": None,
            })
            continue
        if len(covered)==target_size:
            continue
        target_anchored_index=sum(1 for _size,_start,item_covered in placements[:placement_index] if item_covered)
        alternative_starts=[]
        for day in range(project.days):
            for session in range(project.sessions):
                base=day*ppd+session*project.periods_per_session
                for period in range(project.periods_per_session-target_size+1):
                    candidate=base+period
                    candidate_slots=set(range(candidate,candidate+target_size))
                    if current.intersection(candidate_slots)!=covered:
                        continue
                    forced=[]
                    anchored_index=0
                    for size,chosen_start,_chosen_covered in anchored:
                        forced.append((size,candidate if anchored_index==target_anchored_index else chosen_start))
                        anchored_index+=1
                    if _complete_pattern_placement(project,expected,slots,forced) is not None:
                        alternative_starts.append(candidate)
        plan.append({
            "size": target_size,
            "anchor_slots": tuple(sorted(covered)),
            "candidate_starts": tuple(sorted(set(alternative_starts or [start]))),
        })
    return plan


def remaining_pattern_groups(project: Project, assignment: Assignment, slots: list[int] | set[int]):
    """Trả về kích thước đầy đủ của các cụm còn phải hoàn thành."""
    plan = pattern_completion_plan(project, assignment, slots)
    if plan is None:
        return None
    return [item["size"] for item in plan]

def assignment_pattern_matches(project:Project,assignment:Assignment,slots:list[int] | set[int]):
    values=list(slots)
    if len(values)!=len(set(values)) or len(values)!=assignment.periods_per_week:
        return False
    if not assignment_requires_double(assignment):
        return True
    return pattern_slots_match(
        project,
        assignment_generated_pattern(assignment),
        assignment.periods_per_week,
        values,
    )

def partial_assignment_pattern_matches(project:Project,assignment:Assignment,slots:list[int] | set[int]):
    return pattern_completion_plan(project,assignment,slots) is not None


def assignment_completion_feasible(
    db: Session,
    project: Project,
    assignment: Assignment,
    proposed_slots: list[int] | set[int],
) -> bool:
    """Kiểm tra phần lịch thủ công có ít nhất một cách hoàn thành hợp lệ.

    Khác với ``partial_assignment_pattern_matches``, hàm này xét cả các ràng
    buộc thực tế: ô khóa, thời gian tránh, trùng lớp/giáo viên, số tiết tối đa
    trong ngày, giới hạn tiết liên tiếp và các cụm cố định.
    """
    values=list(proposed_slots)
    current=set(values)
    maximum=project.days*project.sessions*project.periods_per_session
    if len(values)!=len(current) or len(current)>assignment.periods_per_week:
        return False
    if any(slot<0 or slot>=maximum for slot in current):
        return False
    expected=assignment_groups(assignment)
    keep_groups_separate=assignment_requires_double(assignment)

    teacher=db.get(Teacher,assignment.teacher_id)
    school_class=db.get(SchoolClass,assignment.class_id)
    subject=db.get(Subject,assignment.subject_id)
    if not teacher or not school_class or not subject:
        return False

    ppd=project.sessions*project.periods_per_session
    global_blocked=parse_slots(project.blocked_slots_json)
    teacher_unavailable=parse_slots(teacher.unavailable_json)
    class_unavailable=parse_slots(school_class.unavailable_json)
    accepted_preferred,accepted_unavailable=accepted_teacher_preferences(db,project.id)
    preferred=accepted_preferred.get(teacher.id,set())
    teacher_unavailable.update(accepted_unavailable.get(teacher.id,set()))

    other_lessons=[]
    for lesson in db.scalars(select(Lesson).where(Lesson.project_id==project.id)).all():
        if lesson.assignment_id!=assignment.id:
            other_lessons.append(lesson)

    teacher_busy=set()
    class_busy=set()
    teacher_day=Counter()
    base_subject_periods=defaultdict(set)
    for lesson in other_lessons:
        other=db.get(Assignment,lesson.assignment_id)
        if not other:
            continue
        day=lesson.slot//ppd
        if other.teacher_id==assignment.teacher_id:
            teacher_busy.add(lesson.slot)
            teacher_day[day]+=1
        if other.class_id==assignment.class_id:
            class_busy.add(lesson.slot)
            if other.subject_id==assignment.subject_id:
                inside=lesson.slot%ppd
                session=inside//project.periods_per_session
                period=inside%project.periods_per_session
                base_subject_periods[(day,session)].add(period)

    forbidden=global_blocked|teacher_unavailable|class_unavailable|teacher_busy|class_busy
    if current.intersection(forbidden):
        return False

    starts_by_size={}
    for size in set(expected):
        candidates=[]
        for day in range(project.days):
            for session in range(project.sessions):
                base=day*ppd+session*project.periods_per_session
                for period in range(project.periods_per_session-size+1):
                    start=base+period
                    group=tuple(range(start,start+size))
                    if not set(group).intersection(forbidden):
                        candidates.append((start,group))
        starts_by_size[size]=candidates

    def compatible(left_start:int,left_size:int,right_start:int,right_size:int):
        left_end=left_start+left_size
        right_end=right_start+right_size
        if left_start<right_end and right_start<left_end:
            return False
        left_day,left_session,_=slot_meta(project,left_start)
        right_day,right_session,_=slot_meta(project,right_start)
        if (left_day,left_session)!=(right_day,right_session):
            return True
        if not keep_groups_separate:
            return True
        return left_end!=right_start and right_end!=left_start

    remaining=Counter(expected)
    selected=[]
    selected_slots=set()
    added_teacher_day=Counter()
    added_subject_periods=defaultdict(set)

    def subject_limit_ok(group:tuple[int,...]):
        touched=set()
        for slot in group:
            day=slot//ppd
            inside=slot%ppd
            session=inside//project.periods_per_session
            period=inside%project.periods_per_session
            touched.add((day,session))
            added_subject_periods[(day,session)].add(period)
        valid=True
        for key in touched:
            periods=sorted(base_subject_periods[key]|added_subject_periods[key])
            longest=run=0
            previous=None
            for period in periods:
                run=run+1 if previous is not None and period==previous+1 else 1
                longest=max(longest,run)
                previous=period
            if longest>subject.max_consecutive:
                valid=False
                break
        for slot in group:
            day=slot//ppd
            inside=slot%ppd
            session=inside//project.periods_per_session
            period=inside%project.periods_per_session
            added_subject_periods[(day,session)].discard(period)
        return valid

    same_assignment_lessons=db.scalars(select(Lesson).where(
        Lesson.project_id==project.id,
        Lesson.assignment_id==assignment.id,
    )).all()
    fixed_rows=db.scalars(select(FixedLesson).where(
        FixedLesson.project_id==project.id,
        FixedLesson.assignment_id==assignment.id,
    )).all()

    fixed_placements=[]
    for row in fixed_rows:
        size=fixed_row_size(project,assignment,row,same_assignment_lessons)
        if remaining[size]<=0:
            return False
        group=tuple(range(row.slot,row.slot+size))
        if (row.slot,group) not in starts_by_size.get(size,[]):
            return False
        if any(not compatible(row.slot,size,start,other_size) for other_size,start,_group in fixed_placements):
            return False
        fixed_placements.append((size,row.slot,group))
        remaining[size]-=1

    def add_group(size:int,start:int,group:tuple[int,...]):
        day_counts=Counter(slot//ppd for slot in group)
        for day,count in day_counts.items():
            if teacher_day[day]+added_teacher_day[day]+count>teacher.max_periods_day:
                return False
        if not subject_limit_ok(group):
            return False
        for day,count in day_counts.items():
            added_teacher_day[day]+=count
        for slot in group:
            day=slot//ppd
            inside=slot%ppd
            session=inside//project.periods_per_session
            period=inside%project.periods_per_session
            added_subject_periods[(day,session)].add(period)
        selected.append((size,start,group))
        selected_slots.update(group)
        return True

    def remove_group(size:int,start:int,group:tuple[int,...]):
        selected.pop()
        for slot in group:
            selected_slots.remove(slot)
            day=slot//ppd
            added_teacher_day[day]-=1
            inside=slot%ppd
            session=inside//project.periods_per_session
            period=inside%project.periods_per_session
            added_subject_periods[(day,session)].discard(period)

    for size,start,group in fixed_placements:
        if set(group).intersection(selected_slots) or not add_group(size,start,group):
            return False

    failed=set()
    def search():
        uncovered=current-selected_slots
        if sum(remaining.values())==0:
            return not uncovered and len(selected_slots)==assignment.periods_per_week
        state=(
            tuple(sorted(remaining.items())),
            tuple(sorted(uncovered)),
            tuple(sorted(selected_slots)),
        )
        if state in failed:
            return False

        candidate_sets=[]
        if uncovered:
            target=min(uncovered)
            for size,count in remaining.items():
                if count<=0:
                    continue
                options=[item for item in starts_by_size[size] if target in item[1]]
                if options:
                    candidate_sets.append((len(options),-size,size,options))
        else:
            for size,count in remaining.items():
                if count<=0:
                    continue
                options=starts_by_size[size]
                candidate_sets.append((len(options),-size,size,options))
        if not candidate_sets:
            failed.add(state)
            return False

        for _count,_neg,size,raw_options in sorted(candidate_sets):
            options=sorted(raw_options,key=lambda item:(
                -sum(slot in preferred for slot in item[1]),
                item[0],
            ))
            for start,group in options:
                group_set=set(group)
                if group_set.intersection(selected_slots):
                    continue
                if any(not compatible(start,size,other_start,other_size) for other_size,other_start,_other_group in selected):
                    continue
                if uncovered and not group_set.intersection(uncovered):
                    continue
                if not add_group(size,start,group):
                    continue
                remaining[size]-=1
                if search():
                    return True
                remaining[size]+=1
                remove_group(size,start,group)
        failed.add(state)
        return False

    return search()


def assignment_pattern_label(assignment:Assignment):
    if assignment_requires_double(assignment):
        return "bắt buộc tiết đôi ("+" + ".join(str(value) for value in assignment_groups(assignment))+")"
    if assignment_prefers_double(assignment):
        return "ưu tiên tiết đôi"
    return "xếp tiết tự do"

def accepted_teacher_preferences(db: Session, project_id: int):
    rows = db.scalars(
        select(TeacherPreference).where(
            TeacherPreference.project_id == project_id,
            TeacherPreference.status == "accepted",
        )
    ).all()
    preferred = defaultdict(set)
    unavailable = defaultdict(set)
    for row in rows:
        preferred[row.teacher_id].update(parse_slots(row.preferred_json))
        unavailable[row.teacher_id].update(parse_slots(row.unavailable_json))
    return preferred, unavailable


def bounded_text(value, label: str, max_length: int, *, required: bool = True) -> str:
    cleaned = str(value or "").strip()
    if required and not cleaned:
        raise HTTPException(400, f"{label} không được để trống")
    if len(cleaned) > max_length:
        raise HTTPException(400, f"{label} không được vượt quá {max_length} ký tự")
    return cleaned

@app.exception_handler(401)
async def auth_error(request: Request, exc):
    if request.url.path.startswith("/api/"):
        return JSONResponse({"ok": False, "message": "Phiên đăng nhập đã hết hạn."}, status_code=401)
    return RedirectResponse("/login", 303)

@app.get("/", response_class=HTMLResponse)
def home(request: Request, db: Session = Depends(db_session)):
    raw = request.cookies.get("session")
    if raw:
        try:
            data=signer.loads(raw,max_age=SESSION_TTL_SECONDS); user=db.get(User,int(data["uid"]))
            if user and int(data.get("sv",-1))==user.session_version:
                destination = "/teacher" if user.role == "teacher" else ("/projects" if user.role == "admin" else "/account-pending")
                return RedirectResponse(destination, 303)
        except (BadSignature,SignatureExpired,KeyError,TypeError,ValueError):
            pass
    return templates.TemplateResponse("landing.html", {"request": request})

@app.get("/login", response_class=HTMLResponse)
def login_page(request: Request):
    return templates.TemplateResponse("auth.html", {"request": request, "mode": "login", "error": None})

@app.post("/login")
def login(request: Request, email: str = Form(...), password: str = Form(...), db: Session = Depends(db_session)):
    user = db.scalar(select(User).where(User.email == email.lower().strip()))
    if not user or not pwd.verify(password, user.password_hash):
        return templates.TemplateResponse("auth.html", {"request": request, "mode": "login", "error": "Email hoặc mật khẩu không đúng"}, status_code=400)
    destination = "/teacher" if user.role == "teacher" else ("/projects" if user.role == "admin" else "/account-pending")
    res = RedirectResponse(destination, 303)
    set_session_cookie(res, user)
    return res

@app.get("/register", response_class=HTMLResponse)
def register_page(request: Request, project: str = "", db: Session = Depends(db_session)):
    target_project = db.scalar(select(Project).where(Project.share_token == project)) if project else None
    return templates.TemplateResponse("auth.html", {
        "request": request, "mode": "register", "error": None,
        "target_project": target_project, "project_token": project if target_project else "",
    })

@app.post("/register")
def register(
    request: Request,
    name: str = Form(...),
    email: str = Form(...),
    teacher_name: str = Form(...),
    password: str = Form(...),
    project_token: str = Form(""),
    db: Session = Depends(db_session),
):
    name = name.strip()
    email = email.lower().strip()
    teacher_name = teacher_name.strip()
    target_project = db.scalar(select(Project).where(Project.share_token == project_token)) if project_token else None
    context = {
        "request": request, "mode": "register", "error": None,
        "target_project": target_project, "project_token": project_token if target_project else "",
    }
    if project_token and not target_project:
        context["error"] = "Liên kết đăng ký không hợp lệ hoặc đã được thay đổi."
        return templates.TemplateResponse("auth.html", context, status_code=400)
    if not name:
        context["error"] = "Họ tên tài khoản không được để trống"
        return templates.TemplateResponse("auth.html", context, status_code=400)
    if not teacher_name:
        context["error"] = "Tên giáo viên mong muốn không được để trống"
        return templates.TemplateResponse("auth.html", context, status_code=400)
    if len(name) > 120:
        context["error"] = "Họ tên tài khoản không được vượt quá 120 ký tự"
        return templates.TemplateResponse("auth.html", context, status_code=400)
    if len(teacher_name) > 120:
        context["error"] = "Tên giáo viên mong muốn không được vượt quá 120 ký tự"
        return templates.TemplateResponse("auth.html", context, status_code=400)
    if len(email) > 255:
        context["error"] = "Email không được vượt quá 255 ký tự"
        return templates.TemplateResponse("auth.html", context, status_code=400)
    if len(password) < 6:
        context["error"] = "Mật khẩu phải có ít nhất 6 ký tự"
        return templates.TemplateResponse("auth.html", context, status_code=400)
    if not email or "@" not in email:
        context["error"] = "Email không hợp lệ"
        return templates.TemplateResponse("auth.html", context, status_code=400)
    if db.scalar(select(User).where(User.email == email)):
        context["error"] = "Email đã tồn tại"
        return templates.TemplateResponse("auth.html", context, status_code=400)
    user = User(
        name=name,
        email=email,
        requested_teacher_name=teacher_name,
        requested_project_id=target_project.id if target_project else None,
        password_hash=pwd.hash(password),
        role="pending",
    )
    db.add(user); db.commit()
    res = RedirectResponse("/account-pending", 303)
    set_session_cookie(res, user)
    return res

@app.get("/forgot-password", response_class=HTMLResponse)
def forgot_password_page(request: Request):
    question, captcha_token = new_captcha()
    return templates.TemplateResponse("forgot_password.html", {
        "request": request,
        "question": question,
        "captcha_token": captcha_token,
        "error": None,
        "submitted": False,
        "dev_reset_link": None,
    })

@app.post("/forgot-password", response_class=HTMLResponse)
def forgot_password(
    request: Request,
    email: str = Form(...),
    not_robot: Optional[str] = Form(None),
    captcha_answer: str = Form(...),
    captcha_token: str = Form(...),
    db: Session = Depends(db_session),
):
    if not_robot != "yes" or not captcha_is_valid(captcha_token, captcha_answer):
        question, fresh_token = new_captcha()
        return templates.TemplateResponse("forgot_password.html", {
            "request": request,
            "question": question,
            "captcha_token": fresh_token,
            "error": "Xác minh Tôi không phải robot chưa đúng. Vui lòng thử lại.",
            "submitted": False,
            "dev_reset_link": None,
        }, status_code=400)

    account = db.scalar(select(User).where(User.email == email.lower().strip()))
    dev_reset_link = None
    allow_local_link = development_reset_links_enabled(request)
    smtp_configured = bool(os.getenv("SMTP_HOST"))
    if account and (smtp_configured or allow_local_link):
        nonce = secrets.token_urlsafe(32)
        account.reset_token_hash = hashlib.sha256(nonce.encode()).hexdigest()
        account.reset_token_expires_at = (
            datetime.now(timezone.utc) + timedelta(seconds=RESET_TOKEN_TTL_SECONDS)
        ).isoformat()
        db.commit()
        token = reset_signer.dumps({"uid": account.id, "nonce": nonce})
        reset_url = f"{str(request.base_url).rstrip('/')}/reset-password/{token}"
        email_sent = False
        if smtp_configured:
            try:
                email_sent = send_password_reset_email(account.email, reset_url)
            except (OSError, smtplib.SMTPException):
                email_sent = False
        if allow_local_link and not email_sent:
            dev_reset_link = reset_url

    question, fresh_token = new_captcha()
    return templates.TemplateResponse("forgot_password.html", {
        "request": request,
        "question": question,
        "captcha_token": fresh_token,
        "error": None,
        "submitted": True,
        "dev_reset_link": dev_reset_link,
    })

def reset_account_for_token(token: str, db: Session) -> Optional[User]:
    try:
        data = reset_signer.loads(token, max_age=RESET_TOKEN_TTL_SECONDS)
        account = db.get(User, int(data["uid"]))
        nonce_hash = hashlib.sha256(str(data["nonce"]).encode()).hexdigest()
        if not account or not account.reset_token_hash:
            return None
        if not hmac.compare_digest(account.reset_token_hash, nonce_hash):
            return None
        expires_at = datetime.fromisoformat(account.reset_token_expires_at or "")
        if expires_at < datetime.now(timezone.utc):
            return None
        return account
    except (BadSignature, SignatureExpired, KeyError, ValueError, TypeError):
        return None

@app.get("/reset-password/{token}", response_class=HTMLResponse)
def reset_password_page(token: str, request: Request, db: Session = Depends(db_session)):
    account = reset_account_for_token(token, db)
    return templates.TemplateResponse("reset_password.html", {
        "request": request,
        "token": token,
        "valid": account is not None,
        "error": None,
        "success": False,
    }, status_code=200 if account else 400)

@app.post("/reset-password/{token}", response_class=HTMLResponse)
def reset_password(
    token: str,
    request: Request,
    password: str = Form(...),
    password_confirm: str = Form(...),
    db: Session = Depends(db_session),
):
    account = reset_account_for_token(token, db)
    error = None
    if not account:
        error = "Liên kết đặt lại mật khẩu không hợp lệ, đã hết hạn hoặc đã được sử dụng."
    elif len(password) < 6:
        error = "Mật khẩu mới phải có ít nhất 6 ký tự."
    elif password != password_confirm:
        error = "Hai lần nhập mật khẩu không khớp."
    if error:
        return templates.TemplateResponse("reset_password.html", {
            "request": request,
            "token": token,
            "valid": account is not None,
            "error": error,
            "success": False,
        }, status_code=400)

    account.password_hash = pwd.hash(password)
    account.session_version += 1
    account.reset_token_hash = None
    account.reset_token_expires_at = None
    db.commit()
    return templates.TemplateResponse("reset_password.html", {
        "request": request,
        "token": token,
        "valid": False,
        "error": None,
        "success": True,
    })

@app.get("/logout")
def logout():
    res = RedirectResponse("/", 303); res.delete_cookie("session"); return res

@app.get("/account-pending", response_class=HTMLResponse)
def account_pending(request: Request, user: User = Depends(current_user)):
    if user.role == "admin":
        return RedirectResponse("/projects", 303)
    if user.role == "teacher":
        return RedirectResponse("/teacher", 303)
    return templates.TemplateResponse("user_pending.html", {"request": request, "user": user})

@app.get("/admin/users", response_class=HTMLResponse)
def admin_users(request: Request, user: User = Depends(current_user), db: Session = Depends(db_session)):
    if user.role != "admin":
        raise HTTPException(403, "Chỉ quản trị viên được quản lý tài khoản")
    projects = db.scalars(select(Project).where(Project.owner_id == user.id)).all()
    project_names = {project.id: project.name for project in projects}
    project_ids = set(project_names)
    managed_teacher_ids = admin_teacher_ids(user, db)
    all_users = db.scalars(select(User).order_by(User.id.asc())).all()
    users = [account for account in all_users if admin_can_manage_account(user, account, db)]
    assigned_teacher_ids = set(db.scalars(select(TeacherAccountLink.teacher_id)).all())
    assigned_teacher_ids.update(db.scalars(
        select(User.teacher_id).where(User.role == "teacher", User.teacher_id.is_not(None))
    ).all())
    managed_account_ids = {account.id for account in users if account.role == "teacher" and admin_can_manage_account(user, account, db)}
    available_teachers = []
    if project_ids:
        teachers = db.scalars(
            select(Teacher).where(Teacher.project_id.in_(project_ids)).order_by(Teacher.name.asc())
        ).all()
        available_teachers = [
            {
                "id": teacher.id,
                "name": teacher.name,
                "project_id": teacher.project_id,
                "project_name": project_names[teacher.project_id],
            }
            for teacher in teachers
            if teacher.id not in assigned_teacher_ids
        ]
    return templates.TemplateResponse("users.html", {
        "request": request,
        "user": user,
        "users": users,
        "available_teachers": available_teachers,
        "managed_projects": projects,
        "managed_teacher_ids": managed_teacher_ids,
        "managed_account_ids": managed_account_ids,
        "project_names": project_names,
    })

@app.post("/admin/users/{account_id}/update")
def update_account(
    account_id: int,
    name: str = Form(...),
    email: str = Form(...),
    requested_teacher_name: str = Form(""),
    password: str = Form(""),
    user: User = Depends(current_user),
    db: Session = Depends(db_session),
):
    if user.role != "admin":
        raise HTTPException(403, "Chỉ quản trị viên được quản lý tài khoản")
    account = db.get(User, account_id)
    if not account or not admin_can_manage_account(user, account, db):
        raise HTTPException(404, "Không tìm thấy tài khoản trong phạm vi quản lý")
    name = bounded_text(name, "Họ tên", 120)
    email = bounded_text(email.lower(), "Email", 255)
    requested_teacher_name = bounded_text(
        requested_teacher_name, "Tên giáo viên mong muốn", 120, required=False
    )
    if "@" not in email:
        raise HTTPException(400, "Email không hợp lệ")
    conflict = db.scalar(select(User).where(User.email == email, User.id != account.id))
    if conflict:
        raise HTTPException(409, "Email đã được dùng cho tài khoản khác")
    account.name = name
    account.email = email
    account.requested_teacher_name = requested_teacher_name or None
    password_changed = bool(password.strip())
    if password_changed:
        if len(password.strip()) < 6:
            raise HTTPException(400, "Mật khẩu mới phải có ít nhất 6 ký tự")
        account.password_hash = pwd.hash(password.strip())
        account.session_version += 1
    db.commit()
    response = RedirectResponse("/admin/users", 303)
    if account.id == user.id and password_changed:
        set_session_cookie(response, account)
    return response

@app.post("/admin/users/{account_id}/delete")
def delete_account(
    account_id: int,
    user: User = Depends(current_user),
    db: Session = Depends(db_session),
):
    if user.role != "admin":
        raise HTTPException(403, "Chỉ quản trị viên được quản lý tài khoản")
    account = db.get(User, account_id)
    if not account or not admin_can_manage_account(user, account, db):
        raise HTTPException(404, "Không tìm thấy tài khoản trong phạm vi quản lý")
    if account.id == user.id:
        raise HTTPException(400, "Không thể xóa chính tài khoản đang đăng nhập")
    if db.scalar(select(Project.id).where(Project.owner_id == account.id)) is not None:
        raise HTTPException(400, "Không thể xóa tài khoản đang sở hữu bộ thời khóa biểu")
    for link in db.scalars(select(TeacherAccountLink).where(TeacherAccountLink.user_id == account.id)).all():
        db.delete(link)
    db.flush()
    db.delete(account)
    db.commit()
    return RedirectResponse("/admin/users", 303)

@app.post("/admin/users/{account_id}/approve")
def approve_teacher_account(
    account_id: int,
    teacher_id: str = Form(""),
    project_id: str = Form(""),
    user: User = Depends(current_user),
    db: Session = Depends(db_session),
):
    if user.role != "admin":
        raise HTTPException(403, "Chỉ quản trị viên được quản lý tài khoản")
    account = db.get(User, account_id)
    if not account or not admin_can_manage_account(user, account, db):
        raise HTTPException(404, "Không tìm thấy tài khoản trong phạm vi quản lý")
    if account.role != "pending":
        raise HTTPException(400, "Chỉ có thể duyệt tài khoản đang chờ")
    teacher_id = teacher_id.strip()
    project_id = project_id.strip()
    if teacher_id and not teacher_id.isdigit():
        raise HTTPException(400, "Giáo viên được chọn không hợp lệ")
    if project_id and not project_id.isdigit():
        raise HTTPException(400, "Bộ thời khóa biểu được chọn không hợp lệ")
    selected_teacher_id = int(teacher_id) if teacher_id else None
    selected_project_id = int(project_id) if project_id else None
    if selected_teacher_id is not None:
        teacher = db.get(Teacher, selected_teacher_id)
        project = db.get(Project, teacher.project_id) if teacher else None
        if not teacher or not project or project.owner_id != user.id:
            raise HTTPException(400, "Giáo viên không hợp lệ hoặc không thuộc bộ thời khóa biểu của bạn")
        if account.requested_project_id is not None and project.id != account.requested_project_id:
            raise HTTPException(403, "Tài khoản này được mời vào một bộ thời khóa biểu khác")
        if account_for_teacher(teacher.id, db):
            raise HTTPException(400, "Giáo viên này đã có tài khoản")
    else:
        teacher_name = (account.requested_teacher_name or "").strip()
        if not teacher_name:
            raise HTTPException(400, "Hãy nhập tên giáo viên mong muốn hoặc chọn một giáo viên có sẵn")
        managed_projects = db.scalars(
            select(Project).where(Project.owner_id == user.id).order_by(Project.id.asc())
        ).all()
        if selected_project_id is None and len(managed_projects) == 1:
            project = managed_projects[0]
        else:
            project = db.get(Project, selected_project_id) if selected_project_id is not None else None
        if not project or project.owner_id != user.id:
            raise HTTPException(400, "Hãy chọn bộ thời khóa biểu để tạo hồ sơ giáo viên mới")
        if account.requested_project_id is not None and project.id != account.requested_project_id:
            raise HTTPException(403, "Tài khoản này được mời vào một bộ thời khóa biểu khác")
        short_name = teacher_name.split()[-1][:30]
        teacher = Teacher(
            project_id=project.id,
            name=teacher_name,
            short_name=short_name,
            max_periods_day=5,
        )
        db.add(teacher)
        db.flush()
    account.role = "teacher"
    account.teacher_id = teacher.id
    ensure_teacher_link(account, teacher, db)
    account.requested_project_id = None
    db.commit()
    return RedirectResponse("/admin/users", 303)

@app.post("/admin/users/{account_id}/promote-admin")
def promote_teacher_to_admin(
    account_id: int,
    user: User = Depends(current_user),
    db: Session = Depends(db_session),
):
    if user.role != "admin":
        raise HTTPException(403, "Chỉ quản trị viên được quản lý tài khoản")
    account = db.get(User, account_id)
    if not account:
        raise HTTPException(404, "Không tìm thấy tài khoản")
    if account.role != "teacher":
        raise HTTPException(400, "Chỉ có thể nâng tài khoản giáo viên lên quản trị viên")
    if not admin_can_manage_account(user, account, db):
        raise HTTPException(403, "Bạn không quản lý tài khoản giáo viên này")
    for link in db.scalars(select(TeacherAccountLink).where(TeacherAccountLink.user_id == account.id)).all():
        db.delete(link)
    account.role = "admin"
    account.teacher_id = None
    db.commit()
    return RedirectResponse("/admin/users", 303)

@app.get("/projects", response_class=HTMLResponse)
def projects(request: Request, user: User = Depends(current_user), db: Session = Depends(db_session)):
    if user.role == "teacher":
        return RedirectResponse("/teacher", 303)
    if user.role != "admin":
        return RedirectResponse("/account-pending", 303)
    rows = db.scalars(select(Project).where(Project.owner_id == user.id).order_by(Project.id.desc())).all()
    return templates.TemplateResponse("projects.html", {"request": request, "user": user, "projects": rows})

@app.post("/projects")
def create_project(name: str = Form(...), school_name: str = Form(...), days: int = Form(6), sessions: int = Form(2), periods: int = Form(5), user: User = Depends(current_user), db: Session = Depends(db_session)):
    if user.role!="admin": raise HTTPException(403)
    clean_name=name.strip(); clean_school_name=school_name.strip()
    if not clean_name: raise HTTPException(400,"Tên bộ thời khóa biểu không được để trống")
    if not clean_school_name: raise HTTPException(400,"Tên trường không được để trống")
    if len(clean_name)>200: raise HTTPException(400,"Tên bộ thời khóa biểu không được vượt quá 200 ký tự")
    if len(clean_school_name)>200: raise HTTPException(400,"Tên trường không được vượt quá 200 ký tự")
    p = Project(owner_id=user.id, name=clean_name, school_name=clean_school_name, days=max(1,min(days,7)), sessions=max(1,min(sessions,2)), periods_per_session=max(1,min(periods,8)))
    db.add(p); db.commit()
    return RedirectResponse(f"/projects/{p.id}", 303)

@app.post("/projects/{pid}/clone")
def clone_project(pid: int, user: User = Depends(current_user), db: Session = Depends(db_session)):
    # Dùng cùng khóa với mọi API chỉnh sửa để bản sao luôn được đọc từ một
    # trạng thái nhất quán, không trộn dữ liệu trước và sau một thay đổi đồng thời.
    src = get_project_for_update(pid,user,db)
    suffix = " (bản sao)"
    clone_name = src.name[: 200 - len(suffix)] + suffix
    p = Project(owner_id=user.id,name=clone_name,school_name=src.school_name,days=src.days,sessions=src.sessions,periods_per_session=src.periods_per_session,blocked_slots_json=src.blocked_slots_json)
    db.add(p); db.flush()
    maps = {"dep":{},"sub":{},"tea":{},"grade":{},"cls":{},"ass":{}}
    for x in db.scalars(select(Department).where(Department.project_id==pid)):
        n=Department(project_id=p.id,name=x.name);db.add(n);db.flush();maps["dep"][x.id]=n.id
    for x in db.scalars(select(Subject).where(Subject.project_id==pid)):
        n=Subject(project_id=p.id,name=x.name,short_name=x.short_name,max_consecutive=x.max_consecutive);db.add(n);db.flush();maps["sub"][x.id]=n.id
    for x in db.scalars(select(Teacher).where(Teacher.project_id==pid)):
        n=Teacher(project_id=p.id,department_id=maps["dep"].get(x.department_id),name=x.name,short_name=x.short_name,max_periods_day=x.max_periods_day,unavailable_json=x.unavailable_json);db.add(n);db.flush();maps["tea"][x.id]=n.id
    source_links=db.scalars(select(TeacherAccountLink).where(TeacherAccountLink.teacher_id.in_(list(maps["tea"])))).all() if maps["tea"] else []
    linked_source_ids={link.teacher_id for link in source_links}
    for link in source_links:
        db.add(TeacherAccountLink(user_id=link.user_id,teacher_id=maps["tea"][link.teacher_id]))
    if maps["tea"]:
        legacy_accounts=db.scalars(select(User).where(User.role=="teacher",User.teacher_id.in_(list(maps["tea"])))).all()
        for account in legacy_accounts:
            if account.teacher_id not in linked_source_ids:
                db.add(TeacherAccountLink(user_id=account.id,teacher_id=maps["tea"][account.teacher_id]))
    for x in db.scalars(select(Grade).where(Grade.project_id==pid)):
        n=Grade(project_id=p.id,name=x.name);db.add(n);db.flush();maps["grade"][x.id]=n.id
    for x in db.scalars(select(SchoolClass).where(SchoolClass.project_id==pid)):
        n=SchoolClass(project_id=p.id,grade_id=maps["grade"].get(x.grade_id),name=x.name,unavailable_json=x.unavailable_json);db.add(n);db.flush();maps["cls"][x.id]=n.id
    for x in db.scalars(select(Assignment).where(Assignment.project_id==pid)):
        n=Assignment(project_id=p.id,class_id=maps["cls"][x.class_id],subject_id=maps["sub"][x.subject_id],teacher_id=maps["tea"][x.teacher_id],periods_per_week=x.periods_per_week,block_mode=x.block_mode,consecutive_pattern="");db.add(n);db.flush();maps["ass"][x.id]=n.id
    for x in db.scalars(select(FixedLesson).where(FixedLesson.project_id==pid)):
        if x.assignment_id in maps["ass"]:
            db.add(FixedLesson(project_id=p.id,assignment_id=maps["ass"][x.assignment_id],slot=x.slot,group_size=x.group_size))
    for x in db.scalars(select(Lesson).where(Lesson.project_id==pid)):
        db.add(Lesson(project_id=p.id,assignment_id=maps["ass"][x.assignment_id],slot=x.slot,locked=x.locked))
    db.commit(); return RedirectResponse(f"/projects/{p.id}",303)

@app.get("/projects/{pid}", response_class=HTMLResponse)
def project_page(pid:int, request:Request, user:User=Depends(current_user), db:Session=Depends(db_session)):
    p=get_project(pid,user,db)
    data=project_data(db,p)
    return templates.TemplateResponse("workspace.html", {"request":request,"user":user,"p":p,"data":data,"days":DAYS})

class EntityIn(BaseModel):
    type: str
    data: dict

def required_text(data: dict, key: str, label: str, max_length: int) -> str:
    return bounded_text(data.get(key, ""), label, max_length)

def required_id(data: dict, key: str, label: str) -> int:
    value = data.get(key)
    try:
        parsed = int(value)
    except (TypeError, ValueError) as exc:
        raise HTTPException(400, f"{label} không hợp lệ") from exc
    if parsed <= 0:
        raise HTTPException(400, f"{label} không hợp lệ")
    return parsed

@app.post("/api/projects/{pid}/entity")
def add_entity(pid:int, payload:EntityIn, user:User=Depends(current_user), db:Session=Depends(db_session)):
    project=get_project_for_update(pid,user,db); d=payload.data
    if payload.type=="department":
        obj=Department(project_id=pid,name=required_text(d,"name","Tên tổ chuyên môn",120))
    elif payload.type=="subject":
        name=required_text(d,"name","Tên môn học",120)
        max_consecutive=bounded_int(d.get("max_consecutive"),2,1,4,"Số tiết liên tiếp tối đa")
        short_name=(str(d.get("short_name") or "").strip() or name[:5])[:20]
        obj=Subject(project_id=pid,name=name,short_name=short_name,max_consecutive=max_consecutive)
    elif payload.type=="teacher":
        name=required_text(d,"name","Tên giáo viên",120)
        department_id=d.get("department_id") or None
        if department_id:
            try: department_id=int(department_id)
            except (TypeError,ValueError) as exc: raise HTTPException(400,"Tổ chuyên môn không hợp lệ") from exc
            department=db.get(Department,department_id)
            if not department or department.project_id!=pid: raise HTTPException(400,"Tổ chuyên môn không hợp lệ")
        max_periods_day=bounded_int(d.get("max_periods_day"),5,1,10,"Số tiết tối đa mỗi ngày")
        short_name=(str(d.get("short_name") or "").strip() or name)[:30]
        obj=Teacher(project_id=pid,name=name,short_name=short_name,department_id=department_id,max_periods_day=max_periods_day,unavailable_json=json.dumps(valid_slots(project,d.get("unavailable",[]))))
    elif payload.type=="grade":
        obj=Grade(project_id=pid,name=required_text(d,"name","Tên khối lớp",80))
    elif payload.type=="class":
        name=required_text(d,"name","Tên lớp học",80)
        grade_id=d.get("grade_id") or None
        if grade_id:
            try: grade_id=int(grade_id)
            except (TypeError,ValueError) as exc: raise HTTPException(400,"Khối lớp không hợp lệ") from exc
            grade=db.get(Grade,grade_id)
            if not grade or grade.project_id!=pid: raise HTTPException(400,"Khối lớp không hợp lệ")
        obj=SchoolClass(project_id=pid,name=name,grade_id=grade_id,unavailable_json=json.dumps(valid_slots(project,d.get("unavailable",[]))))
    elif payload.type=="assignment":
        class_id=required_id(d,"class_id","Lớp học")
        subject_id=required_id(d,"subject_id","Môn học")
        teacher_id=required_id(d,"teacher_id","Giáo viên")
        school_class=db.get(SchoolClass,class_id); subject=db.get(Subject,subject_id); teacher=db.get(Teacher,teacher_id)
        if not school_class or not subject or not teacher or any(x.project_id!=pid for x in (school_class,subject,teacher)):
            raise HTTPException(400,"Lớp, môn hoặc giáo viên không thuộc bộ thời khóa biểu")
        duplicate=db.scalar(select(Assignment.id).where(
            Assignment.project_id==pid,
            Assignment.class_id==school_class.id,
            Assignment.subject_id==subject.id,
            Assignment.teacher_id==teacher.id,
        ))
        if duplicate is not None:
            raise HTTPException(409,"Phân công lớp – môn – giáo viên này đã tồn tại")
        periods=bounded_int(d.get("periods_per_week"),1,1,40,"Số tiết mỗi tuần")
        try: mode=normalized_block_mode(d.get("block_mode","free"),periods,subject,project)
        except ValueError as exc: raise HTTPException(400,str(exc)) from exc
        obj=Assignment(project_id=pid,class_id=school_class.id,subject_id=subject.id,teacher_id=teacher.id,periods_per_week=periods,block_mode=mode,consecutive_pattern="")
    else: raise HTTPException(400,"Loại dữ liệu không hợp lệ")
    db.add(obj); db.commit(); return {"ok":True,"id":obj.id}

@app.put("/api/projects/{pid}/entity/{typ}/{eid}")
def update_entity(
    pid: int,
    typ: str,
    eid: int,
    payload: EntityIn,
    user: User = Depends(current_user),
    db: Session = Depends(db_session),
):
    project = get_project_for_update(pid, user, db)
    if payload.type != typ or typ not in {"subject", "teacher", "class"}:
        raise HTTPException(400, "Loại dữ liệu không hợp lệ")
    model = {"subject": Subject, "teacher": Teacher, "class": SchoolClass}[typ]
    obj = db.get(model, eid)
    if not obj or obj.project_id != pid:
        raise HTTPException(404, "Không tìm thấy dữ liệu cần sửa")
    d = payload.data
    name_limit = {"subject": 120, "teacher": 120, "class": 80}[typ]
    name = bounded_text(d.get("name", ""), "Tên", name_limit)
    if typ == "subject":
        short_name = bounded_text(d.get("short_name", ""), "Tên rút gọn", 20)
        new_max_consecutive = bounded_int(
            d.get("max_consecutive"), 1, 1, 4, "Số tiết liên tiếp tối đa"
        )
        assignments = db.scalars(select(Assignment).where(
            Assignment.project_id == pid,
            Assignment.subject_id == obj.id,
        )).all()
        incompatible_assignments = [
            assignment.id for assignment in assignments
            if getattr(assignment,"block_mode","free") in {"preferred_double","required_double"}
            and assignment.periods_per_week >= 2
            and new_max_consecutive < 2
        ]

        assignment_by_id = {assignment.id: assignment for assignment in assignments}
        periods_by_class_session = defaultdict(list)
        periods_per_day = project.sessions * project.periods_per_session
        if assignment_by_id:
            lessons = db.scalars(select(Lesson).where(
                Lesson.project_id == pid,
                Lesson.assignment_id.in_(list(assignment_by_id)),
            )).all()
            for lesson in lessons:
                assignment = assignment_by_id.get(lesson.assignment_id)
                if not assignment:
                    continue
                day = lesson.slot // periods_per_day
                inside_day = lesson.slot % periods_per_day
                session = inside_day // project.periods_per_session
                period = inside_day % project.periods_per_session
                periods_by_class_session[(assignment.class_id, day, session)].append(period)

        violating_class_sessions = 0
        for periods in periods_by_class_session.values():
            longest = run = 0
            previous = None
            for period in sorted(set(periods)):
                run = run + 1 if previous is not None and period == previous + 1 else 1
                longest = max(longest, run)
                previous = period
            if longest > new_max_consecutive:
                violating_class_sessions += 1

        if incompatible_assignments or violating_class_sessions:
            details = []
            if incompatible_assignments:
                details.append(
                    f"{len(incompatible_assignments)} phân công đang dùng chế độ tiết đôi"
                )
            if violating_class_sessions:
                details.append(
                    f"{violating_class_sessions} buổi của lớp đang có cụm môn học dài hơn"
                )
            raise HTTPException(
                409,
                f"Không thể giảm còn {new_max_consecutive} tiết liên tiếp vì "
                + " và ".join(details)
                + ". Hãy điều chỉnh phân công hoặc lịch hiện tại trước.",
            )
        obj.name = name
        obj.short_name = short_name
        obj.max_consecutive = new_max_consecutive
    elif typ == "teacher":
        short_name = bounded_text(d.get("short_name", ""), "Tên ngắn", 30)
        department_id = d.get("department_id") or None
        if department_id is not None:
            try:
                department_id = int(department_id)
            except (TypeError, ValueError) as exc:
                raise HTTPException(400, "Tổ chuyên môn không hợp lệ") from exc
            department = db.get(Department, department_id)
            if not department or department.project_id != pid:
                raise HTTPException(400, "Tổ chuyên môn không hợp lệ")
        new_max_periods_day = bounded_int(
            d.get("max_periods_day"), 5, 1, 10, "Số tiết tối đa mỗi ngày"
        )
        assignment_ids = set(db.scalars(select(Assignment.id).where(
            Assignment.project_id == pid,
            Assignment.teacher_id == obj.id,
        )).all())
        if assignment_ids:
            ppd = project.sessions * project.periods_per_session
            daily_counts = Counter(
                slot // ppd
                for slot in db.scalars(select(Lesson.slot).where(
                    Lesson.project_id == pid,
                    Lesson.assignment_id.in_(assignment_ids),
                )).all()
            )
            highest_current = max(daily_counts.values(), default=0)
            if highest_current > new_max_periods_day:
                raise HTTPException(
                    409,
                    f"Không thể giảm còn {new_max_periods_day} tiết/ngày vì lịch hiện tại có ngày giáo viên đang dạy {highest_current} tiết. Hãy điều chỉnh lịch trước.",
                )
        obj.name = name
        obj.short_name = short_name
        obj.department_id = department_id
        obj.max_periods_day = new_max_periods_day
    else:
        grade_id = d.get("grade_id") or None
        if grade_id is not None:
            try:
                grade_id = int(grade_id)
            except (TypeError, ValueError) as exc:
                raise HTTPException(400, "Khối lớp không hợp lệ") from exc
            grade = db.get(Grade, grade_id)
            if not grade or grade.project_id != pid:
                raise HTTPException(400, "Khối lớp không hợp lệ")
        obj.name = name
        obj.grade_id = grade_id
    db.commit()
    return {"ok": True}

class AssignmentUpdateIn(BaseModel):
    periods_per_week: int
    block_mode: str = "free"


def normalize_assignment_fixed_rows(
    db: Session,
    project: Project,
    assignment: Assignment,
    lessons: list[Lesson],
) -> str | None:
    """Chuẩn hóa các ghim khi số tiết hoặc chế độ cụm thay đổi.

    Lesson.locked là nguồn dữ liệu thật. Ở chế độ tự do/ưu tiên, mỗi tiết khóa
    có một FixedLesson riêng. Ở chế độ bắt buộc tiết đôi, các tiết khóa phải tự
    tạo thành những cụm hoàn chỉnh được chế độ mới cho phép; không tự ý khóa
    thêm một tiết lân cận vì điều đó sẽ thay đổi lựa chọn của người dùng.
    """
    locked_slots=sorted(lesson.slot for lesson in lessons if lesson.locked)
    replacement_rows=[]
    if assignment_requires_double(assignment):
        expected=Counter(assignment_groups(assignment))
        used=Counter()
        for run in assignment_run_groups(project,locked_slots):
            size=run["size"]
            if expected[size]<=used[size]:
                return (
                    "Các tiết cố định hiện tại không tạo thành các cặp/tiết lẻ "
                    "hợp lệ theo chế độ bắt buộc tiết đôi mới. Hãy bỏ cố định "
                    "hoặc sắp xếp lại các tiết cố định trước."
                )
            used[size]+=1
            replacement_rows.append((run["start"],size))
    else:
        replacement_rows=[(slot,1) for slot in locked_slots]

    fixed_rows=db.scalars(select(FixedLesson).where(
        FixedLesson.project_id==project.id,
        FixedLesson.assignment_id==assignment.id,
    )).all()
    for row in fixed_rows:
        db.delete(row)
    for slot,size in replacement_rows:
        db.add(FixedLesson(
            project_id=project.id,
            assignment_id=assignment.id,
            slot=slot,
            group_size=size,
        ))
    return None


@app.put("/api/projects/{pid}/assignments/{assignment_id}")
def update_assignment(pid:int,assignment_id:int,payload:AssignmentUpdateIn,user:User=Depends(current_user),db:Session=Depends(db_session)):
    project=get_project_for_update(pid,user,db)
    assignment=db.get(Assignment,assignment_id)
    if not assignment or assignment.project_id!=pid: raise HTTPException(404)
    periods=bounded_int(payload.periods_per_week,1,1,40,"Số tiết mỗi tuần")
    lessons=db.scalars(select(Lesson).where(Lesson.assignment_id==assignment.id)).all()
    scheduled=len(lessons)
    if periods<scheduled:
        return JSONResponse({"ok":False,"message":f"Đang có {scheduled} tiết trên lịch. Hãy gỡ bớt tiết trước khi giảm số tiết/tuần."},409)
    subject=db.get(Subject,assignment.subject_id)
    if not subject or subject.project_id!=pid: raise HTTPException(409,"Môn học của phân công không còn tồn tại")
    try: mode=normalized_block_mode(payload.block_mode,periods,subject,project)
    except ValueError as exc: raise HTTPException(400,str(exc)) from exc

    old_periods=assignment.periods_per_week
    old_mode=assignment.block_mode
    periods_changed=periods!=old_periods
    mode_changed=mode!=old_mode

    # Không xóa các Lesson đang có. Thay đổi được áp dụng trong transaction và
    # sẽ rollback nguyên vẹn nếu phần lịch hiện tại không tương thích.
    assignment.periods_per_week=periods
    assignment.block_mode=mode
    assignment.consecutive_pattern=""

    if mode_changed or (periods_changed and assignment_requires_double(assignment)):
        fixed_error=normalize_assignment_fixed_rows(db,project,assignment,lessons)
        if fixed_error:
            db.rollback()
            return JSONResponse({"ok":False,"message":fixed_error},409)

    # Chỉ các thay đổi ảnh hưởng cấu trúc cụm mới cần chứng minh rằng những tiết
    # đã xếp vẫn có ít nhất một cách hoàn thành. Tăng số tiết ở chế độ tự do/
    # ưu tiên chỉ tạo phần còn thiếu trong khay và giữ nguyên toàn bộ lịch cũ.
    if (mode_changed or (periods_changed and assignment_requires_double(assignment))) and not assignment_completion_feasible(
        db,project,assignment,[lesson.slot for lesson in lessons]
    ):
        db.rollback()
        return JSONResponse({
            "ok":False,
            "message":"Không thể áp dụng thay đổi vì các tiết hiện có không thể hoàn thành hợp lệ theo chế độ mới và các ràng buộc hiện tại. Lịch cũ được giữ nguyên.",
        },409)

    db.commit();return {"ok":True,"scheduled_preserved":scheduled}

@app.delete("/api/projects/{pid}/entity/{typ}/{eid}")
def delete_entity(pid:int,typ:str,eid:int,user:User=Depends(current_user),db:Session=Depends(db_session)):
    get_project_for_update(pid,user,db)
    model={"department":Department,"subject":Subject,"teacher":Teacher,"grade":Grade,"class":SchoolClass,"assignment":Assignment}.get(typ)
    if not model: raise HTTPException(400)
    obj=db.get(model,eid)
    if not obj or obj.project_id!=pid: raise HTTPException(404)
    dependency={
        "department":db.scalar(select(Teacher.id).where(Teacher.department_id==eid)),
        "subject":db.scalar(select(Assignment.id).where(Assignment.subject_id==eid)),
        "teacher":db.scalar(select(Assignment.id).where(Assignment.teacher_id==eid)),
        "grade":db.scalar(select(SchoolClass.id).where(SchoolClass.grade_id==eid)),
        "class":db.scalar(select(Assignment.id).where(Assignment.class_id==eid)),
    }.get(typ)
    if dependency is not None:
        return JSONResponse({"ok":False,"message":"Không thể xóa vì dữ liệu đang được sử dụng."},409)
    if typ=="teacher" and account_for_teacher(eid, db) is not None:
        return JSONResponse({"ok":False,"message":"Hãy thu hồi tài khoản giáo viên trước khi xóa."},409)
    if typ=="assignment":
        for l in db.scalars(select(Lesson).where(Lesson.assignment_id==eid)).all(): db.delete(l)
        for fixed_lesson in db.scalars(select(FixedLesson).where(FixedLesson.assignment_id==eid)).all(): db.delete(fixed_lesson)
    if typ=="teacher":
        for preference in db.scalars(select(TeacherPreference).where(TeacherPreference.teacher_id==eid)).all():
            db.delete(preference)
    db.delete(obj); db.commit(); return {"ok":True}

class ConstraintIn(BaseModel):
    entity_type: str
    entity_id: int
    slots: list[int]

@app.post("/api/projects/{pid}/constraints")
def constraints(pid:int,payload:ConstraintIn,user:User=Depends(current_user),db:Session=Depends(db_session)):
    p=get_project_for_update(pid,user,db)
    if payload.entity_type not in {"teacher","class"}: raise HTTPException(400,"Loại ràng buộc không hợp lệ")
    model=Teacher if payload.entity_type=="teacher" else SchoolClass
    obj=db.get(model,payload.entity_id)
    if not obj or obj.project_id!=pid: raise HTTPException(404)
    slots=set(valid_slots(p,payload.slots))
    assignments=db.scalars(select(Assignment).where(
        Assignment.project_id==pid,
        Assignment.teacher_id==obj.id if payload.entity_type=="teacher" else Assignment.class_id==obj.id,
    )).all()
    assignment_ids={assignment.id for assignment in assignments}
    lessons=db.scalars(select(Lesson).where(
        Lesson.project_id==pid,
        Lesson.assignment_id.in_(assignment_ids),
    )).all() if assignment_ids else []
    lessons_by_assignment=defaultdict(list)
    for lesson in lessons:
        lessons_by_assignment[lesson.assignment_id].append(lesson)

    fixed_rows=db.scalars(select(FixedLesson).where(
        FixedLesson.project_id==pid,
        FixedLesson.assignment_id.in_(assignment_ids),
    )).all() if assignment_ids else []
    assignment_by_id={assignment.id:assignment for assignment in assignments}
    for row in fixed_rows:
        assignment=assignment_by_id.get(row.assignment_id)
        if not assignment:
            continue
        size=fixed_row_size(p,assignment,row,lessons_by_assignment[row.assignment_id])
        if slots.intersection(range(row.slot,row.slot+size)):
            return JSONResponse({"ok":False,"message":"Ràng buộc mới xung đột với tiết cố định. Hãy bỏ cố định trước."},409)

    removed_ids=set()
    for assignment in assignments:
        assignment_lessons=lessons_by_assignment[assignment.id]
        if not assignment_requires_double(assignment):
            affected=[lesson for lesson in assignment_lessons if lesson.slot in slots]
            if any(lesson.locked for lesson in affected):
                return JSONResponse({"ok":False,"message":"Ràng buộc mới xung đột với tiết cố định. Hãy bỏ cố định trước."},409)
            removed_ids.update(lesson.id for lesson in affected)
            continue
        for run in assignment_run_groups(p,[lesson.slot for lesson in assignment_lessons]):
            if not slots.intersection(run["slots"]):
                continue
            run_slots=set(run["slots"])
            affected=[lesson for lesson in assignment_lessons if lesson.slot in run_slots]
            if any(lesson.locked for lesson in affected):
                return JSONResponse({"ok":False,"message":"Ràng buộc mới xung đột với tiết cố định. Hãy bỏ cố định trước."},409)
            removed_ids.update(lesson.id for lesson in affected)

    obj.unavailable_json=json.dumps(sorted(slots))
    for lesson in lessons:
        if lesson.id in removed_ids:
            db.delete(lesson)
    db.commit()
    return {"ok":True,"removed":len(removed_ids)}

class SessionLocksIn(BaseModel):
    sessions: list[int] = Field(default_factory=list)
    slots: list[int] = Field(default_factory=list)

@app.post("/api/projects/{pid}/session-locks")
def save_session_locks(pid:int,payload:SessionLocksIn,user:User=Depends(current_user),db:Session=Depends(db_session)):
    project=get_project_for_update(pid,user,db)
    maximum=project.days*project.sessions
    session_keys=sorted({int(value) for value in payload.sessions if 0<=int(value)<maximum})
    blocked=[]
    ppd=project.sessions*project.periods_per_session
    for key in session_keys:
        day=key//project.sessions
        session=key%project.sessions
        start=day*ppd+session*project.periods_per_session
        blocked.extend(range(start,start+project.periods_per_session))
    blocked=valid_slots(project,[*blocked,*payload.slots])
    all_lessons=db.scalars(select(Lesson).where(Lesson.project_id==pid)).all()
    lessons_by_assignment=defaultdict(list)
    for lesson in all_lessons:
        lessons_by_assignment[lesson.assignment_id].append(lesson)
    fixed_rows=db.scalars(select(FixedLesson).where(FixedLesson.project_id==pid)).all()
    blocked_set=set(blocked)
    removed_ids=set()

    # Khóa buổi phải tuân cùng quy tắc với ràng buộc giáo viên/lớp:
    # không âm thầm xóa các cụm đã được cố định.
    for row in fixed_rows:
        assignment=db.get(Assignment,row.assignment_id)
        if not assignment:
            continue
        lessons=lessons_by_assignment.get(row.assignment_id,[])
        size=fixed_row_size(project,assignment,row,lessons)
        row_slots=set(range(row.slot,row.slot+size))
        if blocked_set.intersection(row_slots):
            return JSONResponse({
                "ok":False,
                "message":"Buổi hoặc tiết mới khóa đang chứa tiết cố định. Hãy bỏ cố định trước.",
            },409)
    if any(lesson.locked and lesson.slot in blocked_set for lesson in all_lessons):
        return JSONResponse({
            "ok":False,
            "message":"Buổi hoặc tiết mới khóa đang chứa tiết cố định. Hãy bỏ cố định trước.",
        },409)

    project.blocked_slots_json=json.dumps(blocked)
    for assignment_id,lessons in lessons_by_assignment.items():
        if not any(lesson.slot in blocked_set for lesson in lessons):
            continue
        assignment=db.get(Assignment,assignment_id)
        if not assignment:
            for lesson in lessons:
                if lesson.slot in blocked_set:
                    removed_ids.add(lesson.id)
            continue
        if not assignment_requires_double(assignment):
            removed_ids.update(lesson.id for lesson in lessons if lesson.slot in blocked_set)
            continue
        for run in assignment_run_groups(project,[lesson.slot for lesson in lessons]):
            if blocked_set.intersection(run["slots"]):
                run_slots=set(run["slots"])
                removed_ids.update(lesson.id for lesson in lessons if lesson.slot in run_slots)
    for lesson in all_lessons:
        if lesson.id in removed_ids:
            db.delete(lesson)
    db.commit()
    return {"ok":True,"sessions":session_keys,"removed":len(removed_ids)}

class FixedIn(BaseModel):
    assignment_id:int
    slot:int

def fixed_row_size(project: Project, assignment: Assignment, row: FixedLesson, lessons: list[Lesson] | None = None) -> int:
    if not assignment_requires_double(assignment):
        return 1
    expected = assignment_groups(assignment)
    size = int(getattr(row, "group_size", 1) or 1)
    if lessons:
        for run in assignment_run_groups(project, [lesson.slot for lesson in lessons]):
            if run["start"] == row.slot and row.slot in run["slots"] and run["size"] in expected:
                return run["size"]
    if size in expected and not (size == 1 and 1 not in expected):
        return size
    return expected[0] if expected else 1

@app.post("/api/projects/{pid}/fixed")
def fixed(pid:int,payload:FixedIn,user:User=Depends(current_user),db:Session=Depends(db_session)):
    p=get_project_for_update(pid,user,db)
    assignment=db.get(Assignment,payload.assignment_id)
    if not assignment or assignment.project_id!=pid: raise HTTPException(404)
    if payload.slot not in set(all_slots(p)):
        raise HTTPException(400,"Ô thời khóa biểu không hợp lệ")
    lessons=db.scalars(select(Lesson).where(Lesson.project_id==pid,Lesson.assignment_id==assignment.id)).all()
    if assignment_requires_double(assignment):
        run=next((item for item in assignment_run_groups(p,[lesson.slot for lesson in lessons]) if payload.slot in item["slots"]),None)
    else:
        selected=next((lesson for lesson in lessons if lesson.slot==payload.slot),None)
        run={"start":payload.slot,"size":1,"slots":[payload.slot]} if selected else None
    if not run:
        # Tương thích với client cũ: payload.slot từng được hiểu là ô đích.
        # Chỉ cho phép khi phân công chưa có cụm cố định để tránh làm mất các ghim khác.
        existing_fixed=db.scalars(select(FixedLesson).where(FixedLesson.project_id==pid,FixedLesson.assignment_id==assignment.id)).all()
        if existing_fixed:
            raise HTTPException(409,"Hãy chọn trực tiếp một tiết đang có trên lịch để cố định tiết hoặc cặp đó")
        groups=assignment_groups(assignment)
        size=groups[0] if groups else 1
        day,session,period=slot_meta(p,payload.slot)
        if period+size>p.periods_per_session:
            raise HTTPException(409,"Cặp tiết cố định vượt quá cuối buổi học")
        for lesson in lessons:
            if lesson.locked:
                raise HTTPException(409,"Phân công đang có tiết cố định; không thể dùng chế độ di chuyển cũ")
            db.delete(lesson)
        db.add(FixedLesson(project_id=pid,assignment_id=assignment.id,slot=payload.slot,group_size=size))
        db.flush()
        result=solve_missing(db,p,tries=180,target_assignment_ids={assignment.id})
        target=[row for row in result["lessons"] if row[0]==assignment.id]
        if len(target)<assignment.periods_per_week:
            db.rollback()
            return JSONResponse({"ok":False,"message":"Không thể cố định phân công tại vị trí này vì xung đột lớp, giáo viên hoặc ràng buộc."},409)
        for aid,slot,locked in result["lessons"]:
            db.add(Lesson(project_id=pid,assignment_id=aid,slot=slot,locked=locked))
        db.commit()
        return {"ok":True,"message":f"Đã chuyển và cố định {size} tiết tại ô đã chọn."}
    if remaining_pattern_groups(p,assignment,[lesson.slot for lesson in lessons]) is None:
        raise HTTPException(409,"Lịch hiện tại chưa phù hợp với chế độ xếp tiết; hãy xếp lại trước khi cố định")
    expected=Counter(assignment_groups(assignment))
    if expected[run["size"]] <= 0:
        raise HTTPException(409,"Tiết hoặc cặp đang chọn không phù hợp với chế độ của phân công")
    fixed_rows=db.scalars(select(FixedLesson).where(FixedLesson.project_id==pid,FixedLesson.assignment_id==assignment.id)).all()
    used=Counter()
    for row in fixed_rows:
        size=fixed_row_size(p,assignment,row,lessons)
        if row.slot==run["start"]:
            row.group_size=run["size"]
            for lesson in lessons:
                if lesson.slot in run["slots"]: lesson.locked=True
            db.commit()
            return {"ok":True,"message":f"Đã cố định {run['size']} tiết đang chọn."}
        used[size]+=1
    if used[run["size"]] >= expected[run["size"]]:
        raise HTTPException(409,"Số tiết/cặp cố định đã vượt số lượng cho phép của phân công")
    for lesson in lessons:
        if lesson.slot in run["slots"]:
            error=lesson_slot_error(db,p,assignment,lesson.slot,lesson.id)
            if error: raise HTTPException(409,error)
            lesson.locked=True
    db.add(FixedLesson(project_id=pid,assignment_id=assignment.id,slot=run["start"],group_size=run["size"]))
    db.commit()
    return {"ok":True,"message":f"Đã cố định {run['size']} tiết đang chọn."}

@app.delete("/api/projects/{pid}/fixed/{assignment_id}/{slot}")
def remove_fixed_group(pid:int,assignment_id:int,slot:int,user:User=Depends(current_user),db:Session=Depends(db_session)):
    p=get_project_for_update(pid,user,db)
    assignment=db.get(Assignment,assignment_id)
    if not assignment or assignment.project_id!=pid: raise HTTPException(404)
    lessons=db.scalars(select(Lesson).where(Lesson.project_id==pid,Lesson.assignment_id==assignment_id)).all()
    rows=db.scalars(select(FixedLesson).where(FixedLesson.project_id==pid,FixedLesson.assignment_id==assignment_id)).all()
    targets=[]
    for row in rows:
        size=fixed_row_size(p,assignment,row,lessons)
        if row.slot<=slot<row.slot+size:
            targets.append((row,size))
    if not targets:
        raise HTTPException(404,"Không tìm thấy tiết hoặc cặp cố định")
    unlocked=set()
    for row,size in targets:
        unlocked.update(range(row.slot,row.slot+size));db.delete(row)
    remaining=[]
    for row in rows:
        if all(row.id!=target.id for target,_size in targets):
            size=fixed_row_size(p,assignment,row,lessons)
            remaining.extend(range(row.slot,row.slot+size))
    for lesson in lessons:
        if lesson.slot in unlocked and lesson.slot not in remaining:
            lesson.locked=False
    db.commit()
    return {"ok":True,"message":"Đã bỏ cố định tiết hoặc cặp đang chọn."}

@app.delete("/api/projects/{pid}/fixed/{assignment_id}")
def remove_fixed(pid:int,assignment_id:int,user:User=Depends(current_user),db:Session=Depends(db_session)):
    get_project_for_update(pid,user,db)
    assignment=db.get(Assignment,assignment_id)
    if not assignment or assignment.project_id!=pid: raise HTTPException(404)
    fixed_rows=db.scalars(select(FixedLesson).where(FixedLesson.project_id==pid,FixedLesson.assignment_id==assignment_id)).all()
    for row in fixed_rows: db.delete(row)
    lessons=db.scalars(select(Lesson).where(Lesson.project_id==pid,Lesson.assignment_id==assignment_id)).all()
    for lesson in lessons: lesson.locked=False
    db.commit()
    return {"ok":True,"message":"Đã bỏ toàn bộ cố định của phân công."}

@app.post("/api/projects/{pid}/generate")
def generate(pid:int,user:User=Depends(current_user),db:Session=Depends(db_session)):
    p=get_project_for_update(pid,user,db)
    assignments=db.scalars(select(Assignment).where(Assignment.project_id==pid)).all()
    existing=db.scalars(select(Lesson).where(Lesson.project_id==pid)).all()
    assignment_by_id={assignment.id:assignment for assignment in assignments}

    # Chuẩn hóa dữ liệu cố định cũ và hỗ trợ nhiều cụm cố định trên một phân công.
    lessons_by_assignment=defaultdict(list)
    for lesson in existing:
        lessons_by_assignment[lesson.assignment_id].append(lesson)
    fixed_rows=db.scalars(select(FixedLesson).where(FixedLesson.project_id==pid)).all()
    fixed_changed=False
    assignments_with_unsatisfied_fixed=set()
    for fixed_row in fixed_rows:
        assignment=assignment_by_id.get(fixed_row.assignment_id)
        if not assignment:
            db.delete(fixed_row);fixed_changed=True;continue
        size=fixed_row_size(p,assignment,fixed_row,lessons_by_assignment[assignment.id])
        if fixed_row.group_size!=size:
            fixed_row.group_size=size;fixed_changed=True
        expected_slots=set(range(fixed_row.slot,fixed_row.slot+size))
        matching={lesson.slot:lesson for lesson in lessons_by_assignment[assignment.id] if lesson.slot in expected_slots}
        if expected_slots.issubset(matching):
            for slot in expected_slots:
                if not matching[slot].locked:
                    matching[slot].locked=True;fixed_changed=True
        else:
            assignments_with_unsatisfied_fixed.add(assignment.id)
    for assignment_id in assignments_with_unsatisfied_fixed:
        for lesson in lessons_by_assignment[assignment_id]:
            if not lesson.locked:
                db.delete(lesson);fixed_changed=True
    if fixed_changed:
        db.flush()
        existing=db.scalars(select(Lesson).where(Lesson.project_id==pid)).all()

    # Rà lại lịch cũ trước mỗi lần xếp. Ràng buộc hoặc nguyện vọng có thể đã
    # thay đổi sau khi lịch được tạo, vì vậy không được coi "đủ tiết" là hợp lệ.
    invalid_lessons=[]
    rebuild_assignment_ids=set()
    for lesson in existing:
        assignment=assignment_by_id.get(lesson.assignment_id)
        if not assignment:
            invalid_lessons.append(lesson)
            continue
        if lesson_slot_error(db,p,assignment,lesson.slot,lesson.id):
            if assignment_requires_double(assignment):
                rebuild_assignment_ids.add(assignment.id)
            else:
                invalid_lessons.append(lesson)

    existing_by_assignment=defaultdict(list)
    for lesson in existing:
        existing_by_assignment[lesson.assignment_id].append(lesson)
    for assignment in assignments:
        lessons_for_assignment=existing_by_assignment[assignment.id]
        if not lessons_for_assignment:
            continue
        slots_for_assignment=[lesson.slot for lesson in lessons_for_assignment]
        remaining_groups=remaining_pattern_groups(p,assignment,slots_for_assignment)
        if remaining_groups is None:
            rebuild_assignment_ids.add(assignment.id)

    if rebuild_assignment_ids:
        invalid_lessons.extend(
            lesson for lesson in existing
            if lesson.assignment_id in rebuild_assignment_ids
        )
    invalid_lessons=list({lesson.id:lesson for lesson in invalid_lessons}.values())
    locked_invalid=[lesson for lesson in invalid_lessons if lesson.locked]
    if locked_invalid:
        return JSONResponse({
            "ok":False,
            "message":f"Có {len(locked_invalid)} tiết cố định xung đột với ràng buộc hoặc chế độ xếp tiết mới. Hãy bỏ cố định hoặc điều chỉnh ràng buộc trước khi xếp lại.",
        },409)
    for lesson in invalid_lessons:
        db.delete(lesson)
    if invalid_lessons:
        db.flush()
        existing=db.scalars(select(Lesson).where(Lesson.project_id==pid)).all()

    expected=sum(a.periods_per_week for a in assignments)
    missing=max(0,expected-len(existing))

    # Từ lần xếp thứ hai trở đi, tuyệt đối giữ nguyên các tiết đang có.
    # Chỉ bổ sung những tiết còn thiếu trong khay; nếu lịch đã đủ thì không làm gì.
    if existing:
        if missing == 0:
            if invalid_lessons or fixed_changed:
                db.commit()
            return {
                "ok":True,"score":0,"unscheduled":0,
                "message":"Thời khóa biểu đã đủ tiết. Các vị trí hiện tại được giữ nguyên.",
            }
        result=solve_missing(db,p,tries=160)
        if result["unscheduled"]>0:
            # Các tiết đang có có thể tự chặn phần còn thiếu. Thử giữ nguyên
            # tiết cố định và tái tối ưu toàn bộ phần không cố định.
            rebuild=solve_rebuild(db,p,tries=260)
            if rebuild["unscheduled"]>0:
                if rebuild.get("proven_infeasible"):
                    message="Các ràng buộc và tiết cố định hiện tại không cho phép tạo một thời khóa biểu đầy đủ."
                else:
                    message=(
                        f"Chưa tìm được lịch đầy đủ sau các bước bổ sung và tái tối ưu; "
                        f"còn {rebuild['unscheduled']} tiết chưa xếp. Dữ liệu lịch hiện tại chưa bị thay đổi."
                    )
                return JSONResponse({
                    "ok":False,"score":rebuild["score"],"unscheduled":rebuild["unscheduled"],
                    "message":message,
                },409)
            current_lessons=db.scalars(select(Lesson).where(Lesson.project_id==pid)).all()
            moved_count=sum(1 for lesson in current_lessons if not lesson.locked)
            for lesson in current_lessons:
                if not lesson.locked:
                    db.delete(lesson)
            db.flush()
            for aid,slot,locked in rebuild["lessons"]:
                db.add(Lesson(project_id=pid,assignment_id=aid,slot=slot,locked=locked))
            db.commit()
            return {
                "ok":True,"score":rebuild["score"],"unscheduled":0,
                "message":f"Đã giữ nguyên các tiết cố định và xếp lại {moved_count} tiết không cố định để hoàn thành thời khóa biểu.",
            }
        for aid,slot,locked in result["lessons"]:
            db.add(Lesson(project_id=pid,assignment_id=aid,slot=slot,locked=locked))
        db.commit()
        return {
            "ok":True,"score":result["score"],"unscheduled":0,
            "message":f"Đã xếp bổ sung {len(result['lessons'])} tiết từ khay và giữ nguyên các vị trí còn hợp lệ.",
        }

    # Chỉ khi lịch hoàn toàn trống mới chạy bộ xếp toàn bộ.
    result=solve(db,p,tries=120)
    if result["unscheduled"]>0:
        if result.get("proven_infeasible"):
            message="Các ràng buộc hiện tại không cho phép tạo một thời khóa biểu đầy đủ."
        else:
            message=f"Chưa tìm được lịch đầy đủ sau số lần thử hiện tại; còn {result['unscheduled']} tiết chưa xếp. Lịch hiện tại được giữ nguyên."
        return JSONResponse({
            "ok":False,"score":result["score"],"unscheduled":result["unscheduled"],
            "message":message,
        },409)
    for l in existing: db.delete(l)
    for aid,slot,locked in result["lessons"]:
        db.add(Lesson(project_id=pid,assignment_id=aid,slot=slot,locked=locked))
    db.commit()
    return {"ok":True,"score":result["score"],"unscheduled":0,"message":f"Đã xếp đầy đủ {len(result['lessons'])} tiết."}

def lesson_slot_error(db:Session,project:Project,assignment:Assignment,slot:int,exclude_lesson_id:Optional[int]=None):
    if slot not in all_slots(project): return "Ô thời khóa biểu không hợp lệ."
    if slot in parse_slots(project.blocked_slots_json): return "Buổi này đã bị khóa và không được xếp tiết."
    teacher=db.get(Teacher,assignment.teacher_id);school_class=db.get(SchoolClass,assignment.class_id);subject=db.get(Subject,assignment.subject_id)
    if not teacher or not school_class or not subject: return "Phân công không còn đầy đủ lớp, môn hoặc giáo viên."
    if slot in parse_slots(teacher.unavailable_json): return "Giáo viên không thể dạy ở tiết này theo ràng buộc chính thức."
    _, accepted_unavailable = accepted_teacher_preferences(db, project.id)
    if slot in accepted_unavailable.get(teacher.id, set()):
        return "Tiết này nằm trong nguyện vọng cần tránh đã được duyệt của giáo viên."
    if slot in parse_slots(school_class.unavailable_json): return "Lớp không học ở tiết này."
    existing_lessons=db.scalars(select(Lesson).where(Lesson.project_id==project.id)).all()
    existing_lessons=[lesson for lesson in existing_lessons if lesson.id!=exclude_lesson_id]
    for lesson in existing_lessons:
        if lesson.slot!=slot: continue
        other=db.get(Assignment,lesson.assignment_id)
        if other and (other.class_id==assignment.class_id or other.teacher_id==assignment.teacher_id):
            return "Ô đích bị trùng lớp hoặc giáo viên."
    ppd=project.sessions*project.periods_per_session;target_day=slot//ppd
    target_position=slot%ppd;target_session=target_position//project.periods_per_session
    teacher_periods=0;subject_periods=[]
    for lesson in existing_lessons:
        other=db.get(Assignment,lesson.assignment_id)
        if not other or lesson.slot//ppd!=target_day: continue
        if other.teacher_id==assignment.teacher_id: teacher_periods+=1
        position=lesson.slot%ppd
        if position//project.periods_per_session==target_session and other.class_id==assignment.class_id and other.subject_id==assignment.subject_id:
            subject_periods.append(position%project.periods_per_session)
    if teacher_periods>=teacher.max_periods_day: return "Giáo viên đã đạt số tiết tối đa trong ngày."
    run=sorted(subject_periods+[target_position%project.periods_per_session]);longest=current=1
    for left,right in zip(run,run[1:]):
        current=current+1 if right==left+1 else 1;longest=max(longest,current)
    if longest>subject.max_consecutive: return "Vượt số tiết liên tiếp tối đa của môn học."
    return None

class MoveIn(BaseModel):
    lesson_id:int
    slot:int

@app.post("/api/projects/{pid}/move")
def move(pid:int,payload:MoveIn,user:User=Depends(current_user),db:Session=Depends(db_session)):
    project=get_project_for_update(pid,user,db);lesson=db.get(Lesson,payload.lesson_id)
    if not lesson or lesson.project_id!=pid: raise HTTPException(404)
    if lesson.locked: return JSONResponse({"ok":False,"message":"Tiết cố định không thể di chuyển."},409)
    assignment=db.get(Assignment,lesson.assignment_id)
    if not assignment or assignment.project_id!=pid:
        return JSONResponse({"ok":False,"message":"Phân công của tiết học không còn tồn tại."},409)
    error=lesson_slot_error(db,project,assignment,payload.slot,lesson.id)
    if error: return JSONResponse({"ok":False,"message":error},409)
    assignment_lessons=db.scalars(select(Lesson).where(Lesson.assignment_id==assignment.id)).all()
    proposed_slots=[payload.slot if item.id==lesson.id else item.slot for item in assignment_lessons]
    if not assignment_completion_feasible(db,project,assignment,proposed_slots):
        return JSONResponse({"ok":False,"message":f"Vị trí này không thể hoàn thành hợp lệ theo chế độ {assignment_pattern_label(assignment)} và các ràng buộc hiện tại."},409)
    lesson.slot=payload.slot;db.commit();return {"ok":True}

class ManualLessonIn(BaseModel):
    assignment_id:int
    slot:int

@app.post("/api/projects/{pid}/lessons")
def add_manual_lesson(pid:int,payload:ManualLessonIn,user:User=Depends(current_user),db:Session=Depends(db_session)):
    project=get_project_for_update(pid,user,db);assignment=db.get(Assignment,payload.assignment_id)
    if not assignment or assignment.project_id!=pid: raise HTTPException(404)
    scheduled=db.scalar(select(func.count(Lesson.id)).where(Lesson.assignment_id==assignment.id)) or 0
    if scheduled>=assignment.periods_per_week:
        return JSONResponse({"ok":False,"message":"Phân công này đã đủ số tiết/tuần."},409)
    error=lesson_slot_error(db,project,assignment,payload.slot)
    if error: return JSONResponse({"ok":False,"message":error},409)
    current_slots=db.scalars(select(Lesson.slot).where(Lesson.assignment_id==assignment.id)).all()
    proposed_slots=[*current_slots,payload.slot]
    if not assignment_completion_feasible(db,project,assignment,proposed_slots):
        return JSONResponse({"ok":False,"message":f"Vị trí này không thể hoàn thành hợp lệ theo chế độ {assignment_pattern_label(assignment)} và các ràng buộc hiện tại."},409)
    lesson=Lesson(project_id=pid,assignment_id=assignment.id,slot=payload.slot,locked=False)
    db.add(lesson);db.commit();return {"ok":True,"id":lesson.id}

@app.delete("/api/projects/{pid}/lessons/{lesson_id}")
def remove_manual_lesson(pid:int,lesson_id:int,user:User=Depends(current_user),db:Session=Depends(db_session)):
    get_project_for_update(pid,user,db);lesson=db.get(Lesson,lesson_id)
    if not lesson or lesson.project_id!=pid: raise HTTPException(404)
    if lesson.locked: return JSONResponse({"ok":False,"message":"Tiết cố định không thể gỡ."},409)
    db.delete(lesson);db.commit();return {"ok":True}

@app.delete("/api/projects/{pid}/assignments/{assignment_id}/lessons")
def return_assignment_to_tray(pid:int,assignment_id:int,user:User=Depends(current_user),db:Session=Depends(db_session)):
    get_project_for_update(pid,user,db);assignment=db.get(Assignment,assignment_id)
    if not assignment or assignment.project_id!=pid: raise HTTPException(404)
    lessons=db.scalars(select(Lesson).where(Lesson.project_id==pid,Lesson.assignment_id==assignment_id)).all()
    removable=[lesson for lesson in lessons if not lesson.locked];locked=len(lessons)-len(removable)
    for lesson in removable: db.delete(lesson)
    db.commit()
    message=f"Đã đưa {len(removable)} tiết về khay."
    if locked: message+=f" Còn {locked} tiết cố định được giữ lại."
    return {"ok":True,"removed":len(removable),"locked":locked,"message":message}

@app.delete("/api/projects/{pid}/lessons")
def return_all_to_tray(pid:int,user:User=Depends(current_user),db:Session=Depends(db_session)):
    get_project_for_update(pid,user,db);lessons=db.scalars(select(Lesson).where(Lesson.project_id==pid)).all()
    removable=[lesson for lesson in lessons if not lesson.locked];locked=len(lessons)-len(removable)
    for lesson in removable: db.delete(lesson)
    db.commit()
    message=f"Đã đưa {len(removable)} tiết về khay."
    if locked: message+=f" Còn {locked} tiết cố định được giữ lại."
    return {"ok":True,"removed":len(removable),"locked":locked,"message":message}

@app.get("/api/projects/{pid}/data")
def api_data(pid:int,user:User=Depends(current_user),db:Session=Depends(db_session)):
    p=get_project(pid,user,db);return project_data(db,p)

class TeacherAccountIn(BaseModel):
    teacher_id: int
    email: str
    password: str = ""

@app.get("/api/projects/{pid}/teacher-accounts")
def list_teacher_accounts(pid:int,user:User=Depends(current_user),db:Session=Depends(db_session)):
    get_project(pid,user,db)
    teachers=db.scalars(select(Teacher).where(Teacher.project_id==pid).order_by(Teacher.name)).all()
    items=[]
    for teacher in teachers:
        account=account_for_teacher(teacher.id,db)
        items.append({
            "teacher_id":teacher.id,
            "teacher_name":teacher.name,
            "account_id":account.id if account else None,
            "email":account.email if account else None,
        })
    return {"items":items}

@app.post("/api/projects/{pid}/teacher-accounts")
def save_teacher_account(pid:int,payload:TeacherAccountIn,user:User=Depends(current_user),db:Session=Depends(db_session)):
    # Khóa project để hai yêu cầu cấp cùng một tài khoản trong cùng project
    # không thể đồng thời vượt qua bước kiểm tra liên kết.
    get_project_for_update(pid,user,db)
    teacher=db.get(Teacher,payload.teacher_id)
    if not teacher or teacher.project_id!=pid: raise HTTPException(404,"Giáo viên không thuộc bộ thời khóa biểu")
    email=bounded_text(payload.email.lower(),"Email",255)
    if "@" not in email: raise HTTPException(400,"Email không hợp lệ")
    account=account_for_teacher(teacher.id,db)
    email_owner=db.scalar(select(User).where(User.email==email).with_for_update())
    if email_owner and (not account or email_owner.id!=account.id):
        if email_owner.role!="teacher" or not admin_can_manage_account(user,email_owner,db):
            raise HTTPException(409,"Email đã được sử dụng")
        account=email_owner
    if account:
        linked_teacher_ids=account_teacher_ids(account,db)-{teacher.id}
        linked_in_project=db.scalar(select(Teacher.id).where(
            Teacher.id.in_(linked_teacher_ids),
            Teacher.project_id==pid,
        )) if linked_teacher_ids else None
        if linked_in_project is not None:
            raise HTTPException(
                409,
                "Tài khoản này đã được liên kết với một giáo viên khác trong cùng bộ thời khóa biểu",
            )
        account.email=email; account.name=teacher.name
        if payload.password:
            if len(payload.password)<6: raise HTTPException(400,"Mật khẩu phải có ít nhất 6 ký tự")
            account.password_hash=pwd.hash(payload.password)
            account.session_version+=1
    else:
        if len(payload.password)<6: raise HTTPException(400,"Mật khẩu phải có ít nhất 6 ký tự")
        account=User(email=email,name=teacher.name,password_hash=pwd.hash(payload.password),role="teacher",teacher_id=teacher.id)
        db.add(account);db.flush()
    ensure_teacher_link(account,teacher,db)
    db.commit()
    return {"ok":True,"account_id":account.id}

@app.delete("/api/projects/{pid}/teacher-accounts/{teacher_id}")
def revoke_teacher_account(pid:int,teacher_id:int,user:User=Depends(current_user),db:Session=Depends(db_session)):
    get_project(pid,user,db)
    teacher=db.get(Teacher,teacher_id)
    if not teacher or teacher.project_id!=pid: raise HTTPException(404)
    account=account_for_teacher(teacher.id,db)
    if not account or account.role!="teacher": raise HTTPException(404)
    link=db.scalar(select(TeacherAccountLink).where(
        TeacherAccountLink.user_id==account.id,TeacherAccountLink.teacher_id==teacher.id
    ))
    if link:
        db.delete(link)
    remaining_ids=account_teacher_ids(account,db)-{teacher.id}
    if account.teacher_id==teacher.id:
        account.teacher_id=next(iter(remaining_ids),None)
    if not remaining_ids:
        db.flush()
        db.delete(account)
    db.commit();return {"ok":True}

def teacher_profiles_for_user(user:User,db:Session):
    if user.role!="teacher": raise HTTPException(403,"Tài khoản giáo viên không hợp lệ")
    teacher_ids=account_teacher_ids(user,db)
    if not teacher_ids: raise HTTPException(403,"Tài khoản giáo viên chưa được liên kết với hồ sơ nào")
    teachers=db.scalars(select(Teacher).where(Teacher.id.in_(teacher_ids))).all()
    profiles=[]
    for teacher in teachers:
        project=db.get(Project,teacher.project_id)
        if project:
            profiles.append((teacher,project))
    if not profiles: raise HTTPException(403,"Hồ sơ giáo viên không còn tồn tại")
    return sorted(profiles,key=lambda pair:pair[1].id,reverse=True)

def teacher_for_user(user:User,db:Session,project_id:Optional[int]=None):
    profiles=teacher_profiles_for_user(user,db)
    if project_id is not None:
        matches=[pair for pair in profiles if pair[1].id==project_id]
        if not matches: raise HTTPException(403,"Tài khoản không được cấp quyền cho bộ thời khóa biểu này")
        if len(matches)>1:
            raise HTTPException(
                409,
                "Tài khoản đang liên kết với nhiều giáo viên trong cùng bộ thời khóa biểu; quản trị viên cần thu hồi liên kết bị trùng",
            )
        return matches[0]
    latest_project_id=profiles[0][1].id
    latest_matches=[pair for pair in profiles if pair[1].id==latest_project_id]
    if len(latest_matches)>1:
        raise HTTPException(
            409,
            "Tài khoản đang liên kết với nhiều giáo viên trong cùng bộ thời khóa biểu; quản trị viên cần thu hồi liên kết bị trùng",
        )
    return profiles[0]

def teacher_project_data(db:Session,project:Project,teacher:Teacher):
    assignments=db.scalars(select(Assignment).where(
        Assignment.project_id==project.id,Assignment.teacher_id==teacher.id
    )).all()
    assignment_ids={item.id for item in assignments}
    class_ids={item.class_id for item in assignments}
    subject_ids={item.subject_id for item in assignments}
    classes={item.id:item for item in db.scalars(select(SchoolClass).where(
        SchoolClass.project_id==project.id,SchoolClass.id.in_(class_ids)
    )).all()} if class_ids else {}
    subjects={item.id:item for item in db.scalars(select(Subject).where(
        Subject.project_id==project.id,Subject.id.in_(subject_ids)
    )).all()} if subject_ids else {}
    grade_ids={item.grade_id for item in classes.values() if item.grade_id is not None}
    grades=db.scalars(select(Grade).where(
        Grade.project_id==project.id,Grade.id.in_(grade_ids)
    )).all() if grade_ids else []
    departments=[]
    if teacher.department_id is not None:
        department=db.scalar(select(Department).where(
            Department.project_id==project.id,Department.id==teacher.department_id
        ))
        if department:
            departments=[department]
    lessons=db.scalars(select(Lesson).where(
        Lesson.project_id==project.id,Lesson.assignment_id.in_(assignment_ids)
    )).all() if assignment_ids else []
    return {
        "project":{
            "id":project.id,"name":project.name,"school_name":project.school_name,
            "days":project.days,"sessions":project.sessions,"periods":project.periods_per_session,
            "blocked_slots":valid_slots(project,parse_slots(project.blocked_slots_json)),
        },
        "departments":[{"id":item.id,"name":item.name} for item in departments],
        "subjects":[{
            "id":item.id,"name":item.name,"short_name":item.short_name,
            "max_consecutive":item.max_consecutive,
        } for item in subjects.values()],
        "teachers":[{
            "id":teacher.id,"name":teacher.name,"short_name":teacher.short_name,
            "department_id":teacher.department_id,"max_periods_day":teacher.max_periods_day,
            "unavailable":list(parse_slots(teacher.unavailable_json)),
        }],
        "grades":[{"id":item.id,"name":item.name} for item in grades],
        "classes":[{
            "id":item.id,"name":item.name,"grade_id":item.grade_id,
            "unavailable":list(parse_slots(item.unavailable_json)),
        } for item in classes.values()],
        "assignments":[{
            "id":item.id,"class_id":item.class_id,"subject_id":item.subject_id,
            "teacher_id":item.teacher_id,"periods_per_week":item.periods_per_week,
            "block_mode":item.block_mode,
            "class_name":classes[item.class_id].name if item.class_id in classes else "?",
            "subject_name":subjects[item.subject_id].name if item.subject_id in subjects else "?",
            "subject_short":subjects[item.subject_id].short_name if item.subject_id in subjects else "?",
            "teacher_name":teacher.name,"teacher_short":teacher.short_name,
        } for item in assignments],
        "lessons":[{
            "id":item.id,"assignment_id":item.assignment_id,"slot":item.slot,"locked":item.locked,
        } for item in lessons],
        "coverage":{
            "unassigned_teachers":[],"unassigned_subjects":[],"unassigned_classes":[],
        },
    }

@app.get("/teacher",response_class=HTMLResponse)
def teacher_portal(request:Request,project_id:Optional[int]=None,user:User=Depends(current_user),db:Session=Depends(db_session)):
    teacher,project=teacher_for_user(user,db,project_id)
    profiles=teacher_profiles_for_user(user,db)
    preferences=[x for x in preference_payload(db,project) if x["teacher_id"]==teacher.id]
    return templates.TemplateResponse("teacher_portal.html",{
        "request":request,"user":user,"teacher":teacher,"p":project,
        "data":teacher_project_data(db,project,teacher),"preferences":preferences,"days":DAYS,
        "teacher_projects":[{"teacher":item_teacher,"project":item_project} for item_teacher,item_project in profiles],
    })

@app.get("/api/teacher/data")
def api_teacher_data(project_id:Optional[int]=None,user:User=Depends(current_user),db:Session=Depends(db_session)):
    teacher,project=teacher_for_user(user,db,project_id)
    return teacher_project_data(db,project,teacher)

@app.get("/teacher/account",response_class=HTMLResponse)
def teacher_account_page(request:Request,project_id:Optional[int]=None,user:User=Depends(current_user),db:Session=Depends(db_session)):
    teacher,project=teacher_for_user(user,db,project_id)
    return templates.TemplateResponse("teacher_account.html",{
        "request":request,"user":user,"teacher":teacher,"p":project,"error":None,"success":None,
    })

@app.post("/teacher/account",response_class=HTMLResponse)
def update_teacher_account(
    request:Request,
    email:str=Form(...),
    current_password:str=Form(...),
    new_password:str=Form(""),
    confirm_password:str=Form(""),
    project_id:Optional[int]=None,
    user:User=Depends(current_user),
    db:Session=Depends(db_session),
):
    teacher,project=teacher_for_user(user,db,project_id)
    context={"request":request,"user":user,"teacher":teacher,"p":project,"error":None,"success":None}
    if not pwd.verify(current_password,user.password_hash):
        context["error"]="Mật khẩu hiện tại không đúng."
        return templates.TemplateResponse("teacher_account.html",context,status_code=400)
    normalized_email=email.lower().strip()
    if not normalized_email or "@" not in normalized_email:
        context["error"]="Email không hợp lệ."
        return templates.TemplateResponse("teacher_account.html",context,status_code=400)
    if len(normalized_email)>255:
        context["error"]="Email không được vượt quá 255 ký tự."
        return templates.TemplateResponse("teacher_account.html",context,status_code=400)
    email_owner=db.scalar(select(User).where(User.email==normalized_email,User.id!=user.id))
    if email_owner:
        context["error"]="Email đã được sử dụng bởi tài khoản khác."
        return templates.TemplateResponse("teacher_account.html",context,status_code=409)
    password_changed=bool(new_password)
    if password_changed:
        if len(new_password)<6:
            context["error"]="Mật khẩu mới phải có ít nhất 6 ký tự."
            return templates.TemplateResponse("teacher_account.html",context,status_code=400)
        if new_password!=confirm_password:
            context["error"]="Xác nhận mật khẩu mới không khớp."
            return templates.TemplateResponse("teacher_account.html",context,status_code=400)
        user.password_hash=pwd.hash(new_password)
        user.session_version+=1
    user.email=normalized_email
    db.commit()
    context["user"]=user
    context["success"]="Thông tin tài khoản đã được cập nhật."
    response=templates.TemplateResponse("teacher_account.html",context)
    if password_changed:
        set_session_cookie(response,user)
    return response

@app.get("/share/{token}",response_class=HTMLResponse)
def shared(token:str,request:Request,db:Session=Depends(db_session)):
    p=db.scalar(select(Project).where(Project.share_token==token))
    if not p: raise HTTPException(404)
    return templates.TemplateResponse("share.html",{"request":request,"p":p,"data":public_project_data(db,p),"days":DAYS})

@app.get("/preferences/{token}",response_class=HTMLResponse)
def teacher_preferences_page(token:str,request:Request,user:User=Depends(current_user),db:Session=Depends(db_session)):
    p=db.scalar(select(Project).where(Project.share_token==token))
    if not p: raise HTTPException(404)
    teacher,teacher_project=teacher_for_user(user,db,p.id)
    latest_preference=db.scalar(
        select(TeacherPreference).where(
            TeacherPreference.project_id==p.id,
            TeacherPreference.teacher_id==teacher.id,
        ).order_by(TeacherPreference.id.desc())
    )
    latest_payload=None
    if latest_preference:
        latest_payload={
            "id":latest_preference.id,
            "preferred_slots":valid_slots(p,parse_slots(latest_preference.preferred_json)),
            "unavailable_slots":valid_slots(p,parse_slots(latest_preference.unavailable_json)),
            "note":latest_preference.note,
            "status":latest_preference.status,
            "created_at":latest_preference.created_at,
        }
    return templates.TemplateResponse(
        "teacher_preferences.html",
        {"request":request,"p":p,"teacher":teacher,"days":DAYS,"latest_preference":latest_payload},
    )

class TeacherPreferenceIn(BaseModel):
    teacher_id: int
    preferred_slots: list[int] = Field(default_factory=list)
    unavailable_slots: list[int] = Field(default_factory=list)
    note: str = ""

@app.post("/api/preferences/{token}")
def submit_teacher_preference(token:str,payload:TeacherPreferenceIn,user:User=Depends(current_user),db:Session=Depends(db_session)):
    # Dùng cùng khóa project với luồng duyệt để nội dung nguyện vọng không thể
    # bị thay đổi giữa lúc quản trị viên đang áp dụng nó vào lịch.
    p=db.scalar(select(Project).where(Project.share_token==token).with_for_update())
    if not p: raise HTTPException(404)
    teacher,teacher_project=teacher_for_user(user,db,p.id)
    if payload.teacher_id!=teacher.id: raise HTTPException(403,"Không thể gửi nguyện vọng thay giáo viên khác")
    preferred=set(valid_slots(p,payload.preferred_slots))
    unavailable=set(valid_slots(p,payload.unavailable_slots))
    preferred-=unavailable
    note=payload.note.strip()[:1000]
    pending=db.scalar(
        select(TeacherPreference).where(
            TeacherPreference.project_id==p.id,
            TeacherPreference.teacher_id==teacher.id,
            TeacherPreference.status=="pending",
        ).order_by(TeacherPreference.id.desc()).with_for_update()
    )
    if pending:
        # Trạng thái được đọc lại sau khi đã lấy khóa giao dịch.
        if pending.status!="pending":
            raise HTTPException(409,"Nguyện vọng này đang được xử lý, vui lòng tải lại trang")
        pending.preferred_json=json.dumps(sorted(preferred))
        pending.unavailable_json=json.dumps(sorted(unavailable))
        pending.note=note
        pending.created_at=datetime.now().isoformat(timespec="seconds")
        preference=pending
    else:
        preference=TeacherPreference(
            project_id=p.id,
            teacher_id=teacher.id,
            preferred_json=json.dumps(sorted(preferred)),
            unavailable_json=json.dumps(sorted(unavailable)),
            note=note,
        )
        db.add(preference)
    db.commit()
    return {"ok":True,"message":"Nguyện vọng đã được gửi đến người xếp thời khóa biểu."}

def preference_payload(db:Session,p:Project):
    rows=db.scalars(
        select(TeacherPreference).where(TeacherPreference.project_id==p.id).order_by(TeacherPreference.id.desc())
    ).all()
    teachers={x.id:x for x in db.scalars(select(Teacher).where(Teacher.project_id==p.id))}
    def label(slot:int):
        day,session,period=slot_meta(p,slot)
        session_text=f"{'Sáng' if session==0 else 'Chiều'} · " if p.sessions>1 else ""
        return f"{DAYS[day]} · {session_text}Tiết {period+1}"
    items=[]
    for row in rows:
        preferred_slots=valid_slots(p,parse_slots(row.preferred_json))
        unavailable_slots=valid_slots(p,parse_slots(row.unavailable_json))
        items.append({
            "id":row.id,
            "teacher_id":row.teacher_id,
            "teacher_name":teachers[row.teacher_id].name if row.teacher_id in teachers else "?",
            "preferred_slots":preferred_slots,
            "unavailable_slots":unavailable_slots,
            "preferred_labels":[label(slot) for slot in preferred_slots],
            "unavailable_labels":[label(slot) for slot in unavailable_slots],
            "note":row.note,
            "status":row.status,
            "created_at":row.created_at,
            "reviewed_at":row.reviewed_at,
        })
    return items

@app.get("/api/projects/{pid}/preferences")
def list_teacher_preferences(pid:int,user:User=Depends(current_user),db:Session=Depends(db_session)):
    p=get_project(pid,user,db)
    return {"items":preference_payload(db,p),"submission_url":f"/preferences/{p.share_token}"}

class PreferenceReviewIn(BaseModel):
    action: str

@app.post("/api/projects/{pid}/preferences/{preference_id}/review")
def review_teacher_preference(pid:int,preference_id:int,payload:PreferenceReviewIn,user:User=Depends(current_user),db:Session=Depends(db_session)):
    p=get_project_for_update(pid,user,db)
    preference=db.get(TeacherPreference,preference_id)
    if not preference or preference.project_id!=pid: raise HTTPException(404)
    if payload.action not in {"accept","reject"}: raise HTTPException(400,"Thao tác không hợp lệ")
    if preference.status!="pending":
        raise HTTPException(409,"Nguyện vọng này đã được xử lý")
    if payload.action=="reject":
        preference.status="rejected"
        preference.reviewed_at=datetime.now().isoformat(timespec="seconds")
        db.commit()
        return {"ok":True,"message":"Đã từ chối nguyện vọng."}

    preferred=set(valid_slots(p,parse_slots(preference.preferred_json)))
    unavailable=set(valid_slots(p,parse_slots(preference.unavailable_json)))
    teacher_assignments=db.scalars(select(Assignment).where(
        Assignment.project_id==pid,
        Assignment.teacher_id==preference.teacher_id,
    )).all()
    assignment_ids={assignment.id for assignment in teacher_assignments}
    teacher_lessons=db.scalars(select(Lesson).where(
        Lesson.project_id==pid,
        Lesson.assignment_id.in_(assignment_ids),
    )).all() if assignment_ids else []

    original_slots=defaultdict(list)
    for lesson in teacher_lessons:
        original_slots[lesson.assignment_id].append(lesson.slot)

    locked_conflicts=[
        lesson for lesson in teacher_lessons
        if lesson.locked and lesson.slot in unavailable
    ]
    if locked_conflicts:
        return JSONResponse({
            "ok":False,
            "message":f"Không thể duyệt vì có {len(locked_conflicts)} tiết cố định nằm trong các tiết cần tránh. Hãy bỏ cố định hoặc sửa nguyện vọng trước.",
        },409)

    def preference_cost(slots):
        values=list(slots)
        preferred_count=sum(1 for slot in values if slot in preferred)
        return -preferred_count

    old_cost=preference_cost(lesson.slot for lesson in teacher_lessons)
    old_assignment_valid=all(
        len(original_slots[assignment.id])==assignment.periods_per_week
        and assignment_pattern_matches(p,assignment,original_slots[assignment.id])
        for assignment in teacher_assignments
    )
    old_hard_valid=all(lesson.slot not in unavailable for lesson in teacher_lessons)

    previous=db.scalars(select(TeacherPreference).where(
        TeacherPreference.project_id==pid,
        TeacherPreference.teacher_id==preference.teacher_id,
        TeacherPreference.status=="accepted",
        TeacherPreference.id!=preference.id,
    )).all()

    # Thử đồng thời cập nhật trạng thái và xếp lại lịch trong savepoint. Nếu
    # không thể thỏa ràng buộc cứng, toàn bộ trạng thái được khôi phục.
    schedule_attempt=db.begin_nested()
    try:
        for row in previous:
            row.status="superseded"
        preference.status="accepted"
        preference.reviewed_at=datetime.now().isoformat(timespec="seconds")
        db.flush()

        removable=[lesson for lesson in teacher_lessons if not lesson.locked]
        for lesson in removable:
            db.delete(lesson)
        db.flush()

        result=solve_missing(db,p,tries=240,target_assignment_ids=assignment_ids)
        locked_slots_by_assignment=defaultdict(list)
        for lesson in teacher_lessons:
            if lesson.locked:
                locked_slots_by_assignment[lesson.assignment_id].append(lesson.slot)
        result_slots=defaultdict(list)
        for assignment_id,slot,_locked in result["lessons"]:
            result_slots[assignment_id].append(slot)

        final_slots={
            assignment.id:[*locked_slots_by_assignment[assignment.id],*result_slots[assignment.id]]
            for assignment in teacher_assignments
        }
        new_assignment_valid=(
            result.get("unscheduled",0)==0
            and not result.get("invalid_assignments")
            and all(
                len(final_slots[assignment.id])==assignment.periods_per_week
                and assignment_pattern_matches(p,assignment,final_slots[assignment.id])
                and all(slot not in unavailable for slot in final_slots[assignment.id])
                for assignment in teacher_assignments
            )
        )
        new_cost=preference_cost(
            slot for assignment in teacher_assignments for slot in final_slots[assignment.id]
        )
        should_apply=new_assignment_valid and (
            not old_assignment_valid or not old_hard_valid or new_cost<old_cost
        )

        if should_apply:
            for aid,slot,locked in result["lessons"]:
                db.add(Lesson(project_id=pid,assignment_id=aid,slot=slot,locked=locked))
            db.flush()
            schedule_attempt.commit()
            db.commit()
            moved=added=removed_count=0
            for assignment in teacher_assignments:
                old_set=set(original_slots[assignment.id])
                new_set=set(final_slots[assignment.id])
                removed_here=len(old_set-new_set)
                added_here=len(new_set-old_set)
                moved_here=min(removed_here,added_here)
                moved+=moved_here
                added+=added_here-moved_here
                removed_count+=removed_here-moved_here
            changed=moved+added+removed_count
            details=[]
            if moved: details.append(f"di chuyển {moved}")
            if added: details.append(f"bổ sung {added}")
            if removed_count: details.append(f"loại bỏ {removed_count}")
            detail_text=f" ({', '.join(details)})" if details else ""
            return {
                "ok":True,
                "message":f"Đã duyệt nguyện vọng và điều chỉnh {changed} tiết{detail_text}. Các tiết cần tránh hiện là ràng buộc cứng.",
            }

        schedule_attempt.rollback()
        if old_assignment_valid and old_hard_valid:
            for row in previous:
                row.status="superseded"
            preference.status="accepted"
            preference.reviewed_at=datetime.now().isoformat(timespec="seconds")
            db.commit()
            return {
                "ok":True,
                "message":"Đã duyệt nguyện vọng. Lịch hiện tại đã thỏa toàn bộ tiết cần tránh nên không cần di chuyển tiết học.",
            }

        db.commit()
        return JSONResponse({
            "ok":False,
            "message":"Không thể duyệt nguyện vọng vì chưa tìm được lịch đầy đủ thỏa toàn bộ tiết cần tránh. Nguyện vọng vẫn ở trạng thái chờ duyệt.",
        },409)
    except Exception:
        if schedule_attempt.is_active:
            schedule_attempt.rollback()
        db.rollback()
        raise

@app.get("/projects/{pid}/export.csv", include_in_schema=False)
def export_csv_legacy(pid:int,user:User=Depends(current_user),db:Session=Depends(db_session)):
    get_project(pid,user,db)
    return RedirectResponse(f"/projects/{pid}/export.xlsx",303)

@app.get("/projects/{pid}/export.xlsx")
def export_excel(pid:int,user:User=Depends(current_user),db:Session=Depends(db_session)):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.worksheet.table import Table, TableStyleInfo

    p=get_project(pid,user,db); data=project_data(db,p)
    workbook=Workbook(); sheet=workbook.active; sheet.title="Thời khóa biểu"
    sheet.sheet_view.showGridLines=False
    sheet.merge_cells("A1:F1"); sheet["A1"]=p.name
    sheet.merge_cells("A2:F2"); sheet["A2"]=p.school_name
    sheet["A1"].font=Font(name="Aptos Display",size=18,bold=True,color="FFFFFF")
    sheet["A1"].fill=PatternFill("solid",fgColor="1D4ED8")
    sheet["A1"].alignment=Alignment(horizontal="left",vertical="center")
    sheet["A2"].font=Font(name="Aptos",size=11,color="475467")
    sheet["A2"].alignment=Alignment(horizontal="left",vertical="center")
    sheet.row_dimensions[1].height=30; sheet.row_dimensions[2].height=22

    headers=["Lớp","Thứ","Buổi","Tiết","Môn","Giáo viên"]
    sheet.append([]); sheet.append(headers)
    header_fill=PatternFill("solid",fgColor="DBEAFE")
    header_border=Border(bottom=Side(style="thin",color="93C5FD"))
    for cell in sheet[4]:
        cell.font=Font(name="Aptos",bold=True,color="1E3A8A")
        cell.fill=header_fill; cell.border=header_border
        cell.alignment=Alignment(vertical="center")
    sheet.row_dimensions[4].height=24

    assignments={item["id"]:item for item in data["assignments"]}
    rows=[]
    for lesson in data["lessons"]:
        assignment=assignments.get(lesson["assignment_id"])
        if not assignment:
            continue
        day,session,period=slot_meta(p,lesson["slot"])
        rows.append((
            lesson["slot"],assignment["class_name"],DAYS[day],
            "Cả buổi" if p.sessions==1 else ("Sáng" if session==0 else "Chiều"),period+1,
            assignment["subject_name"],assignment["teacher_name"],
        ))
    rows.sort(key=lambda row:(row[0],row[1]))
    for _,class_name,day_name,session_name,period,subject_name,teacher_name in rows:
        sheet.append([class_name,day_name,session_name,period,subject_name,teacher_name])

    last_row=4+len(rows)
    if rows:
        table=Table(displayName="ThoiKhoaBieu",ref=f"A4:F{last_row}")
        table.tableStyleInfo=TableStyleInfo(
            name="TableStyleMedium2",showFirstColumn=False,showLastColumn=False,
            showRowStripes=True,showColumnStripes=False,
        )
        sheet.add_table(table)
    else:
        sheet.merge_cells("A5:F5"); sheet["A5"]="Chưa có tiết học nào được xếp."
        sheet["A5"].font=Font(name="Aptos",italic=True,color="667085")
        sheet["A5"].alignment=Alignment(horizontal="center")
        sheet.auto_filter.ref="A4:F4"

    widths={"A":16,"B":14,"C":12,"D":10,"E":24,"F":24}
    for column,width in widths.items(): sheet.column_dimensions[column].width=width
    sheet.freeze_panes="A5"; sheet.auto_filter.ref=f"A4:F{last_row}"
    sheet.print_title_rows="1:4"; sheet.page_setup.orientation="landscape"
    sheet.page_setup.fitToWidth=1; sheet.sheet_properties.pageSetUpPr.fitToPage=True
    sheet.print_area=f"A1:F{max(last_row,5)}"
    sheet["D5" if rows else "D4"].number_format="0"

    output=io.BytesIO(); workbook.save(output); output.seek(0)
    encoded_filename=quote(f"{p.name}.xlsx",safe="")
    disposition=f"attachment; filename=thoi-khoa-bieu.xlsx; filename*=UTF-8''{encoded_filename}"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition":disposition},
    )

def project_data(db:Session,p:Project):
    deps=db.scalars(select(Department).where(Department.project_id==p.id)).all()
    subs=db.scalars(select(Subject).where(Subject.project_id==p.id)).all()
    teas=db.scalars(select(Teacher).where(Teacher.project_id==p.id)).all()
    grades=db.scalars(select(Grade).where(Grade.project_id==p.id)).all()
    classes=db.scalars(select(SchoolClass).where(SchoolClass.project_id==p.id)).all()
    assignments=db.scalars(select(Assignment).where(Assignment.project_id==p.id)).all()
    lessons=db.scalars(select(Lesson).where(Lesson.project_id==p.id)).all()
    sm={x.id:x for x in subs};tm={x.id:x for x in teas};cm={x.id:x for x in classes}
    assigned_teacher_ids={x.teacher_id for x in assignments};assigned_subject_ids={x.subject_id for x in assignments};assigned_class_ids={x.class_id for x in assignments}
    return {
      "project":{"id":p.id,"name":p.name,"school_name":p.school_name,"days":p.days,"sessions":p.sessions,"periods":p.periods_per_session,"share_token":p.share_token,"blocked_slots":valid_slots(p,parse_slots(p.blocked_slots_json))},
      "departments":[{"id":x.id,"name":x.name} for x in deps],
      "subjects":[{"id":x.id,"name":x.name,"short_name":x.short_name,"max_consecutive":x.max_consecutive} for x in subs],
      "teachers":[{"id":x.id,"name":x.name,"short_name":x.short_name,"department_id":x.department_id,"max_periods_day":x.max_periods_day,"unavailable":list(parse_slots(x.unavailable_json))} for x in teas],
      "grades":[{"id":x.id,"name":x.name} for x in grades],
      "classes":[{"id":x.id,"name":x.name,"grade_id":x.grade_id,"unavailable":list(parse_slots(x.unavailable_json))} for x in classes],
      "assignments":[{"id":x.id,"class_id":x.class_id,"subject_id":x.subject_id,"teacher_id":x.teacher_id,"periods_per_week":x.periods_per_week,"block_mode":x.block_mode,"class_name":cm.get(x.class_id).name if cm.get(x.class_id) else "?","subject_name":sm.get(x.subject_id).name if sm.get(x.subject_id) else "?","subject_short":sm.get(x.subject_id).short_name if sm.get(x.subject_id) else "?","teacher_name":tm.get(x.teacher_id).name if tm.get(x.teacher_id) else "?","teacher_short":tm.get(x.teacher_id).short_name if tm.get(x.teacher_id) else "?"} for x in assignments],
      "lessons":[{"id":x.id,"assignment_id":x.assignment_id,"slot":x.slot,"locked":x.locked} for x in lessons],
      "coverage":{
        "unassigned_teachers":[{"id":x.id,"name":x.name} for x in teas if x.id not in assigned_teacher_ids],
        "unassigned_subjects":[{"id":x.id,"name":x.name} for x in subs if x.id not in assigned_subject_ids],
        "unassigned_classes":[{"id":x.id,"name":x.name} for x in classes if x.id not in assigned_class_ids],
      }
    }

def public_project_data(db:Session,p:Project):
    subjects={x.id:x for x in db.scalars(select(Subject).where(Subject.project_id==p.id)).all()}
    teachers={x.id:x for x in db.scalars(select(Teacher).where(Teacher.project_id==p.id)).all()}
    classes={x.id:x for x in db.scalars(select(SchoolClass).where(SchoolClass.project_id==p.id)).all()}
    assignments=db.scalars(select(Assignment).where(Assignment.project_id==p.id)).all()
    lessons=db.scalars(select(Lesson).where(Lesson.project_id==p.id)).all()
    return {
        "project":{
            "id":p.id,"name":p.name,"school_name":p.school_name,"days":p.days,
            "sessions":p.sessions,"periods":p.periods_per_session,
        },
        "classes":[{"id":item.id,"name":item.name} for item in classes.values()],
        "teachers":[{"id":item.id,"name":item.name,"short_name":item.short_name} for item in teachers.values()],
        "subjects":[{"id":item.id,"name":item.name,"short_name":item.short_name} for item in subjects.values()],
        "assignments":[{
            "id":item.id,"class_id":item.class_id,"subject_id":item.subject_id,"teacher_id":item.teacher_id,
            "periods_per_week":item.periods_per_week,"block_mode":item.block_mode,
            "class_name":classes[item.class_id].name if item.class_id in classes else "?",
            "subject_name":subjects[item.subject_id].name if item.subject_id in subjects else "?",
            "subject_short":subjects[item.subject_id].short_name if item.subject_id in subjects else "?",
            "teacher_name":teachers[item.teacher_id].name if item.teacher_id in teachers else "?",
            "teacher_short":teachers[item.teacher_id].short_name if item.teacher_id in teachers else "?",
        } for item in assignments],
        "lessons":[{"id":item.id,"assignment_id":item.assignment_id,"slot":item.slot} for item in lessons],
    }

def ga_schedule(db:Session,p:Project,mode:str,tries:int=120,target_assignment_ids:Optional[set[int]]=None):
    assignments=db.scalars(select(Assignment).where(Assignment.project_id==p.id)).all()
    teachers={x.id:x for x in db.scalars(select(Teacher).where(Teacher.project_id==p.id))}
    classes={x.id:x for x in db.scalars(select(SchoolClass).where(SchoolClass.project_id==p.id))}
    subjects={x.id:x for x in db.scalars(select(Subject).where(Subject.project_id==p.id))}
    all_existing=db.scalars(select(Lesson).where(Lesson.project_id==p.id)).all()
    if mode=="missing":
        existing=list(all_existing)
    elif mode=="rebuild":
        existing=[lesson for lesson in all_existing if lesson.locked]
    elif mode=="full":
        existing=[]
    else:
        raise ValueError(f"Chế độ xếp lịch không hợp lệ: {mode}")
    existing_counts=Counter(x.assignment_id for x in existing)
    existing_slots=defaultdict(set)
    locked_slots=defaultdict(set)
    for lesson in existing:
        existing_slots[lesson.assignment_id].add(lesson.slot)
        if lesson.locked:
            locked_slots[lesson.assignment_id].add(lesson.slot)
    fixed_rows_by_assignment=defaultdict(list)
    for row in db.scalars(select(FixedLesson).where(FixedLesson.project_id==p.id)).all():
        assignment=next((item for item in assignments if item.id==row.assignment_id),None)
        if assignment:
            size=fixed_row_size(p,assignment,row,[lesson for lesson in all_existing if lesson.assignment_id==assignment.id])
            fixed_rows_by_assignment[row.assignment_id].append((row.slot,size))
    requested_preferred,requested_unavailable=accepted_teacher_preferences(db,p.id)
    global_blocked=parse_slots(p.blocked_slots_json)
    slots=all_slots(p)
    ppd=p.sessions*p.periods_per_session
    if not assignments:
        return {"lessons":[],"unscheduled":0,"score":0}

    task_rows=[]
    invalid_assignment_ids=[]
    for assignment in assignments:
        if target_assignment_ids is not None and assignment.id not in target_assignment_ids:
            continue
        current_slots=existing_slots[assignment.id] if mode in {"missing","rebuild"} else set()
        plan=pattern_completion_plan(p,assignment,current_slots)
        if plan is None:
            invalid_assignment_ids.append(assignment.id)
            continue
        pending=[dict(item) for item in plan]
        task_index=0
        for fixed_slot,fixed_size in fixed_rows_by_assignment[assignment.id]:
            expected=set(range(fixed_slot,fixed_slot+fixed_size))
            if expected.issubset(locked_slots[assignment.id]):
                continue
            match_index=next((
                index for index,item in enumerate(pending)
                if item["size"]==fixed_size and not item["anchor_slots"]
            ),None)
            if match_index is None:
                continue
            pending.pop(match_index)
            task_rows.append((assignment,task_index,fixed_size,assignment_requires_double(assignment),fixed_slot,tuple(),(fixed_slot,)))
            task_index+=1
        for item in pending:
            task_rows.append((
                assignment,
                task_index,
                item["size"],
                assignment_requires_double(assignment),
                None,
                tuple(item["anchor_slots"]),
                item["candidate_starts"],
            ))
            task_index+=1

    if invalid_assignment_ids:
        missing=sum(
            max(0,assignment.periods_per_week-len(existing_slots[assignment.id]))
            for assignment in assignments if assignment.id in invalid_assignment_ids
        )
        return {
            "lessons":[],"unscheduled":missing,"score":missing*10000,
            "invalid_assignments":invalid_assignment_ids,
        }

    if not task_rows:
        return {"lessons":[],"unscheduled":0,"score":0}

    random.shuffle(task_rows)
    task_rows.sort(key=lambda task:(
        1 if task[4] is not None else 0,
        task[2]-len(task[5]),
        len(parse_slots(teachers[task[0].teacher_id].unavailable_json))+len(parse_slots(classes[task[0].class_id].unavailable_json)),
        -existing_counts[task[0].id],
    ),reverse=True)

    def valid_start_slots(size:int):
        return [slot for slot in slots if (slot % ppd) % p.periods_per_session + size <= p.periods_per_session]

    starts_by_size={}
    def start_pool(size:int):
        pool=starts_by_size.get(size)
        if pool is None:
            pool=valid_start_slots(size)
            starts_by_size[size]=pool
        return pool

    def evaluate(genes:list[int|None]):
        teacher_busy=defaultdict(set)
        class_busy=defaultdict(set)
        assignment_busy=defaultdict(set)
        teacher_day=Counter()
        class_sub_day=Counter()
        class_sub_slots=defaultdict(set)
        placed=[]
        chosen_starts=[None]*len(task_rows)
        unscheduled=0
        gene_miss=0.0
        preference_score=0.0

        for lesson in existing:
            assignment=next((x for x in assignments if x.id==lesson.assignment_id),None)
            if not assignment:
                continue
            day=lesson.slot//ppd
            teacher_busy[assignment.teacher_id].add(lesson.slot)
            class_busy[assignment.class_id].add(lesson.slot)
            assignment_busy[assignment.id].add(lesson.slot)
            teacher_day[(assignment.teacher_id,day)]+=1
            class_sub_day[(assignment.class_id,assignment.subject_id,day)]+=1
            class_sub_slots[(assignment.class_id,assignment.subject_id,day)].add(lesson.slot%ppd)

        for index,task in enumerate(task_rows):
            assignment,group_index,size,explicit,forced,anchor_slots,planned_starts=task
            anchor=set(anchor_slots)
            missing_size=size-len(anchor)
            gene=forced if forced is not None else genes[index]
            if forced is not None:
                candidate_pool=[forced]
            elif planned_starts is not None:
                candidate_pool=list(planned_starts)
            else:
                candidate_pool=start_pool(size)
            tu=parse_slots(teachers[assignment.teacher_id].unavailable_json)
            requested_avoid=requested_unavailable.get(assignment.teacher_id,set())
            preferred=requested_preferred.get(assignment.teacher_id,set())
            cu=parse_slots(classes[assignment.class_id].unavailable_json)

            def soft_preference_score(candidate_slots):
                value=0.0
                if preferred:
                    value+=sum(-4 if candidate in preferred else 1.5 for candidate in candidate_slots)
                return value
            best_slot=None
            best_score=None
            for slot in candidate_pool:
                if slot is None or slot not in slots:
                    continue
                day=slot//ppd
                position=slot%ppd
                session=position//p.periods_per_session
                period=position%p.periods_per_session
                if period+size>p.periods_per_session:
                    continue
                group_slots=list(range(slot,slot+size))
                group_set=set(group_slots)
                if anchor and not anchor.issubset(group_set):
                    continue
                new_slots=[candidate for candidate in group_slots if candidate not in anchor]
                if len(new_slots)!=missing_size:
                    continue
                if any(candidate//ppd!=day or (candidate%ppd)//p.periods_per_session!=session for candidate in group_slots):
                    continue
                if any(candidate in global_blocked or candidate in tu or candidate in requested_avoid or candidate in cu or candidate in teacher_busy[assignment.teacher_id] or candidate in class_busy[assignment.class_id] for candidate in new_slots):
                    continue
                if teacher_day[(assignment.teacher_id,day)]+missing_size>teachers[assignment.teacher_id].max_periods_day:
                    continue
                if explicit:
                    neighbors=[]
                    if period>0:
                        neighbors.append(slot-1)
                    if period+size<p.periods_per_session:
                        neighbors.append(slot+size)
                    if any(neighbor in assignment_busy[assignment.id] for neighbor in neighbors):
                        continue
                existing_periods=[
                    candidate%p.periods_per_session
                    for candidate in class_sub_slots[(assignment.class_id,assignment.subject_id,day)]
                    if candidate//p.periods_per_session==session
                ]
                run=sorted(set(existing_periods+list(range(period,period+size))))
                longest=current=1
                for left,right in zip(run,run[1:]):
                    current=current+1 if right==left+1 else 1
                    longest=max(longest,current)
                if longest>subjects[assignment.subject_id].max_consecutive:
                    continue
                score=class_sub_day[(assignment.class_id,assignment.subject_id,day)]*8+sum((candidate%p.periods_per_session)*0.15 for candidate in new_slots)
                score+=soft_preference_score(new_slots)
                neighbors=[]
                if period>0:
                    neighbors.append(slot-1)
                if period+size<p.periods_per_session:
                    neighbors.append(slot+size)
                for neighbor in neighbors:
                    if neighbor in teacher_busy[assignment.teacher_id]:
                        score-=1.2
                    if assignment_prefers_double(assignment) and neighbor in assignment_busy[assignment.id]:
                        score-=7.0
                if gene is not None:
                    if slot==gene:
                        score-=8
                    else:
                        score+=abs(slot-gene)*0.05
                if best_score is None or score<best_score:
                    best_score=score
                    best_slot=slot
            if best_slot is None:
                unscheduled+=missing_size
                continue
            chosen_starts[index]=best_slot
            day=best_slot//ppd
            gene_value=forced if forced is not None else genes[index]
            if gene_value is not None and best_slot!=gene_value:
                gene_miss+=abs(best_slot-gene_value)
            for slot in range(best_slot,best_slot+size):
                if slot in anchor:
                    continue
                teacher_busy[assignment.teacher_id].add(slot)
                class_busy[assignment.class_id].add(slot)
                assignment_busy[assignment.id].add(slot)
                class_sub_slots[(assignment.class_id,assignment.subject_id,day)].add(slot%ppd)
                placed.append((assignment.id,slot,forced is not None))
            teacher_day[(assignment.teacher_id,day)]+=missing_size
            class_sub_day[(assignment.class_id,assignment.subject_id,day)]+=missing_size
            preference_score+=soft_preference_score(
                [slot for slot in range(best_slot,best_slot+size) if slot not in anchor]
            )

        score=unscheduled*10000+gene_miss*0.05+preference_score
        for (cid,sid,day),n in class_sub_day.items():
            score+=max(0,n-1)*10
        for tid,busy in teacher_busy.items():
            for day in range(p.days):
                xs=sorted(slot%ppd for slot in busy if slot//ppd==day)
                if xs:
                    score+=(xs[-1]-xs[0]+1-len(xs))*2
        for assignment in assignments:
            if not assignment_prefers_double(assignment):
                continue
            runs=assignment_run_groups(p,assignment_busy[assignment.id])
            formed_pairs=sum(run["size"]//2 for run in runs)
            target_pairs=assignment.periods_per_week//2
            score+=max(0,target_pairs-formed_pairs)*14
        return {"lessons":placed,"unscheduled":unscheduled,"score":round(score,2),"genes":chosen_starts}

    def genes_from_candidate(candidate):
        genes=list(candidate.get("genes",[]))
        if len(genes)!=len(task_rows):
            genes=[None]*len(task_rows)
        return genes

    def random_gene(task):
        assignment,group_index,size,explicit,forced,anchor_slots,planned_starts=task
        if forced is not None:
            return forced
        pool=list(planned_starts) if planned_starts is not None else start_pool(size)
        return random.choice(pool) if pool else None

    def mutate(genes):
        child=genes[:]
        for index,task in enumerate(task_rows):
            assignment,group_index,size,explicit,forced,anchor_slots,planned_starts=task
            if forced is not None:
                child[index]=forced
                continue
            if random.random()<0.15:
                child[index]=random_gene(task) if random.random()<0.9 else None
        return child

    def crossover(left,right):
        child=[]
        for index,task in enumerate(task_rows):
            assignment,group_index,size,explicit,forced,anchor_slots,planned_starts=task
            if forced is not None:
                child.append(forced)
            elif random.random()<0.5:
                child.append(left[index])
            else:
                child.append(right[index])
        return child

    def exact_fallback(node_limit:int):
        """Thử tìm lời giải đầy đủ bằng backtracking cho bài toán vừa/nhỏ.

        GA vẫn được dùng trước để có tốc độ tốt. Khi GA bỏ sót lời giải, bước
        này duyệt có hệ thống và có thể chứng minh vô nghiệm nếu hoàn tất toàn
        bộ cây tìm kiếm trước giới hạn nút.
        """
        teacher_busy=defaultdict(set)
        class_busy=defaultdict(set)
        assignment_busy=defaultdict(set)
        teacher_day=Counter()
        class_sub_day=Counter()
        class_sub_slots=defaultdict(set)
        for lesson in existing:
            assignment=next((x for x in assignments if x.id==lesson.assignment_id),None)
            if not assignment:
                continue
            day=lesson.slot//ppd
            teacher_busy[assignment.teacher_id].add(lesson.slot)
            class_busy[assignment.class_id].add(lesson.slot)
            assignment_busy[assignment.id].add(lesson.slot)
            teacher_day[(assignment.teacher_id,day)]+=1
            class_sub_day[(assignment.class_id,assignment.subject_id,day)]+=1
            class_sub_slots[(assignment.class_id,assignment.subject_id,day)].add(lesson.slot%ppd)

        raw_candidates=[]
        for task in task_rows:
            assignment,group_index,size,explicit,forced,anchor_slots,planned_starts=task
            if forced is not None:
                pool=[forced]
            elif planned_starts is not None:
                pool=list(planned_starts)
            else:
                pool=start_pool(size)
            raw_candidates.append([slot for slot in pool if slot is not None])

        placed_by_task={}
        remaining=set(range(len(task_rows)))
        nodes=0
        limit_hit=False

        def options_for(index:int):
            assignment,group_index,size,explicit,forced,anchor_slots,planned_starts=task_rows[index]
            anchor=set(anchor_slots)
            missing_size=size-len(anchor)
            tu=parse_slots(teachers[assignment.teacher_id].unavailable_json)
            requested_avoid=requested_unavailable.get(assignment.teacher_id,set())
            preferred=requested_preferred.get(assignment.teacher_id,set())
            cu=parse_slots(classes[assignment.class_id].unavailable_json)
            options=[]
            for slot in raw_candidates[index]:
                if slot not in slots:
                    continue
                day=slot//ppd
                position=slot%ppd
                session=position//p.periods_per_session
                period=position%p.periods_per_session
                if period+size>p.periods_per_session:
                    continue
                group_slots=tuple(range(slot,slot+size))
                group_set=set(group_slots)
                if anchor and not anchor.issubset(group_set):
                    continue
                new_slots=tuple(candidate for candidate in group_slots if candidate not in anchor)
                if len(new_slots)!=missing_size:
                    continue
                if any(candidate//ppd!=day or (candidate%ppd)//p.periods_per_session!=session for candidate in group_slots):
                    continue
                if any(candidate in global_blocked or candidate in tu or candidate in requested_avoid or candidate in cu or candidate in teacher_busy[assignment.teacher_id] or candidate in class_busy[assignment.class_id] for candidate in new_slots):
                    continue
                if teacher_day[(assignment.teacher_id,day)]+missing_size>teachers[assignment.teacher_id].max_periods_day:
                    continue
                if explicit:
                    neighbors=[]
                    if period>0:
                        neighbors.append(slot-1)
                    if period+size<p.periods_per_session:
                        neighbors.append(slot+size)
                    if any(neighbor in assignment_busy[assignment.id] for neighbor in neighbors):
                        continue
                existing_periods=[
                    candidate%p.periods_per_session
                    for candidate in class_sub_slots[(assignment.class_id,assignment.subject_id,day)]
                    if candidate//p.periods_per_session==session
                ]
                run=sorted(set(existing_periods+list(range(period,period+size))))
                longest=current=1
                for left,right in zip(run,run[1:]):
                    current=current+1 if right==left+1 else 1
                    longest=max(longest,current)
                if longest>subjects[assignment.subject_id].max_consecutive:
                    continue
                preference_cost=sum(-4 if candidate in preferred else 1.5 for candidate in new_slots) if preferred else 0
                adjacent_same=0
                if assignment_prefers_double(assignment):
                    if period>0 and slot-1 in assignment_busy[assignment.id]:
                        adjacent_same+=1
                    if period+size<p.periods_per_session and slot+size in assignment_busy[assignment.id]:
                        adjacent_same+=1
                score=(
                    class_sub_day[(assignment.class_id,assignment.subject_id,day)]*8
                    +sum((candidate%p.periods_per_session)*0.15 for candidate in new_slots)
                    +preference_cost
                    -adjacent_same*7
                )
                options.append((score,slot,group_slots,new_slots,day))
            options.sort(key=lambda item:(item[0],item[1]))
            return options

        def apply(index:int,option):
            _score,slot,group_slots,new_slots,day=option
            assignment,group_index,size,explicit,forced,anchor_slots,planned_starts=task_rows[index]
            for candidate in new_slots:
                teacher_busy[assignment.teacher_id].add(candidate)
                class_busy[assignment.class_id].add(candidate)
                assignment_busy[assignment.id].add(candidate)
                class_sub_slots[(assignment.class_id,assignment.subject_id,day)].add(candidate%ppd)
            teacher_day[(assignment.teacher_id,day)]+=len(new_slots)
            class_sub_day[(assignment.class_id,assignment.subject_id,day)]+=len(new_slots)
            placed_by_task[index]=(assignment.id,tuple(new_slots),forced is not None)

        def undo(index:int,option):
            _score,slot,group_slots,new_slots,day=option
            assignment,group_index,size,explicit,forced,anchor_slots,planned_starts=task_rows[index]
            placed_by_task.pop(index,None)
            teacher_day[(assignment.teacher_id,day)]-=len(new_slots)
            class_sub_day[(assignment.class_id,assignment.subject_id,day)]-=len(new_slots)
            for candidate in new_slots:
                teacher_busy[assignment.teacher_id].remove(candidate)
                class_busy[assignment.class_id].remove(candidate)
                assignment_busy[assignment.id].remove(candidate)
                class_sub_slots[(assignment.class_id,assignment.subject_id,day)].remove(candidate%ppd)

        def search():
            nonlocal nodes,limit_hit
            if not remaining:
                return True
            nodes+=1
            if nodes>node_limit:
                limit_hit=True
                return False
            selected_index=None
            selected_options=None
            for index in tuple(remaining):
                options=options_for(index)
                if not options:
                    return False
                if selected_options is None or len(options)<len(selected_options):
                    selected_index=index
                    selected_options=options
                    if len(options)==1:
                        break
            remaining.remove(selected_index)
            for option in selected_options:
                apply(selected_index,option)
                if search():
                    return True
                undo(selected_index,option)
                if limit_hit:
                    break
            remaining.add(selected_index)
            return False

        solved=search()
        if not solved:
            return None,not limit_hit,nodes
        lessons=[]
        for index in range(len(task_rows)):
            assignment_id,new_slots,locked=placed_by_task[index]
            lessons.extend((assignment_id,slot,locked) for slot in new_slots)
        preference_score=0.0
        final_slots=defaultdict(set)
        for lesson in existing:
            final_slots[lesson.assignment_id].add(lesson.slot)
        for assignment_id,slot,_locked in lessons:
            final_slots[assignment_id].add(slot)
            assignment=next(item for item in assignments if item.id==assignment_id)
            preferred=requested_preferred.get(assignment.teacher_id,set())
            if preferred:
                preference_score+=-4 if slot in preferred else 1.5
        for assignment in assignments:
            if assignment_prefers_double(assignment):
                formed_pairs=sum(run["size"]//2 for run in assignment_run_groups(p,final_slots[assignment.id]))
                preference_score+=max(0,assignment.periods_per_week//2-formed_pairs)*14
        return {
            "lessons":lessons,
            "unscheduled":0,
            "score":round(preference_score,2),
            "exact":True,
        },True,nodes

    population_size=max(18,min(40,max(12,len(task_rows))))
    generations=max(12,min(60,max(tries//3,18)))
    elite_count=max(2,population_size//5)

    seed_candidate=evaluate([None]*len(task_rows))
    best_candidate=seed_candidate
    population=[genes_from_candidate(seed_candidate)]
    for _ in range(population_size-1):
        population.append([random_gene(task) for task in task_rows])

    evaluated=[]
    for genes in population:
        candidate=evaluate(genes)
        evaluated.append((candidate,genes))
        if candidate["score"]<best_candidate["score"]:
            best_candidate=candidate

    for _ in range(generations):
        evaluated.sort(key=lambda item:(item[0]["score"],item[0]["unscheduled"]))
        elites=[genes for _candidate,genes in evaluated[:elite_count]]
        if evaluated[0][0]["score"]<best_candidate["score"]:
            best_candidate=evaluated[0][0]
        next_population=elites[:]
        while len(next_population)<population_size:
            pool=evaluated[:max(6,population_size//2)]
            parent1=random.choice(pool)[1]
            parent2=random.choice(pool)[1]
            child=mutate(crossover(parent1,parent2))
            next_population.append(child)
        evaluated=[]
        for genes in next_population:
            candidate=evaluate(genes)
            evaluated.append((candidate,genes))
            if candidate["score"]<best_candidate["score"]:
                best_candidate=candidate

    if evaluated:
        evaluated.sort(key=lambda item:(item[0]["score"],item[0]["unscheduled"]))
        if evaluated[0][0]["score"]<best_candidate["score"]:
            best_candidate=evaluated[0][0]

    if best_candidate["unscheduled"]>0:
        missing_periods=sum(task[2]-len(task[5]) for task in task_rows)
        exact_allowed=len(task_rows)<=30 and missing_periods<=36
        if exact_allowed:
            exact_result,exhausted,nodes=exact_fallback(max(50000,tries*1500))
            if exact_result is not None:
                return exact_result
            best_candidate["proven_infeasible"]=exhausted
            best_candidate["search_limited"]=not exhausted
            best_candidate["exact_nodes"]=nodes
        else:
            best_candidate["proven_infeasible"]=False
            best_candidate["search_limited"]=True
    return best_candidate

def solve_missing(db:Session,p:Project,tries=120,target_assignment_ids:Optional[set[int]]=None):
    """Giữ nguyên lịch hiện có và chỉ xếp phần còn thiếu của các phân công đích."""
    return ga_schedule(
        db,p,mode="missing",tries=tries,
        target_assignment_ids=target_assignment_ids,
    )

def solve_rebuild(db:Session,p:Project,tries=220):
    """Giữ tiết cố định, xếp lại toàn bộ phần còn lại."""
    return ga_schedule(db,p,mode="rebuild",tries=tries)


def solve(db:Session,p:Project,tries=80):
    return ga_schedule(db,p,mode="full",tries=tries)

def seed_project(db:Session,p:Project):
    d1=Department(project_id=p.id,name="Tổ Toán - Tin");d2=Department(project_id=p.id,name="Tổ Ngữ văn")
    db.add_all([d1,d2]);db.flush()
    s1=Subject(project_id=p.id,name="Toán",short_name="TOÁN",max_consecutive=2);s2=Subject(project_id=p.id,name="Ngữ văn",short_name="VĂN",max_consecutive=2);s3=Subject(project_id=p.id,name="Tin học",short_name="TIN",max_consecutive=1)
    db.add_all([s1,s2,s3]);db.flush()
    t1=Teacher(project_id=p.id,department_id=d1.id,name="Nguyễn Văn An",short_name="An",max_periods_day=4);t2=Teacher(project_id=p.id,department_id=d2.id,name="Trần Thị Bình",short_name="Bình",max_periods_day=4);t3=Teacher(project_id=p.id,department_id=d1.id,name="Lê Minh Châu",short_name="Châu",max_periods_day=4)
    db.add_all([t1,t2,t3]);db.flush()
    g=Grade(project_id=p.id,name="Khối 10");db.add(g);db.flush()
    c1=SchoolClass(project_id=p.id,grade_id=g.id,name="10A1");c2=SchoolClass(project_id=p.id,grade_id=g.id,name="10A2");db.add_all([c1,c2]);db.flush()
    for c in [c1,c2]:
        db.add_all([Assignment(project_id=p.id,class_id=c.id,subject_id=s1.id,teacher_id=t1.id,periods_per_week=4),Assignment(project_id=p.id,class_id=c.id,subject_id=s2.id,teacher_id=t2.id,periods_per_week=3),Assignment(project_id=p.id,class_id=c.id,subject_id=s3.id,teacher_id=t3.id,periods_per_week=2)])
    db.commit()

def ensure_demo():
    db=SessionLocal()
    try:
        # Chỉ bỏ qua bootstrap khi hệ thống thực sự đã có quản trị viên.
        # Trước đây chỉ cần tồn tại một user (kể cả pending/teacher) là hàm
        # dừng, khiến database không có admin không thể tự phục hồi.
        if db.scalar(select(User.id).where(User.role == "admin").limit(1)) is not None:
            return
        if not BOOTSTRAP_ADMIN_EMAIL or len(BOOTSTRAP_ADMIN_PASSWORD) < 8:
            raise RuntimeError(
                "Database chưa có quản trị viên. Hãy cấu hình BOOTSTRAP_ADMIN_EMAIL và "
                "BOOTSTRAP_ADMIN_PASSWORD (ít nhất 8 ký tự) trong .env để khôi phục quyền quản trị."
            )

        # Nếu email bootstrap đã tồn tại, nâng chính tài khoản đó lên admin
        # thay vì tạo user trùng email. Áp dụng mật khẩu bootstrap để chủ hệ
        # thống chắc chắn có thể đăng nhập lại và tăng session_version để vô
        # hiệu hóa các phiên cũ của tài khoản vừa được khôi phục.
        user = db.scalar(
            select(User)
            .where(func.lower(User.email) == BOOTSTRAP_ADMIN_EMAIL)
            .limit(1)
        )
        if user is None:
            user=User(
                email=BOOTSTRAP_ADMIN_EMAIL,
                name="Quản trị viên",
                password_hash=pwd.hash(BOOTSTRAP_ADMIN_PASSWORD),
                role="admin",
                is_superadmin=True,
            )
            db.add(user)
        else:
            user.role = "admin"
            user.is_superadmin = True
            user.password_hash = pwd.hash(BOOTSTRAP_ADMIN_PASSWORD)
            user.session_version = max(1, user.session_version or 1) + 1
            if not (user.name or "").strip():
                user.name = "Quản trị viên"

        db.commit()
        if SEED_DEMO_DATA and db.scalar(select(Project.id).limit(1)) is None:
            p=Project(owner_id=user.id,name="TKB học kỳ I",school_name="THPT Demo",days=6,sessions=2,periods_per_session=5)
            db.add(p);db.commit();seed_project(db,p)
    finally:
        db.close()
ensure_demo()
