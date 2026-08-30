from __future__ import annotations

import re
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def _natural_key(value: str):
    return [int(part) if part.isdigit() else part.casefold() for part in re.split(r"(\d+)", str(value or ""))]


def _safe_excel_text(value: object) -> str:
    text = str(value or "")
    # Prevent user-provided names from accidentally becoming Excel formulas.
    if text.startswith(("=", "+", "-", "@", "\t", "\r")):
        return "'" + text
    return text


def _day_label(day_index: int) -> str:
    return "CN" if day_index == 6 else str(day_index + 2)


def _session_label(day_index: int, session_index: int, session_count: int) -> str:
    day = _day_label(day_index)
    if session_count == 1:
        return f"Cả ngày {day}"
    if session_count == 2:
        return f"{'Sáng' if session_index == 0 else 'Chiều'} {day}"
    return f"Buổi {session_index + 1} - {day}"


def build_timetable_workbook(project, data: dict) -> Workbook:
    """Build a school timetable matrix similar to the supplied legacy Excel template."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Thời khóa biểu"
    sheet.sheet_view.showGridLines = False

    classes = sorted(data.get("classes", []), key=lambda item: _natural_key(item.get("name", "")))
    teachers = {item["id"]: item for item in data.get("teachers", [])}
    assignments = {item["id"]: item for item in data.get("assignments", [])}
    lessons = data.get("lessons", [])

    class_column = {item["id"]: 4 + index for index, item in enumerate(classes)}
    last_col = 4 + len(classes)
    last_col_letter = get_column_letter(last_col)

    thin = Side(style="thin", color="000000")
    medium = Side(style="medium", color="000000")
    table_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    separator_border = Border(top=medium, bottom=medium)
    white_fill = PatternFill(fill_type="solid", fgColor="FFFFFF")

    # Title area follows the structure of the provided sample: school name, then timetable title.
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    sheet.cell(1, 1, _safe_excel_text(project.school_name or "TRƯỜNG HỌC"))
    sheet.cell(1, 1).font = Font(name="Times New Roman", size=10, bold=True)
    sheet.cell(1, 1).alignment = Alignment(horizontal="left", vertical="center")
    sheet.row_dimensions[1].height = 18

    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_col)
    sheet.cell(2, 1, _safe_excel_text(project.name or "THỜI KHÓA BIỂU"))
    sheet.cell(2, 1).font = Font(name="Times New Roman", size=12, bold=True)
    sheet.cell(2, 1).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.row_dimensions[2].height = 32

    headers = ["Thứ", "Buổi", "Tiết"] + [_safe_excel_text(item["name"]) for item in classes] + ["GV nghỉ"]
    for col, value in enumerate(headers, start=1):
        cell = sheet.cell(3, col, value)
        cell.font = Font(name="Times New Roman", size=8, bold=True, italic=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = table_border
        cell.fill = white_fill
    sheet.row_dimensions[3].height = 24

    # Build a quick lookup from (slot, class) to the assignment taught in that slot.
    by_slot_class = {}
    busy_teachers = defaultdict(set)
    for lesson in lessons:
        assignment = assignments.get(lesson.get("assignment_id"))
        if not assignment:
            continue
        slot = int(lesson["slot"])
        class_id = assignment.get("class_id")
        teacher_id = assignment.get("teacher_id")
        if class_id is not None:
            by_slot_class[(slot, class_id)] = assignment
        if teacher_id is not None:
            busy_teachers[slot].add(teacher_id)

    active_teacher_ids = {
        item.get("teacher_id") for item in assignments.values()
        if item.get("teacher_id") in teachers
    }

    periods_per_day = project.sessions * project.periods_per_session
    current_row = 4
    for day in range(project.days):
        day_start = current_row
        for session in range(project.sessions):
            session_start = current_row
            for period in range(project.periods_per_session):
                slot = day * periods_per_day + session * project.periods_per_session + period
                sheet.cell(current_row, 3, period + 1)

                for class_item in classes:
                    assignment = by_slot_class.get((slot, class_item["id"]))
                    if not assignment:
                        continue
                    subject = str(assignment.get("subject_short") or assignment.get("subject_name") or "").strip()
                    teacher = str(assignment.get("teacher_short") or assignment.get("teacher_name") or "").strip()
                    lesson_text = " ".join(part for part in (subject, teacher) if part)
                    sheet.cell(current_row, class_column[class_item["id"]], _safe_excel_text(lesson_text))

                free_teacher_names = []
                for teacher_id in sorted(
                    active_teacher_ids - busy_teachers.get(slot, set()),
                    key=lambda tid: _natural_key(teachers[tid].get("short_name") or teachers[tid].get("name", "")),
                ):
                    teacher = teachers[teacher_id]
                    free_teacher_names.append(str(teacher.get("short_name") or teacher.get("name") or "").strip())
                sheet.cell(current_row, last_col, _safe_excel_text(", ".join(name for name in free_teacher_names if name)))

                for col in range(1, last_col + 1):
                    cell = sheet.cell(current_row, col)
                    cell.font = Font(name="Times New Roman", size=10)
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    cell.border = table_border
                    cell.fill = white_fill
                sheet.cell(current_row, last_col).font = Font(name="Times New Roman", size=9, bold=True)
                sheet.row_dimensions[current_row].height = 24
                current_row += 1

            session_end = current_row - 1
            if session_end >= session_start:
                sheet.merge_cells(start_row=session_start, start_column=2, end_row=session_end, end_column=2)
                session_cell = sheet.cell(session_start, 2, _session_label(day, session, project.sessions))
                session_cell.font = Font(name="Times New Roman", size=9)
                session_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                session_cell.border = table_border

            # The reference workbook leaves a thin blank separator row after each session.
            for col in range(1, last_col + 1):
                cell = sheet.cell(current_row, col)
                cell.border = separator_border
                cell.fill = white_fill
            sheet.row_dimensions[current_row].height = 8
            current_row += 1

        day_end = current_row - 1
        if day_end >= day_start:
            sheet.merge_cells(start_row=day_start, start_column=1, end_row=day_end, end_column=1)
            day_cell = sheet.cell(day_start, 1, _day_label(day))
            day_cell.font = Font(name="Times New Roman", size=9)
            day_cell.alignment = Alignment(horizontal="center", vertical="center")
            day_cell.border = table_border

    last_row = current_row - 1

    # Widths are based closely on the supplied .xls sample (BIFF column widths / 256).
    sheet.column_dimensions["A"].width = 3.6
    sheet.column_dimensions["B"].width = 6.2
    sheet.column_dimensions["C"].width = 3.6
    for col in range(4, last_col):
        sheet.column_dimensions[get_column_letter(col)].width = 15.7
    sheet.column_dimensions[last_col_letter].width = 37.8

    sheet.freeze_panes = "D4"
    sheet.print_title_rows = "1:3"
    sheet.print_area = f"A1:{last_col_letter}{last_row}"
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.page_setup.paperSize = sheet.PAPERSIZE_A3 if len(classes) > 10 else sheet.PAPERSIZE_A4
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_margins.left = 0.2
    sheet.page_margins.right = 0.2
    sheet.page_margins.top = 0.35
    sheet.page_margins.bottom = 0.35
    sheet.page_margins.header = 0.1
    sheet.page_margins.footer = 0.1
    sheet.sheet_properties.pageSetUpPr.autoPageBreaks = False

    return workbook
