from __future__ import annotations

import io
import zipfile

from openpyxl import Workbook

from app.schedule_audit import analyze_schedule_file, analyze_standalone_schedule_file, read_tables


def sample_context():
    project = {"days": 5, "sessions": 2, "periods": 5, "blocked_slots": []}
    classes = [
        {"id": 1, "name": "10A1", "unavailable": []},
        {"id": 2, "name": "10A2", "unavailable": []},
    ]
    subjects = [
        {"id": 1, "name": "Toán", "short_name": "TOAN", "max_consecutive": 2},
        {"id": 2, "name": "Văn", "short_name": "VAN", "max_consecutive": 2},
    ]
    teachers = [
        {"id": 1, "name": "Nguyễn An", "short_name": "AN", "max_periods_day": 5, "unavailable": []},
        {"id": 2, "name": "Trần Bình", "short_name": "BINH", "max_periods_day": 5, "unavailable": []},
    ]
    assignments = [
        {"id": 1, "class_id": 1, "subject_id": 1, "teacher_id": 1, "periods_per_week": 2, "block_mode": "free", "class_name": "10A1", "subject_name": "Toán"},
        {"id": 2, "class_id": 2, "subject_id": 2, "teacher_id": 2, "periods_per_week": 2, "block_mode": "free", "class_name": "10A2", "subject_name": "Văn"},
    ]
    return project, classes, subjects, teachers, assignments


def workbook_bytes(rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "Thời khóa biểu"
    for row in rows:
        ws.append(row)
    stream = io.BytesIO()
    wb.save(stream)
    return stream.getvalue()


def test_wide_excel_clean_schedule():
    project, classes, subjects, teachers, assignments = sample_context()
    content = workbook_bytes([
        ["TRƯỜNG"],
        ["TKB"],
        ["Thứ", "Buổi", "Tiết", "10A1", "10A2", "GV nghỉ"],
        ["2", "Sáng", 1, "TOAN AN", "", ""],
        ["", "", 2, "", "VAN BINH", ""],
        ["3", "Sáng", 1, "TOAN AN", "", ""],
        ["", "", 2, "", "VAN BINH", ""],
    ])
    report = analyze_schedule_file(
        filename="tkb.xlsx", content=content, project=project, classes=classes,
        subjects=subjects, teachers=teachers, assignments=assignments,
    )
    assert report["summary"]["recognized_lessons"] == 4
    assert report["summary"]["errors"] == 0
    assert report["status"] == "clean"


def test_detects_teacher_collision_and_missing_periods():
    project, classes, subjects, teachers, assignments = sample_context()
    assignments[1]["teacher_id"] = 1
    assignments[1]["subject_name"] = "Văn"
    content = workbook_bytes([
        ["Thứ", "Buổi", "Tiết", "10A1", "10A2"],
        ["2", "Sáng", 1, "TOAN AN", "VAN AN"],
        ["3", "Sáng", 1, "TOAN AN", ""],
    ])
    report = analyze_schedule_file(
        filename="tkb.xlsx", content=content, project=project, classes=classes,
        subjects=subjects, teachers=teachers, assignments=assignments,
    )
    codes = [item["code"] for item in report["issues"]]
    assert "teacher_collision" in codes
    assert "missing_lessons" in codes
    assert report["summary"]["missing_periods"] == 1


def test_reads_docx_table_xml():
    xml = '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
    <w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main"><w:body>
      <w:p><w:r><w:t>Thời khóa biểu</w:t></w:r></w:p>
      <w:tbl>
        <w:tr><w:tc><w:p><w:r><w:t>Thứ</w:t></w:r></w:p></w:tc><w:tc><w:p><w:r><w:t>Tiết</w:t></w:r></w:p></w:tc><w:tc><w:p><w:r><w:t>10A1</w:t></w:r></w:p></w:tc></w:tr>
        <w:tr><w:tc><w:p><w:r><w:t>2</w:t></w:r></w:p></w:tc><w:tc><w:p><w:r><w:t>1</w:t></w:r></w:p></w:tc><w:tc><w:p><w:r><w:t>TOAN AN</w:t></w:r></w:p></w:tc></w:tr>
      </w:tbl>
    </w:body></w:document>'''
    stream = io.BytesIO()
    with zipfile.ZipFile(stream, "w") as zf:
        zf.writestr("word/document.xml", xml)
    file_format, tables = read_tables("tkb.docx", stream.getvalue())
    assert file_format == "DOCX"
    assert tables[0][0] == "Thời khóa biểu"
    assert tables[0][1][1][2] == "TOAN AN"


def teacher_workbook_bytes(title: str, rows):
    wb = Workbook()
    ws = wb.active
    ws.title = title
    for row in rows:
        ws.append(row)
    stream = io.BytesIO()
    wb.save(stream)
    return stream.getvalue()


def test_teacher_sheet_day_grid_is_detected_and_scoped():
    project, classes, subjects, teachers, assignments = sample_context()
    content = teacher_workbook_bytes("Nguyễn An", [
        ["Tiết", "Thứ 2", "Thứ 3", "Thứ 4", "Thứ 5", "Thứ 6"],
        [1, "10A1 - Toán", "10A1 - Toán", "", "", ""],
    ])
    report = analyze_schedule_file(
        filename="tkb-gv.xlsx", content=content, project=project, classes=classes,
        subjects=subjects, teachers=teachers, assignments=assignments,
    )
    assert report["summary"]["recognized_lessons"] == 2
    assert report["summary"]["errors"] == 0
    assert report["detection"]["layouts"] == ["Theo giáo viên"]
    assert report["detection"]["teachers"] == ["Nguyễn An"]
    assert report["detection"]["scope_label"] == "Giáo viên: Nguyễn An"
    assert report["status"] == "clean"


def test_teacher_sheet_can_infer_subject_from_class_and_teacher():
    project, classes, subjects, teachers, assignments = sample_context()
    content = teacher_workbook_bytes("TKB GV Nguyễn An", [
        ["Tiết", "Thứ 2", "Thứ 3", "Thứ 4", "Thứ 5", "Thứ 6"],
        [1, "10A1", "10A1", "", "", ""],
    ])
    report = analyze_schedule_file(
        filename="tkb-gv.xlsx", content=content, project=project, classes=classes,
        subjects=subjects, teachers=teachers, assignments=assignments,
    )
    assert report["summary"]["recognized_lessons"] == 2
    assert report["summary"]["missing_periods"] == 0
    assert report["summary"]["errors"] == 0


def test_teacher_wide_table_is_detected():
    project, classes, subjects, teachers, assignments = sample_context()
    content = workbook_bytes([
        ["Thứ", "Buổi", "Tiết", "Nguyễn An", "Trần Bình"],
        ["2", "Sáng", 1, "10A1 - Toán", ""],
        ["", "", 2, "", "10A2 - Văn"],
        ["3", "Sáng", 1, "10A1 - Toán", ""],
        ["", "", 2, "", "10A2 - Văn"],
    ])
    report = analyze_schedule_file(
        filename="tkb-giao-vien.xlsx", content=content, project=project, classes=classes,
        subjects=subjects, teachers=teachers, assignments=assignments,
    )
    assert report["summary"]["recognized_lessons"] == 4
    assert report["summary"]["errors"] == 0
    assert report["detection"]["layouts"] == ["Bảng tổng hợp theo giáo viên"]
    assert report["detection"]["full_project"] is True


def test_class_sheet_day_grid_remains_supported():
    project, classes, subjects, teachers, assignments = sample_context()
    content = teacher_workbook_bytes("TKB lớp 10A1", [
        ["Tiết", "Thứ 2", "Thứ 3", "Thứ 4", "Thứ 5", "Thứ 6"],
        [1, "TOAN AN", "TOAN AN", "", "", ""],
    ])
    report = analyze_schedule_file(
        filename="tkb-lop.xlsx", content=content, project=project, classes=classes,
        subjects=subjects, teachers=teachers, assignments=assignments,
    )
    assert report["summary"]["recognized_lessons"] == 2
    assert report["summary"]["errors"] == 0
    assert report["detection"]["layouts"] == ["Theo lớp/học sinh"]
    assert report["detection"]["classes"] == ["10A1"]


def test_same_lessons_in_class_and_teacher_views_are_not_double_counted():
    project, classes, subjects, teachers, assignments = sample_context()
    wb = Workbook()
    ws_class = wb.active
    ws_class.title = "10A1"
    for row in [
        ["Tiết", "Thứ 2", "Thứ 3", "Thứ 4", "Thứ 5", "Thứ 6"],
        [1, "TOAN AN", "TOAN AN", "", "", ""],
    ]:
        ws_class.append(row)
    ws_teacher = wb.create_sheet("Nguyễn An")
    for row in [
        ["Tiết", "Thứ 2", "Thứ 3", "Thứ 4", "Thứ 5", "Thứ 6"],
        [1, "10A1 - Toán", "10A1 - Toán", "", "", ""],
    ]:
        ws_teacher.append(row)
    stream = io.BytesIO()
    wb.save(stream)

    report = analyze_schedule_file(
        filename="tkb-hai-goc-nhin.xlsx", content=stream.getvalue(), project=project, classes=classes,
        subjects=subjects, teachers=teachers, assignments=assignments,
    )
    assert report["summary"]["read_lessons"] == 4
    assert report["summary"]["recognized_lessons"] == 2
    assert report["summary"]["extra_periods"] == 0
    assert report["summary"]["collisions"] == 0
    assert set(report["detection"]["layouts"]) == {"Theo lớp/học sinh", "Theo giáo viên"}


def test_docx_teacher_day_grid_is_detected():
    project, classes, subjects, teachers, assignments = sample_context()
    xml = '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
    <w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main"><w:body>
      <w:p><w:r><w:t>Nguyễn An</w:t></w:r></w:p>
      <w:tbl>
        <w:tr>
          <w:tc><w:p><w:r><w:t>Tiết</w:t></w:r></w:p></w:tc>
          <w:tc><w:p><w:r><w:t>Thứ 2</w:t></w:r></w:p></w:tc>
          <w:tc><w:p><w:r><w:t>Thứ 3</w:t></w:r></w:p></w:tc>
        </w:tr>
        <w:tr>
          <w:tc><w:p><w:r><w:t>1</w:t></w:r></w:p></w:tc>
          <w:tc><w:p><w:r><w:t>10A1</w:t></w:r></w:p><w:p><w:r><w:t>Toán</w:t></w:r></w:p></w:tc>
          <w:tc><w:p><w:r><w:t>10A1</w:t></w:r></w:p><w:p><w:r><w:t>Toán</w:t></w:r></w:p></w:tc>
        </w:tr>
      </w:tbl>
    </w:body></w:document>'''
    stream = io.BytesIO()
    with zipfile.ZipFile(stream, "w") as zf:
        zf.writestr("word/document.xml", xml)
    report = analyze_schedule_file(
        filename="tkb-gv.docx", content=stream.getvalue(), project=project, classes=classes,
        subjects=subjects, teachers=teachers, assignments=assignments,
    )
    assert report["summary"]["recognized_lessons"] == 2
    assert report["summary"]["errors"] == 0
    assert report["detection"]["layouts"] == ["Theo giáo viên"]


def test_standalone_keeps_vietnamese_diacritics_distinct_for_teacher_identity():
    content = workbook_bytes([
        ["Thứ", "Tiết", "10A1", "10A2"],
        ["2", 1, "Toán Hà", "Toán Hạ"],
    ])
    report = analyze_standalone_schedule_file(filename="tkb-dau.xlsx", content=content)
    assert report["summary"]["recognized_lessons"] == 2
    assert report["summary"]["collisions"] == 0
    assert {teacher["name"] for teacher in report["data"]["teachers"]} == {"Hà", "Hạ"}
    cells = {cell["raw_text"]: cell for cell in report["viewer"]["cells"]}
    assert cells["Toán Hà"]["teacher_name"] == "Hà"
    assert cells["Toán Hạ"]["teacher_name"] == "Hạ"
    assert not cells["Toán Hà"]["conflicts"]
    assert not cells["Toán Hạ"]["conflicts"]


def test_standalone_parses_common_vietnamese_subject_teacher_cells_without_losing_marks():
    content = workbook_bytes([
        ["Thứ", "Tiết", "10A1", "10A2", "10A3", "10A4", "10A5"],
        ["2", 1, "Hóa Cô Lan", "Sinh Thầy Minh", "Lý Cô Hà", "Địa Cô Hương", "Sử Nguyễn Văn An"],
    ])
    report = analyze_standalone_schedule_file(filename="tkb-mon-gv.xlsx", content=content)
    cells = {cell["raw_text"]: cell for cell in report["viewer"]["cells"]}
    expected = {
        "Hóa Cô Lan": ("Hóa học", "Cô Lan"),
        "Sinh Thầy Minh": ("Sinh học", "Thầy Minh"),
        "Lý Cô Hà": ("Vật lý", "Cô Hà"),
        "Địa Cô Hương": ("Địa lý", "Cô Hương"),
        "Sử Nguyễn Văn An": ("Lịch sử", "Nguyễn Văn An"),
    }
    assert report["summary"]["recognized_lessons"] == len(expected)
    for raw_text, (subject, teacher) in expected.items():
        assert cells[raw_text]["subject_name"] == subject
        assert cells[raw_text]["teacher_name"] == teacher
        assert cells[raw_text]["raw_text"] == raw_text


def test_standalone_does_not_treat_teacher_middle_name_van_as_subject():
    content = workbook_bytes([
        ["Thứ", "Tiết", "10A1"],
        ["2", 1, "Sử Nguyễn Văn An"],
    ])
    report = analyze_standalone_schedule_file(filename="tkb-su.xlsx", content=content)
    cell = report["viewer"]["cells"][0]
    assert cell["subject_name"] == "Lịch sử"
    assert cell["teacher_name"] == "Nguyễn Văn An"


def test_standalone_wide_ignores_auxiliary_columns_instead_of_creating_fake_classes():
    content = workbook_bytes([
        ["Thứ", "Buổi", "Tiết", "10A1", "10A2", "GV nghỉ", "Ghi chú", "Phòng"],
        ["2", "Sáng", 1, "Toán Cô Hà", "Văn Thầy Minh", "Nguyễn Văn An", "Họp tổ", "P.201"],
    ])
    report = analyze_standalone_schedule_file(filename="tkb-cot-phu.xlsx", content=content)

    assert report["summary"]["read_lessons"] == 2
    assert report["summary"]["recognized_lessons"] == 2
    assert {item["name"] for item in report["data"]["classes"]} == {"10A1", "10A2"}
    raw_texts = {cell["raw_text"] for cell in report["viewer"]["cells"]}
    assert raw_texts == {"Toán Cô Hà", "Văn Thầy Minh"}
    assert "Nguyễn Văn An" not in raw_texts
    assert "Họp tổ" not in raw_texts
    assert "P.201" not in raw_texts


def test_standalone_wide_ignores_auxiliary_header_variants():
    content = workbook_bytes([
        ["THỨ", "TIẾT", "10A1", "Giáo viên vắng", "Chú thích", "ROOM", "Số thứ tự"],
        ["2", 1, "Hóa Cô Lan", "Cô Hạ", "Đổi tiết", "A101", 1],
    ])
    report = analyze_standalone_schedule_file(filename="tkb-cot-phu-bien-the.xlsx", content=content)

    assert report["summary"]["read_lessons"] == 1
    assert report["summary"]["recognized_lessons"] == 1
    assert [item["name"] for item in report["data"]["classes"]] == ["10A1"]
    assert report["viewer"]["cells"][0]["raw_text"] == "Hóa Cô Lan"
